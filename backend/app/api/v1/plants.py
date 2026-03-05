"""Plant management endpoints."""

import json
from typing import Annotated, Any
from uuid import UUID

from fastapi import APIRouter, BackgroundTasks, Depends, HTTPException, Query, Request

from app.api.v1.auth import get_client_ip
from app.core.exceptions import NotFoundError, ValidationError
from app.core.pool import fetch, fetchrow, fetchval, execute, fetch_insert
from app.core.security import (
    CurrentUser,
    require_admin,
    require_management_or_admin,
)
from app.models.plant import (
    PlantCreate,
    PlantUpdate,
    PlantSummary,
    PlantListResponse,
    PlantTransferRequest,
)
from app.core.events import broadcast
from app.monitoring.logging import get_logger
from app.services.audit_service import audit_service

router = APIRouter()
logger = get_logger(__name__)

# Whitelist of columns allowed for ORDER BY to prevent SQL injection
_ALLOWED_SORT_COLUMNS = {
    "fleet_number", "description", "fleet_type", "make", "model",
    "condition", "current_location", "state", "chassis_number",
    "year_of_manufacture", "purchase_year", "purchase_cost",
    "total_maintenance_cost", "parts_replaced_count", "last_maintenance_date",
    "created_at", "updated_at", "physical_verification", "remarks",
    "shared_po_count", "last_verified_date",
}


# ============================================================================
# Non-parametric routes MUST come before /{plant_id} routes
# ============================================================================


@router.get("/events")
async def list_plant_events(
    current_user: Annotated[CurrentUser, Depends(require_management_or_admin)],
    event_type: str | None = Query(None, pattern="^(movement|missing|new|returned|verification_failed)$"),
    plant_id: UUID | None = None,
    location_id: UUID | None = None,
    acknowledged: bool | None = None,
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=100),
) -> dict[str, Any]:
    """List plant events (movements, new plants, missing plants, etc.)

    Args:
        current_user: The authenticated user.
        event_type: Filter by event type.
        plant_id: Filter by specific plant.
        location_id: Filter by location involved.
        acknowledged: Filter by acknowledgement status.
        page: Page number.
        limit: Items per page.

    Returns:
        Paginated list of plant events.
    """
    conditions: list[str] = []
    params: list[Any] = []

    if event_type:
        params.append(event_type)
        conditions.append(f"pe.event_type = ${len(params)}")
    if plant_id:
        params.append(str(plant_id))
        conditions.append(f"pe.plant_id = ${len(params)}::uuid")
    if location_id:
        params.append(str(location_id))
        n = len(params)
        conditions.append(f"(pe.from_location_id = ${n}::uuid OR pe.to_location_id = ${n}::uuid)")
    if acknowledged is not None:
        params.append(acknowledged)
        conditions.append(f"pe.acknowledged = ${len(params)}")

    where = " AND ".join(conditions) if conditions else "TRUE"
    offset = (page - 1) * limit

    params.append(limit)
    params.append(offset)
    events = await fetch(
        f"""SELECT pe.*, pm.fleet_number, pm.description AS plant_description,
                   count(*) OVER() AS _total_count
            FROM plant_events pe
            LEFT JOIN plants_master pm ON pm.id = pe.plant_id
            WHERE {where}
            ORDER BY pe.created_at DESC
            LIMIT ${len(params) - 1} OFFSET ${len(params)}""",
        *params,
    )

    total = events[0].pop("_total_count", 0) if events else 0
    for row in events[1:]:
        row.pop("_total_count", None)

    return {
        "success": True,
        "data": events,
        "meta": {
            "page": page,
            "limit": limit,
            "total": total,
            "total_pages": (total + limit - 1) // limit if total > 0 else 0,
        },
    }


@router.patch("/events/{event_id}/acknowledge")
async def acknowledge_event(
    event_id: UUID,
    request: Request,
    background_tasks: BackgroundTasks,
    current_user: Annotated[CurrentUser, Depends(require_admin)],
    remarks: str | None = None,
) -> dict[str, Any]:
    """Acknowledge a plant event.

    Args:
        event_id: The event UUID.
        request: The HTTP request.
        background_tasks: Background task runner.
        current_user: The authenticated admin user.
        remarks: Optional remarks about acknowledgement.

    Returns:
        Updated event.
    """
    existing = await fetchrow(
        "SELECT * FROM plant_events WHERE id = $1::uuid",
        str(event_id),
    )

    if not existing:
        raise NotFoundError("Plant event", str(event_id))

    old_values = {"acknowledged": existing.get("acknowledged", False)}

    # Build dynamic SET clause
    set_parts = [
        "acknowledged = true",
        f"acknowledged_by = ${1}",
        "acknowledged_at = now()",
    ]
    params: list[Any] = [current_user.id]

    if remarks:
        params.append(remarks)
        set_parts.append(f"remarks = ${len(params)}")

    params.append(str(event_id))
    set_clause = ", ".join(set_parts)

    updated = await fetchrow(
        f"UPDATE plant_events SET {set_clause} WHERE id = ${len(params)}::uuid RETURNING *",
        *params,
    )

    logger.info(
        "Plant event acknowledged",
        event_id=str(event_id),
        user_id=current_user.id,
    )

    background_tasks.add_task(
        audit_service.log,
        user_id=current_user.id,
        user_email=current_user.email,
        action="update",
        table_name="plant_events",
        record_id=str(event_id),
        old_values=old_values,
        new_values={"acknowledged": True, "remarks": remarks},
        ip_address=get_client_ip(request),
        description=f"Acknowledged plant event {event_id}",
    )

    return {
        "success": True,
        "data": updated,
    }


@router.get("/search/{query}")
async def search_plants(
    query: str,
    current_user: Annotated[CurrentUser, Depends(require_management_or_admin)],
    condition: str | None = Query(None, pattern="^(working|standby|under_repair|breakdown|faulty|scrap|missing|off_hire|gpm_assessment|unverified)$"),
    location_id: UUID | None = None,
    fleet_type: str | None = Query(None, description="Filter by fleet type name"),
    limit: int = Query(20, ge=1, le=100),
) -> dict[str, Any]:
    """Full-text search for plants.

    Args:
        query: Search query.
        current_user: The authenticated user.
        condition: Filter by condition (working, standby, breakdown, etc.).
        location_id: Filter by location.
        fleet_type: Filter by fleet type name.
        limit: Maximum results.

    Returns:
        Search results ranked by relevance.
    """
    data = await fetch(
        "SELECT * FROM search_plants($1, $2, $3, $4, $5, $6)",
        query,
        condition,
        str(location_id) if location_id else None,
        fleet_type,
        limit,
        0,
    )

    return {
        "success": True,
        "data": data,
        "meta": {"query": query, "count": len(data)},
    }


@router.get("/usage/summary")
async def get_usage_summary(
    current_user: Annotated[CurrentUser, Depends(require_management_or_admin)],
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=100),
    year: int | None = None,
    month: int | None = Query(None, ge=1, le=12),
    week_number: int | None = Query(None, ge=1, le=53),
    location_id: UUID | None = None,
) -> dict[str, Any]:
    """Get plant usage summary across the fleet.

    Args:
        current_user: The authenticated user.
        page: Page number.
        limit: Items per page.
        year: Filter by year.
        month: Filter by month (1-12).
        week_number: Filter by week number (1-53).
        location_id: Filter by location.

    Returns:
        Paginated usage summary for plants.
    """
    offset = (page - 1) * limit

    data = await fetch(
        "SELECT * FROM get_plant_usage_summary($1, $2, $3, $4, $5, $6, $7)",
        None,  # p_plant_id
        year,
        month,
        week_number,
        str(location_id) if location_id else None,
        limit,
        offset,
    )

    # total_count is returned in each row via window function
    total = data[0]["total_count"] if data else 0
    # Strip the total_count field from response data
    clean_data = [{k: v for k, v in row.items() if k != "total_count"} for row in data]

    return {
        "success": True,
        "data": clean_data,
        "meta": {
            "page": page,
            "limit": limit,
            "total": total,
            "total_pages": (total + limit - 1) // limit if total > 0 else 0,
            "year": year,
            "month": month,
            "week_number": week_number,
        },
    }


@router.get("/usage/breakdowns")
async def get_breakdown_report(
    current_user: Annotated[CurrentUser, Depends(require_management_or_admin)],
    year: int | None = None,
    week_number: int | None = Query(None, ge=1, le=53),
    location_id: UUID | None = None,
) -> dict[str, Any]:
    """Get breakdown report - plants with breakdown hours.

    Args:
        current_user: The authenticated user.
        year: Filter by year.
        week_number: Filter by week number.
        location_id: Filter by location.

    Returns:
        List of plants with breakdowns.
    """
    data = await fetch(
        "SELECT * FROM get_breakdown_report($1, $2, $3)",
        year,
        week_number,
        str(location_id) if location_id else None,
    )

    return {
        "success": True,
        "data": data,
    }


@router.get("/utilization")
async def get_fleet_utilization(
    current_user: Annotated[CurrentUser, Depends(require_management_or_admin)],
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=100),
    location_id: UUID | None = None,
    fleet_type: str | None = Query(None, description="Filter by fleet type name"),
    condition: str | None = Query(None, pattern="^(working|standby|under_repair|breakdown|faulty|scrap|missing|off_hire|gpm_assessment|unverified)$", description="Filter by condition"),
    search: str | None = None,
) -> dict[str, Any]:
    """Get fleet utilization view with comprehensive stats.

    Args:
        current_user: The authenticated user.
        page: Page number.
        limit: Items per page.
        location_id: Filter by location.
        fleet_type: Filter by fleet type name.
        condition: Filter by condition (working, standby, breakdown, off_hire, etc.).
        search: Search in fleet_number or description.

    Returns:
        Paginated plant utilization data with hours, rates, and costs.
    """
    conditions: list[str] = []
    params: list[Any] = []

    if location_id:
        params.append(str(location_id))
        conditions.append(f"current_location_id = ${len(params)}::uuid")
    if fleet_type:
        params.append(f"%{fleet_type}%")
        conditions.append(f"fleet_type ILIKE ${len(params)}")
    if condition:
        params.append(condition)
        conditions.append(f"condition = ${len(params)}")
    if search:
        params.append(f"%{search}%")
        n = len(params)
        conditions.append(f"(fleet_number ILIKE ${n} OR description ILIKE ${n})")

    where = " AND ".join(conditions) if conditions else "TRUE"
    offset = (page - 1) * limit

    params.append(limit)
    params.append(offset)
    data = await fetch(
        f"""SELECT *, count(*) OVER() AS _total_count FROM v_plant_utilization
            WHERE {where}
            ORDER BY total_hours_worked DESC
            LIMIT ${len(params) - 1} OFFSET ${len(params)}""",
        *params,
    )

    total = data[0].pop("_total_count", 0) if data else 0
    for row in data[1:]:
        row.pop("_total_count", None)

    return {
        "success": True,
        "data": data,
        "meta": {
            "page": page,
            "limit": limit,
            "total": total,
            "total_pages": (total + limit - 1) // limit if total > 0 else 0,
        },
    }


@router.get("/export/excel")
async def export_plants_excel(
    current_user: Annotated[CurrentUser, Depends(require_management_or_admin)],
    exclude_not_seen: bool = Query(True, description="Exclude plants with 'not seen' in remarks"),
    location_id: UUID | None = Query(None, description="Filter by location"),
    state: str | None = Query(None, description="Filter by state (e.g., 'Kaduna', 'FCT', 'Ogun')"),
    fleet_type: str | None = Query(None, description="Filter by fleet type(s), comma-separated"),
    condition: str | None = Query(None, description="Filter by condition(s), comma-separated: working,standby,breakdown"),
    search: str | None = Query(None, description="Search in fleet_number or description"),
    verified_only: bool = Query(False, description="Only include physically verified plants"),
    columns: str | None = Query(None, description="Comma-separated columns to export. Default: all standard columns"),
) -> Any:
    """Export plants to Excel file with optional column and filter selection.

    Args:
        current_user: The authenticated user.
        exclude_not_seen: If true, exclude plants with 'not seen' in remarks.
        location_id: Filter by location.
        state: Filter by state.
        fleet_type: Filter by fleet type(s), comma-separated.
        condition: Filter by condition(s), comma-separated.
        search: Search in fleet_number or description.
        verified_only: Only include physically verified plants.
        columns: Comma-separated columns to include in export.

    Returns:
        Excel file download.
    """
    import io
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    select_fields = (
        "fleet_number, description, fleet_type, make, model, "
        "current_location, state, condition, physical_verification, "
        "chassis_number, year_of_manufacture, purchase_year, purchase_cost, "
        "total_maintenance_cost, parts_replaced_count, last_maintenance_date, remarks"
    )

    # Build dynamic WHERE for the query
    conds: list[str] = []
    params: list[Any] = []

    if exclude_not_seen:
        conds.append("(remarks NOT ILIKE '%not seen%' OR remarks IS NULL)")

    if location_id:
        params.append(str(location_id))
        conds.append(f"current_location_id = ${len(params)}::uuid")

    if state:
        params.append(f"%{state}%")
        conds.append(f"state ILIKE ${len(params)}")

    if fleet_type:
        fleet_type_list = [f.strip() for f in fleet_type.split(",")]
        ft_conds = []
        for ft in fleet_type_list:
            params.append(f"%{ft}%")
            ft_conds.append(f"fleet_type ILIKE ${len(params)}")
        conds.append(f"({' OR '.join(ft_conds)})")

    if condition:
        condition_list = [c.strip() for c in condition.split(",")]
        if len(condition_list) == 1:
            params.append(condition_list[0])
            conds.append(f"condition = ${len(params)}")
        else:
            placeholders = ", ".join(f"${len(params) + i + 1}" for i in range(len(condition_list)))
            params.extend(condition_list)
            conds.append(f"condition IN ({placeholders})")

    if search:
        params.append(f"%{search}%")
        n = len(params)
        conds.append(f"(fleet_number ILIKE ${n} OR description ILIKE ${n})")

    if verified_only:
        conds.append("physical_verification = true")

    where = " AND ".join(conds) if conds else "TRUE"

    plants = await fetch(
        f"""SELECT {select_fields} FROM v_plants_summary
            WHERE {where}
            ORDER BY fleet_type, fleet_number""",
        *params,
    )

    # All available export columns with their config
    all_export_columns = [
        {"key": "fleet_number", "header": "Fleet Number", "width": 15, "getter": lambda p: p.get("fleet_number")},
        {"key": "description", "header": "Description", "width": 30, "getter": lambda p: p.get("description")},
        {"key": "fleet_type", "header": "Fleet Type", "width": 25, "getter": lambda p: p.get("fleet_type")},
        {"key": "make", "header": "Make", "width": 15, "getter": lambda p: p.get("make")},
        {"key": "model", "header": "Model", "width": 15, "getter": lambda p: p.get("model")},
        {"key": "current_location", "header": "Location", "width": 20, "getter": lambda p: p.get("current_location")},
        {"key": "state", "header": "State", "width": 12, "getter": lambda p: p.get("state")},
        {"key": "condition", "header": "Condition", "width": 15, "getter": lambda p: (p.get("condition") or "").replace("_", " ").title()},
        {"key": "physical_verification", "header": "Physical Verification", "width": 18, "getter": lambda p: "Yes" if p.get("physical_verification") else "No"},
        {"key": "chassis_number", "header": "Chassis Number", "width": 20, "getter": lambda p: p.get("chassis_number")},
        {"key": "year_of_manufacture", "header": "Year of Manufacture", "width": 18, "getter": lambda p: p.get("year_of_manufacture")},
        {"key": "purchase_year", "header": "Purchase Year", "width": 14, "getter": lambda p: p.get("purchase_year")},
        {"key": "purchase_cost", "header": "Purchase Cost", "width": 15, "getter": lambda p: p.get("purchase_cost")},
        {"key": "total_maintenance_cost", "header": "Maintenance Cost", "width": 16, "getter": lambda p: p.get("total_maintenance_cost")},
        {"key": "parts_replaced_count", "header": "Parts Replaced", "width": 14, "getter": lambda p: p.get("parts_replaced_count")},
        {"key": "last_maintenance_date", "header": "Last Maintenance", "width": 16, "getter": lambda p: p.get("last_maintenance_date")},
        {"key": "remarks", "header": "Remarks", "width": 40, "getter": lambda p: p.get("remarks")},
    ]

    # Filter to requested columns (or use default set)
    if columns:
        requested = {c.strip() for c in columns.split(",")}
        export_cols = [c for c in all_export_columns if c["key"] in requested]
    else:
        default_keys = {"fleet_number", "description", "fleet_type", "make", "model", "current_location", "state", "physical_verification", "remarks"}
        export_cols = [c for c in all_export_columns if c["key"] in default_keys]

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Plants"

    # Style definitions
    title_font = Font(bold=True, size=14, color="101415")
    header_font = Font(bold=True, size=10, color="101415")
    header_fill = PatternFill(start_color="FFBF36", end_color="FFBF36", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_border = Border(
        left=Side(style="thin", color="808080"),
        right=Side(style="thin", color="808080"),
        top=Side(style="thin", color="808080"),
        bottom=Side(style="thin", color="808080"),
    )
    header_border = Border(
        left=Side(style="medium", color="606060"),
        right=Side(style="medium", color="606060"),
        top=Side(style="medium", color="606060"),
        bottom=Side(style="medium", color="606060"),
    )

    # Row 1: P.W. NIGERIA LTD. branding with logo
    ws.row_dimensions[1].height = 55
    logo_added = False
    try:
        from pathlib import Path
        from openpyxl.drawing.image import Image as XlImage
        logo_path = Path(__file__).resolve().parents[4] / "frontend" / "public" / "images" / "logo.png"
        if logo_path.exists():
            logo = XlImage(str(logo_path))
            logo.width = 50
            logo.height = 50
            ws.add_image(logo, "A1")
            logo_added = True
    except Exception as exc:
        logger.warning("Failed to add logo to Excel export", error=str(exc))

    text_col = 2 if logo_added else 1
    ws.merge_cells(start_row=1, start_column=text_col, end_row=1, end_column=len(export_cols))
    brand_cell = ws.cell(row=1, column=text_col, value="P.W. NIGERIA LTD. — Plant Register")
    brand_cell.font = title_font
    brand_cell.alignment = Alignment(vertical="center")

    # Row 2: Column headers
    for col_idx, col_def in enumerate(export_cols, 1):
        cell = ws.cell(row=2, column=col_idx, value=col_def["header"])
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = header_border

    # Write data
    for row_idx, plant in enumerate(plants, 3):
        for col_idx, col_def in enumerate(export_cols, 1):
            value = col_def["getter"](plant)
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = cell_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    # Set column widths
    for col_idx, col_def in enumerate(export_cols, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_def["width"]

    ws.freeze_panes = "A3"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    from datetime import datetime
    filename = f"plants_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    logger.info(
        "Plants exported to Excel",
        plants_count=len(plants),
        user_id=current_user.id,
    )

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/stats")
async def get_plant_stats(
    current_user: Annotated[CurrentUser, Depends(require_management_or_admin)],
    location_id: UUID | None = None,
) -> dict[str, Any]:
    """Get plant counts grouped by condition.

    Uses a single RPC that scans plants_master once with COUNT FILTER
    instead of 3 separate queries.

    Args:
        current_user: The authenticated user.
        location_id: Optional filter by location.

    Returns:
        Counts by condition (working, standby, breakdown, off_hire, etc.).
    """
    raw = await fetchval(
        "SELECT get_plant_filter_stats($1)",
        str(location_id) if location_id else None,
    )

    stats = (json.loads(raw) if isinstance(raw, str) else raw) if raw else {}

    return {
        "success": True,
        "data": {
            "total": stats.get("total", 0),
            "by_condition": stats.get("by_condition", {}),
            "unknown_location": stats.get("unknown_location", 0),
            "pending_transfers": stats.get("pending_transfers", 0),
        },
    }


@router.get("/filtered-stats")
async def get_filtered_plant_stats(
    current_user: Annotated[CurrentUser, Depends(require_management_or_admin)],
    condition: str | None = Query(None, description="Filter by condition(s), comma-separated"),
    location_id: UUID | None = None,
    state: str | None = Query(None, description="Filter by state"),
    fleet_type: str | None = Query(None, description="Filter by fleet type(s), comma-separated"),
    search: str | None = None,
    verified_only: bool = False,
    unknown_location: bool = Query(False),
    pending_transfer: bool = Query(False),
) -> dict[str, Any]:
    """Get aggregated plant stats matching current filters.

    Returns condition breakdown, location breakdown, and fleet type breakdown
    for the filtered plant set. Uses a database RPC for fast GROUP BY aggregation.
    """
    # Parse multi-value filters
    condition_list = [c.strip() for c in condition.split(",")] if condition else None
    fleet_type_list = [f.strip() for f in fleet_type.split(",")] if fleet_type else None

    raw = await fetchval(
        "SELECT get_filtered_plant_stats($1, $2, $3, $4, $5, $6, $7, $8)",
        condition_list,
        str(location_id) if location_id else None,
        fleet_type_list,
        state,
        search,
        verified_only,
        unknown_location,
        pending_transfer,
    )

    stats = (json.loads(raw) if isinstance(raw, str) else raw) if raw else {}
    stats = stats or {"total": 0, "by_condition": {}, "by_location": {}, "by_fleet_type": {}, "by_state_fleet_type": {}}

    return {
        "success": True,
        "data": stats,
    }


# ============================================================================
# Parametric routes with /{plant_id}
# ============================================================================


@router.get("")
async def list_plants(
    current_user: Annotated[CurrentUser, Depends(require_management_or_admin)],
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=500),
    condition: str | None = Query(None, description="Filter by condition(s). Comma-separated: working,standby,breakdown,off_hire"),
    location_id: UUID | None = None,
    state: str | None = Query(None, description="Filter by state (e.g., 'Kaduna', 'FCT', 'Ogun')"),
    fleet_type: str | None = Query(None, description="Filter by fleet type(s). Comma-separated: TRUCKS,EXCAVATOR"),
    search: str | None = None,
    verified_only: bool = False,
    unknown_location: bool = Query(False, description="Filter for plants with unknown/NULL location"),
    pending_transfer: bool = Query(False, description="Filter for plants with pending transfers"),
    columns: str | None = Query(
        None,
        description="Comma-separated list of columns to return. Default: all. Example: fleet_number,condition,current_location,fleet_type",
    ),
    sort_by: str = Query("fleet_number", description="Column to sort by"),
    sort_order: str = Query("asc", pattern="^(asc|desc)$", description="Sort order: asc or desc"),
) -> dict[str, Any]:
    """List plants with filtering, pagination, and column selection.

    **Multi-value filters:** Use comma-separated values to filter by multiple options.
    - `condition=working,standby` - Plants that are working OR standby
    - `fleet_type=TRUCKS,EXCAVATOR` - Trucks OR Excavators

    **Column selection:** Use `columns` to select which fields to return.
    - `columns=fleet_number,condition,current_location` - Only these 3 fields

    **Condition values:**
    working, standby, under_repair, breakdown, scrap, missing, off_hire, gpm_assessment, unverified
    """
    # Parse multi-value filters
    condition_list = [c.strip() for c in condition.split(",")] if condition else None
    fleet_type_list = [f.strip() for f in fleet_type.split(",")] if fleet_type else None

    # Build select clause based on columns parameter
    if columns:
        column_list = [c.strip() for c in columns.split(",")]
        if "id" not in column_list:
            column_list.insert(0, "id")
        select_clause = ", ".join(column_list)
    else:
        select_clause = "*"

    # Build WHERE clause
    conds: list[str] = []
    params: list[Any] = []

    # If filtering by state, get location IDs for that state
    if state:
        params.append(f"%{state}%")
        state_locs = await fetch(
            f"""SELECT l.id FROM locations l
                JOIN states s ON s.id = l.state_id
                WHERE s.name ILIKE ${len(params)}""",
            *params,
        )
        state_location_ids = [loc["id"] for loc in state_locs]
        if not state_location_ids:
            return {
                "success": True,
                "data": [],
                "meta": {"page": page, "limit": limit, "total": 0, "total_pages": 0, "has_more": False},
            }
        # Reset params — the state lookup was a separate query
        params = []
        placeholders = ", ".join(f"${i+1}::uuid" for i in range(len(state_location_ids)))
        params.extend(state_location_ids)
        conds.append(f"current_location_id IN ({placeholders})")

    # Condition filter
    if pending_transfer:
        conds.append("pending_transfer_to_id IS NOT NULL")
    elif condition_list:
        if len(condition_list) == 1:
            params.append(condition_list[0])
            conds.append(f"condition = ${len(params)}")
        else:
            placeholders = ", ".join(f"${len(params) + i + 1}" for i in range(len(condition_list)))
            params.extend(condition_list)
            conds.append(f"condition IN ({placeholders})")

    # Fleet type filter (ILIKE for partial matching)
    if fleet_type_list:
        ft_conds = []
        for ft in fleet_type_list:
            params.append(f"%{ft}%")
            ft_conds.append(f"fleet_type ILIKE ${len(params)}")
        conds.append(f"({' OR '.join(ft_conds)})")

    # Location filters
    if unknown_location:
        conds.append("current_location_id IS NULL")
    elif location_id and not state:
        params.append(str(location_id))
        conds.append(f"current_location_id = ${len(params)}::uuid")

    if search:
        params.append(f"%{search}%")
        n = len(params)
        conds.append(f"(fleet_number ILIKE ${n} OR description ILIKE ${n})")

    if verified_only:
        conds.append("physical_verification = true")

    where = " AND ".join(conds) if conds else "TRUE"

    # Validate sort column
    safe_sort = sort_by if sort_by in _ALLOWED_SORT_COLUMNS else "fleet_number"
    safe_order = "DESC" if sort_order == "desc" else "ASC"

    # Single query: data + total count in one round-trip (saves ~500ms network latency)
    offset = (page - 1) * limit
    params.append(limit)
    params.append(offset)
    data = await fetch(
        f"""SELECT {select_clause}, count(*) OVER() AS _total_count
            FROM v_plants_summary
            WHERE {where}
            ORDER BY {safe_sort} {safe_order}
            LIMIT ${len(params) - 1} OFFSET ${len(params)}""",
        *params,
    )

    # Extract total from first row, strip the helper column
    total = data[0].pop("_total_count", 0) if data else 0
    for row in data[1:]:
        row.pop("_total_count", None)

    return {
        "success": True,
        "data": data,
        "meta": {
            "page": page,
            "limit": limit,
            "total": total,
            "total_pages": (total + limit - 1) // limit if total > 0 else 0,
            "has_more": page * limit < total,
            "columns": columns.split(",") if columns else "all",
            "filters": {
                "condition": condition_list,
                "fleet_type": fleet_type_list,
                "location_id": str(location_id) if location_id else None,
                "state": state,
                "verified_only": verified_only,
                "unknown_location": unknown_location,
                "pending_transfer": pending_transfer,
            },
        },
    }


@router.get("/by-fleet/{fleet_number}")
async def get_plant_by_fleet(
    fleet_number: str,
    current_user: Annotated[CurrentUser, Depends(require_management_or_admin)],
) -> dict[str, Any]:
    """Get a single plant by fleet number.

    Args:
        fleet_number: The fleet number (e.g., P453, T468, E25).
        current_user: The authenticated user.

    Returns:
        Plant details with related data.
    """
    normalized = " ".join(fleet_number.upper().split())

    result = await fetchrow(
        "SELECT * FROM v_plants_summary WHERE fleet_number = $1",
        normalized,
    )

    if not result:
        raise NotFoundError("Plant with fleet number", fleet_number)

    return {
        "success": True,
        "data": result,
    }


@router.get("/{plant_id}")
async def get_plant(
    plant_id: UUID,
    current_user: Annotated[CurrentUser, Depends(require_management_or_admin)],
) -> dict[str, Any]:
    """Get a single plant by ID.

    Args:
        plant_id: The plant UUID.
        current_user: The authenticated user.

    Returns:
        Plant details with related data.
    """
    result = await fetchrow(
        "SELECT * FROM v_plants_summary WHERE id = $1::uuid",
        str(plant_id),
    )

    if not result:
        raise NotFoundError("Plant", str(plant_id))

    return {
        "success": True,
        "data": result,
    }


@router.post("", status_code=201)
async def create_plant(
    plant: PlantCreate,
    request: Request,
    background_tasks: BackgroundTasks,
    current_user: Annotated[CurrentUser, Depends(require_admin)],
) -> dict[str, Any]:
    """Create a new plant.

    Args:
        plant: Plant data.
        request: The HTTP request.
        background_tasks: Background task runner.
        current_user: The authenticated admin user.

    Returns:
        Created plant with ID.
    """
    plant_data = plant.model_dump(exclude_none=True, mode="json")

    # Build INSERT dynamically from provided fields
    cols = list(plant_data.keys())
    vals = list(plant_data.values())
    placeholders = ", ".join(f"${i+1}" for i in range(len(cols)))
    col_names = ", ".join(cols)

    try:
        created = await fetchrow(
            f"INSERT INTO plants_master ({col_names}) VALUES ({placeholders}) RETURNING *",
            *vals,
        )
    except Exception as e:
        error_msg = str(e).lower()
        if "duplicate" in error_msg or "unique" in error_msg or "already exists" in error_msg:
            raise ValidationError(
                "Plant with this fleet number already exists",
                details=[{"field": "fleet_number", "message": "Already exists", "code": "DUPLICATE"}],
            )
        if "violates check constraint" in error_msg:
            raise ValidationError(
                "Invalid data: a field value is not allowed. Please check all fields and try again.",
                details=[{"message": str(e), "code": "CONSTRAINT_VIOLATION"}],
            )
        raise

    # Record initial location history if a location was provided
    if plant_data.get("current_location_id"):
        await execute(
            """INSERT INTO plant_location_history (plant_id, location_id, start_date, transfer_reason, created_by)
               VALUES ($1::uuid, $2::uuid, now(), $3, $4)""",
            created["id"],
            plant_data["current_location_id"],
            "Initial assignment",
            current_user.id,
        )

    logger.info(
        "Plant created",
        plant_id=created["id"],
        fleet_number=plant.fleet_number,
        user_id=current_user.id,
    )

    background_tasks.add_task(
        audit_service.log,
        user_id=current_user.id,
        user_email=current_user.email,
        action="create",
        table_name="plants_master",
        record_id=created["id"],
        new_values=plant.model_dump(exclude_none=True, mode="json"),
        ip_address=get_client_ip(request),
        description=f"Created plant {plant.fleet_number}",
    )

    broadcast("plants", "create")
    return {
        "success": True,
        "data": created,
    }


@router.patch("/{plant_id}")
async def update_plant(
    plant_id: UUID,
    request: Request,
    background_tasks: BackgroundTasks,
    current_user: Annotated[CurrentUser, Depends(require_admin)],
    description: str | None = Query(None, description="Equipment description"),
    fleet_type: str | None = Query(None, description="Fleet type (e.g., TRUCKS, EXCAVATOR)"),
    make: str | None = Query(None, description="Manufacturer (e.g., TOYOTA, CAT)"),
    model: str | None = Query(None, description="Model name/number"),
    chassis_number: str | None = Query(None, description="Chassis/VIN number"),
    year_of_manufacture: int | None = Query(None, ge=1900, le=2100, description="Year manufactured"),
    purchase_year: int | None = Query(None, ge=1900, le=2100, description="Year purchased"),
    purchase_cost: float | None = Query(None, ge=0, description="Purchase cost"),
    serial_m: str | None = Query(None, description="Mechanical serial number"),
    serial_e: str | None = Query(None, description="Electrical serial number"),
    remarks: str | None = Query(None, description="Additional remarks"),
    current_location_id: UUID | None = Query(None, description="Current location UUID"),
    condition: str | None = Query(None, pattern="^(working|standby|under_repair|breakdown|faulty|scrap|missing|off_hire|gpm_assessment|unverified)$", description="Plant condition"),
    physical_verification: bool | None = Query(None, description="Has been physically verified"),
) -> dict[str, Any]:
    """Update an existing plant - only provide the fields you want to change.

    **Example:** To update just purchase_cost and purchase_year:
    ```
    PATCH /plants/{id}?purchase_cost=5000000&purchase_year=2023
    ```
    """
    # Build update data from provided parameters only
    update_data: dict[str, Any] = {}
    if description is not None:
        update_data["description"] = description
    if fleet_type is not None:
        update_data["fleet_type"] = fleet_type
    if make is not None:
        update_data["make"] = make
    if model is not None:
        update_data["model"] = model
    if chassis_number is not None:
        update_data["chassis_number"] = chassis_number
    if year_of_manufacture is not None:
        update_data["year_of_manufacture"] = year_of_manufacture
    if purchase_year is not None:
        update_data["purchase_year"] = purchase_year
    if purchase_cost is not None:
        update_data["purchase_cost"] = purchase_cost
    if serial_m is not None:
        update_data["serial_m"] = serial_m
    if serial_e is not None:
        update_data["serial_e"] = serial_e
    if remarks is not None:
        update_data["remarks"] = remarks
    if current_location_id is not None:
        update_data["current_location_id"] = str(current_location_id)
    if condition is not None:
        update_data["condition"] = condition
    if physical_verification is not None:
        update_data["physical_verification"] = physical_verification

    if not update_data:
        raise ValidationError("No fields to update. Provide at least one field.")

    # Fetch current values for audit diff
    existing = await fetchrow(
        "SELECT * FROM plants_master WHERE id = $1::uuid",
        str(plant_id),
    )

    if not existing:
        raise NotFoundError("Plant", str(plant_id))

    old_values = {k: existing.get(k) for k in update_data if k in existing}

    # Build SET clause
    update_data["updated_at"] = "now()"
    set_parts: list[str] = []
    params: list[Any] = []
    for key, val in update_data.items():
        if key == "updated_at":
            set_parts.append("updated_at = now()")
        elif key == "current_location_id":
            params.append(val)
            set_parts.append(f"{key} = ${len(params)}::uuid")
        else:
            params.append(val)
            set_parts.append(f"{key} = ${len(params)}")

    params.append(str(plant_id))
    set_clause = ", ".join(set_parts)

    await execute(
        f"UPDATE plants_master SET {set_clause} WHERE id = ${len(params)}::uuid",
        *params,
    )

    logger.info(
        "Plant updated",
        plant_id=str(plant_id),
        updated_fields=list(update_data.keys()),
        user_id=current_user.id,
    )

    fleet_number = existing.get("fleet_number", str(plant_id))
    background_tasks.add_task(
        audit_service.log,
        user_id=current_user.id,
        user_email=current_user.email,
        action="update",
        table_name="plants_master",
        record_id=str(plant_id),
        old_values=old_values,
        new_values={k: v for k, v in update_data.items() if k != "updated_at"},
        ip_address=get_client_ip(request),
        description=f"Updated plant {fleet_number}: {', '.join(update_data.keys())}",
    )

    # Return full plant data from view
    updated = await fetchrow(
        "SELECT * FROM v_plants_summary WHERE id = $1::uuid",
        str(plant_id),
    )

    broadcast("plants", "update")

    return {
        "success": True,
        "data": updated,
        "meta": {
            "updated_fields": [k for k in update_data.keys() if k != "updated_at"],
        },
    }


@router.post("/{plant_id}/transfer")
async def transfer_plant(
    plant_id: UUID,
    transfer: PlantTransferRequest,
    request: Request,
    background_tasks: BackgroundTasks,
    current_user: Annotated[CurrentUser, Depends(require_admin)],
) -> dict[str, Any]:
    """Transfer a plant to a new location.

    Args:
        plant_id: The plant UUID.
        transfer: Transfer details.
        request: The HTTP request.
        background_tasks: Background task runner.
        current_user: The authenticated admin user.

    Returns:
        Transfer result.
    """
    raw = await fetchval(
        "SELECT transfer_plant($1, $2, $3, $4)",
        str(plant_id),
        str(transfer.new_location_id),
        transfer.transfer_reason,
        current_user.id,
    )

    result = (json.loads(raw) if isinstance(raw, str) else raw) if raw else None
    if not result or not result.get("success"):
        raise ValidationError(result.get("error", "Transfer failed") if result else "Transfer failed")

    from_loc = result.get("from_location", "Unknown")
    to_loc = result.get("to_location", "Unknown")

    logger.info(
        "Plant transferred",
        plant_id=str(plant_id),
        from_location=from_loc,
        to_location=to_loc,
        user_id=current_user.id,
    )

    background_tasks.add_task(
        audit_service.log,
        user_id=current_user.id,
        user_email=current_user.email,
        action="transfer",
        table_name="plants_master",
        record_id=str(plant_id),
        old_values={"location": from_loc},
        new_values={"location": to_loc, "transfer_reason": transfer.transfer_reason},
        ip_address=get_client_ip(request),
        description=f"Transferred plant from {from_loc} to {to_loc}",
    )

    broadcast("plants", "transfer")

    return {
        "success": True,
        "data": result,
    }


@router.delete("/{plant_id}")
async def delete_plant(
    plant_id: UUID,
    request: Request,
    background_tasks: BackgroundTasks,
    current_user: Annotated[CurrentUser, Depends(require_admin)],
) -> dict[str, Any]:
    """Delete a plant record.

    Args:
        plant_id: The plant UUID.
        request: The HTTP request.
        background_tasks: Background task runner.
        current_user: The authenticated admin user.

    Returns:
        Success message.
    """
    existing = await fetchrow(
        "SELECT * FROM plants_master WHERE id = $1::uuid",
        str(plant_id),
    )

    if not existing:
        raise NotFoundError("Plant", str(plant_id))

    fleet_number = existing.get("fleet_number", str(plant_id))

    await execute(
        "DELETE FROM plants_master WHERE id = $1::uuid",
        str(plant_id),
    )

    logger.info(
        "Plant deleted",
        plant_id=str(plant_id),
        fleet_number=fleet_number,
        user_id=current_user.id,
    )

    background_tasks.add_task(
        audit_service.log,
        user_id=current_user.id,
        user_email=current_user.email,
        action="delete",
        table_name="plants_master",
        record_id=str(plant_id),
        old_values=existing,
        ip_address=get_client_ip(request),
        description=f"Deleted plant {fleet_number}",
    )

    broadcast("plants", "delete")

    return {
        "success": True,
        "message": f"Plant {fleet_number} deleted successfully",
    }


@router.get("/{plant_id}/maintenance-history")
async def get_plant_maintenance_history(
    plant_id: UUID,
    current_user: Annotated[CurrentUser, Depends(require_management_or_admin)],
    limit: int = Query(50, ge=1, le=200),
) -> dict[str, Any]:
    """Get maintenance history for a plant.

    Args:
        plant_id: The plant UUID.
        current_user: The authenticated user.
        limit: Maximum records to return.

    Returns:
        List of maintenance/spare parts records.
    """
    data = await fetch(
        "SELECT * FROM get_plant_maintenance_history($1, $2)",
        str(plant_id),
        limit,
    )

    return {
        "success": True,
        "data": data,
    }


@router.get("/{plant_id}/location-history")
async def get_plant_location_history(
    plant_id: UUID,
    current_user: Annotated[CurrentUser, Depends(require_management_or_admin)],
) -> dict[str, Any]:
    """Get location history for a plant.

    Args:
        plant_id: The plant UUID.
        current_user: The authenticated user.

    Returns:
        List of location records showing where the plant has been.
    """
    data = await fetch(
        "SELECT * FROM get_plant_location_history($1)",
        str(plant_id),
    )

    return {
        "success": True,
        "data": data,
    }


@router.get("/{plant_id}/weekly-records")
async def get_plant_weekly_records(
    plant_id: UUID,
    current_user: Annotated[CurrentUser, Depends(require_management_or_admin)],
    year: int | None = Query(None, description="Filter by year"),
    page: int = Query(1, ge=1, description="Page number"),
    limit: int = Query(52, ge=1, le=200, description="Items per page (default 52 = 1 year)"),
) -> dict[str, Any]:
    """Get weekly tracking records for a plant.

    Shows where the plant was reported each week.

    Args:
        plant_id: The plant UUID.
        current_user: The authenticated user.
        year: Filter by year.
        page: Page number.
        limit: Items per page (default 52 = 1 year).

    Returns:
        Paginated list of weekly records with location info.
    """
    conditions = ["pwr.plant_id = $1::uuid"]
    params: list[Any] = [str(plant_id)]

    if year:
        params.append(year)
        conditions.append(f"pwr.year = ${len(params)}")

    where = " AND ".join(conditions)
    offset = (page - 1) * limit

    params.append(limit)
    params.append(offset)
    records = await fetch(
        f"""SELECT pwr.*, l.name AS location_name, count(*) OVER() AS _total_count
            FROM plant_weekly_records pwr
            LEFT JOIN locations l ON l.id = pwr.location_id
            WHERE {where}
            ORDER BY pwr.year DESC, pwr.week_number DESC
            LIMIT ${len(params) - 1} OFFSET ${len(params)}""",
        *params,
    )

    total = records[0].pop("_total_count", 0) if records else 0
    for row in records[1:]:
        row.pop("_total_count", None)

    return {
        "success": True,
        "data": records,
        "meta": {
            "page": page,
            "limit": limit,
            "total": total,
            "total_pages": (total + limit - 1) // limit if total > 0 else 0,
            "has_more": page * limit < total,
            "year": year,
        },
    }


@router.get("/{plant_id}/events")
async def get_plant_events(
    plant_id: UUID,
    current_user: Annotated[CurrentUser, Depends(require_management_or_admin)],
    limit: int = Query(50, ge=1, le=200),
) -> dict[str, Any]:
    """Get events for a specific plant.

    Args:
        plant_id: The plant UUID.
        current_user: The authenticated user.
        limit: Maximum records.

    Returns:
        List of events for this plant.
    """
    events = await fetch(
        """SELECT pe.*,
                  fl.name AS from_location_name,
                  tl.name AS to_location_name
           FROM plant_events pe
           LEFT JOIN locations fl ON fl.id = pe.from_location_id
           LEFT JOIN locations tl ON tl.id = pe.to_location_id
           WHERE pe.plant_id = $1::uuid
           ORDER BY pe.created_at DESC
           LIMIT $2""",
        str(plant_id),
        limit,
    )

    return {
        "success": True,
        "data": events,
    }


@router.get("/{plant_id}/usage")
async def get_plant_usage(
    plant_id: UUID,
    current_user: Annotated[CurrentUser, Depends(require_management_or_admin)],
    year: int | None = None,
    month: int | None = Query(None, ge=1, le=12),
) -> dict[str, Any]:
    """Get usage summary for a specific plant.

    Args:
        plant_id: The plant UUID.
        current_user: The authenticated user.
        year: Filter by year.
        month: Filter by month.

    Returns:
        Usage summary for the plant.
    """
    data = await fetch(
        "SELECT * FROM get_plant_usage_summary($1, $2, $3, $4, $5)",
        str(plant_id),
        year,
        month,
        None,  # p_week_number
        None,  # p_location_id
    )

    if not data:
        raise NotFoundError("Plant usage data", str(plant_id))

    return {
        "success": True,
        "data": data[0] if data else None,
    }
