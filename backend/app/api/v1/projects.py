"""Project management endpoints.

CRUD operations for the projects registry + Award Letters Excel import.
All mutations are admin-only. All reads are available to any authenticated user.
"""

from datetime import date
from typing import Annotated, Any
from uuid import UUID

from fastapi import (
    APIRouter,
    BackgroundTasks,
    Depends,
    File,
    Form,
    HTTPException,
    Query,
    Request,
    UploadFile,
)

from app.api.v1.auth import get_client_ip
from app.core.exceptions import NotFoundError, ValidationError
from app.core.pool import fetch, fetchrow, fetchval, execute, get_pool
from app.core.security import (
    CurrentUser,
    require_admin,
    require_projects_access,
)
from app.models.project import ProjectCreate, ProjectUpdate
from app.core.events import broadcast
from app.monitoring.logging import get_logger
from app.services.audit_service import audit_service

router = APIRouter()
logger = get_logger(__name__)

_ALLOWED_SORT_COLUMNS = {
    "project_name",
    "client",
    "state_name",
    "status",
    "original_contract_sum",
    "current_contract_sum",
    "award_date",
    "created_at",
    "updated_at",
}

# Fields that require ::uuid cast in parameterized queries
_UUID_FIELDS = {"state_id", "client_id", "location_id", "created_by", "updated_by", "import_batch_id"}


# ============================================================================
# Non-parametric routes (must come before /{project_id})
# ============================================================================


@router.post("/upload-weekly-report")
async def upload_weekly_report(
    request: Request,
    background_tasks: BackgroundTasks,
    current_user: Annotated[CurrentUser, Depends(require_admin)],
    file: UploadFile = File(...),
    project_id: UUID = Form(...),
    year: int = Form(...),
    week_number: int = Form(...),
) -> dict[str, Any]:
    """Upload a 16-sheet project weekly report (admin). Mirrors the plant
    upload PATTERN (pick project + week, drop file, poll submission) but
    shares no tables or code with it."""
    import hashlib

    if not file.filename:
        raise ValidationError("File name is required")
    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
    if ext != "xlsx":
        raise ValidationError("Only .xlsx weekly reports are accepted")
    if not (1 <= week_number <= 53):
        raise ValidationError("week_number must be 1–53")
    if not (2020 <= year <= 2100):
        raise ValidationError("year out of range")

    file_content = await file.read()
    if len(file_content) > 25 * 1024 * 1024:
        raise ValidationError("File too large (max 25MB)")

    project = await fetchrow(
        "SELECT id, short_name FROM projects WHERE id = $1::uuid", str(project_id)
    )
    if project is None:
        raise NotFoundError("Project not found")

    # ── store the original file ─────────────────────────────────────────
    from app.core.database import get_supabase_admin_client  # Storage only

    storage_path = (
        f"weekly-reports/projects/{project_id}/"
        f"{year}-W{week_number:02d}/{file.filename}"
    )
    client = get_supabase_admin_client()
    try:
        client.storage.from_("reports").upload(
            storage_path, file_content,
            {"content-type":
             "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
        )
    except Exception as e:
        if "already exists" in str(e).lower() or "duplicate" in str(e).lower():
            client.storage.from_("reports").update(storage_path, file_content)
        else:
            logger.error("Storage upload failed",
                         path=storage_path, error=f"{type(e).__name__}: {e}")
            raise HTTPException(
                status_code=502,
                detail=f"Could not store the file in Supabase Storage: {e}",
            ) from e

    submission_id = await fetchval(
        """INSERT INTO project_report_submissions
           (project_id, year, week_number, file_name, file_hash, file_path,
            file_size, source, status, uploaded_by)
           VALUES ($1::uuid, $2, $3, $4, $5, $6, $7, 'excel', 'queued', $8::uuid)
           RETURNING id""",
        str(project_id), year, week_number, file.filename,
        hashlib.sha256(file_content).hexdigest(), storage_path,
        len(file_content), current_user.id,
    )

    from app.workers.project_report_worker import process_project_weekly_report

    background_tasks.add_task(process_project_weekly_report, str(submission_id))

    background_tasks.add_task(
        audit_service.log,
        user_id=current_user.id,
        user_email=current_user.email,
        action="create",
        table_name="project_report_submissions",
        record_id=str(submission_id),
        new_values={"project": project["short_name"], "year": year,
                    "week": week_number, "file": file.filename},
        ip_address=get_client_ip(request),
        description=f"Uploaded weekly report {year}-W{week_number} for {project['short_name']}",
    )
    return {
        "success": True,
        "data": {"submission_id": str(submission_id), "status": "queued"},
    }


@router.post("/preview-weekly-report")
async def preview_weekly_report(
    current_user: Annotated[CurrentUser, Depends(require_admin)],
    file: UploadFile = File(...),
    project_id: UUID | None = Form(None),
) -> dict[str, Any]:
    """Parse a weekly report IN MEMORY — nothing is stored. Returns every
    sheet for review: parsed sheets as structured tables with warnings and
    cross-checks, stored-only sheets as a raw grid. The admin inspects,
    then confirms via the normal upload endpoint."""
    import io

    import openpyxl

    from app.services.weekly_report_sheets import (
        STORED_ONLY_SHEETS,
        parse_workbook,
    )

    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise ValidationError("Only .xlsx weekly reports are accepted")
    content = await file.read()
    if len(content) > 25 * 1024 * 1024:
        raise ValidationError("File too large (max 25MB)")

    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        raise ValidationError(f"Not a readable Excel workbook: {e}") from e

    import time as _time
    _t0 = _time.monotonic()
    try:
        parsed = parse_workbook(wb)
    except Exception as e:
        logger.error("Preview parse failed", error=f"{type(e).__name__}: {e}")
        raise HTTPException(
            status_code=422, detail=f"Workbook could not be parsed: {e}"
        ) from e
    parse_ms = int((_time.monotonic() - _t0) * 1000)

    # identity guard against the selected project (advisory)
    identity_warning = None
    if project_id:
        sel = await fetchrow(
            "SELECT short_name FROM projects WHERE id = $1::uuid", str(project_id))
        if sel is None:
            raise NotFoundError("Project not found")
        wb_short = ((parsed.get("identity") or {}).get("short_name") or "").strip().upper()
        sel_short = (sel["short_name"] or "").strip().upper()
        if wb_short and sel_short and wb_short != sel_short:
            identity_warning = (
                f"Workbook says {wb_short!r} but you selected {sel_short!r} — "
                "verify you picked the right project"
            )

    sheets_out: dict[str, Any] = {}
    for name, s in parsed["sheets"].items():
        entry: dict[str, Any] = {
            "kind": "parsed",
            "status": s.get("status"),
            "warnings": s.get("warnings", []),
        }
        rows = s.get("rows")
        if isinstance(rows, list):
            entry["rows"] = rows          # full sheet — they are small
            entry["total_rows"] = len(rows)
        # sheet-specific extras the preview renders as summary panels
        for extra in ("bills", "tail", "summary_table", "cross_checks",
                      "footer", "stock", "sheet_totals", "sheet_total",
                      "snapshot", "stock_maintained", "totals"):
            if s.get(extra) is not None:
                entry[extra] = s[extra]
        if name == "Lists":
            entry["rows"] = []
            entry["total_rows"] = len(s.get("reference", []) or [])
            entry["calendar_weeks"] = len(s.get("week_endings", {}) or {})
        sheets_out[name] = entry

    # stored-only sheets: raw grid so the admin can eyeball them
    for name in STORED_ONLY_SHEETS:
        target = next((n for n in wb.sheetnames
                       if n.strip().lower() == name.strip().lower()), None)
        if target is None:
            sheets_out[name] = {"kind": "stored_only", "status": "missing",
                                "warnings": [], "grid": []}
            continue
        ws = wb[target]
        grid = []
        for row in ws.iter_rows(min_row=1, max_row=30, max_col=12,
                                values_only=True):
            cells = ["" if v is None else str(v)[:60] for v in row]
            if any(c for c in cells):
                grid.append(cells)
        sheets_out[name] = {"kind": "stored_only", "status": "stored",
                            "warnings": [], "grid": grid}

    # ── detect-and-confirm: the workbook is the truth, the form is a
    # claim. Match its short name against the register and read its own
    # week declaration (consensus of 11+ sheet headers).
    declared = parsed.get("declared") or {}
    wb_short = ((parsed.get("identity") or {}).get("short_name") or "").strip()
    matched_project = None
    if wb_short:
        matched_project = await fetchrow(
            """SELECT id, short_name, project_name FROM projects
               WHERE upper(trim(short_name)) = upper($1) AND is_legacy = false
               LIMIT 1""",
            wb_short,
        )
    already_ingested = False
    if matched_project and declared.get("week_number") and declared.get("year"):
        already_ingested = bool(await fetchval(
            """SELECT 1 FROM project_weekly_reports
               WHERE project_id = $1::uuid AND year = $2 AND week_number = $3""",
            str(matched_project["id"]), declared["year"], declared["week_number"],
        ))

    return {"success": True, "data": {
        "identity": parsed.get("identity"),
        "identity_warning": identity_warning,
        "drift": parsed["drift"],
        "sheets": sheets_out,
        "parse_ms": parse_ms,
        "file_name": file.filename,
        "file_size": len(content),
        "declared": declared,
        "matched_project": matched_project,
        "already_ingested": already_ingested,
    }}


@router.get("/submissions")
async def list_project_submissions(
    current_user: Annotated[CurrentUser, Depends(require_projects_access)],
    status: str | None = Query(None, pattern="^(queued|parsing|success|partial|failed|deleted)$"),
    project_id: UUID | None = None,
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=100),
) -> dict[str, Any]:
    conds, params = ["TRUE"], []
    if status:
        params.append(status)
        conds.append(f"s.status = ${len(params)}")
    if project_id:
        params.append(str(project_id))
        conds.append(f"s.project_id = ${len(params)}::uuid")
    params += [limit, (page - 1) * limit]
    rows = await fetch(
        f"""SELECT s.*, p.short_name, p.project_name,
                   count(*) OVER() AS _total_count
            FROM project_report_submissions s
            JOIN projects p ON p.id = s.project_id
            WHERE {' AND '.join(conds)}
            ORDER BY s.uploaded_at DESC
            LIMIT ${len(params) - 1} OFFSET ${len(params)}""",
        *params,
    )
    total = rows[0].pop("_total_count", 0) if rows else 0
    for r in rows[1:]:
        r.pop("_total_count", None)
    return {"success": True, "data": rows,
            "meta": {"page": page, "limit": limit, "total": total}}


@router.get("/submissions/{submission_id}")
async def get_project_submission(
    submission_id: UUID,
    current_user: Annotated[CurrentUser, Depends(require_projects_access)],
) -> dict[str, Any]:
    row = await fetchrow(
        """SELECT s.*, p.short_name, p.project_name
           FROM project_report_submissions s
           JOIN projects p ON p.id = s.project_id
           WHERE s.id = $1::uuid""",
        str(submission_id),
    )
    if row is None:
        raise NotFoundError("Submission not found")
    return {"success": True, "data": row}


@router.post("/submissions/{submission_id}/retry")
async def retry_project_submission(
    submission_id: UUID,
    background_tasks: BackgroundTasks,
    current_user: Annotated[CurrentUser, Depends(require_admin)],
) -> dict[str, Any]:
    row = await fetchrow(
        "SELECT id, status FROM project_report_submissions WHERE id = $1::uuid",
        str(submission_id),
    )
    if row is None:
        raise NotFoundError("Submission not found")
    if row["status"] in ("queued", "parsing"):
        raise ValidationError("Submission is already being processed")
    await execute(
        """UPDATE project_report_submissions
           SET status = 'queued', error_message = NULL,
               retry_count = retry_count + 1, updated_at = now()
           WHERE id = $1::uuid""",
        str(submission_id),
    )
    from app.workers.project_report_worker import process_project_weekly_report

    background_tasks.add_task(process_project_weekly_report, str(submission_id))
    return {"success": True, "data": {"status": "queued"}}


@router.get("/unmapped-fleet-numbers")
async def list_unmapped_fleet_numbers(
    current_user: Annotated[CurrentUser, Depends(require_projects_access)],
) -> dict[str, Any]:
    """Fleet numbers in project reports that don't match plants_master
    and have no alias verdict yet (external ones are settled — hidden)."""
    rows = await fetch(
        """SELECT fleet_number_raw,
                  count(*)::int AS occurrences,
                  count(DISTINCT project_id)::int AS projects,
                  min(week_number)::int AS first_week,
                  max(week_number)::int AS last_week,
                  max(description) AS description
           FROM (
               SELECT fleet_number_raw, project_id, week_number, description
               FROM project_plant_utilization WHERE plant_id IS NULL
               UNION ALL
               SELECT fleet_number_raw, project_id, week_number, description
               FROM project_diesel_consumption WHERE plant_id IS NULL
           ) u
           WHERE fleet_number_raw IS NOT NULL
             AND NOT EXISTS (
                 SELECT 1 FROM project_fleet_aliases a
                 WHERE a.raw_normalized =
                       upper(replace(trim(u.fleet_number_raw), ' ', ''))
             )
           GROUP BY fleet_number_raw
           ORDER BY occurrences DESC"""
    )
    return {"success": True, "data": rows}


@router.get("/fleet-aliases")
async def list_fleet_aliases(
    current_user: Annotated[CurrentUser, Depends(require_projects_access)],
) -> dict[str, Any]:
    """All durable fleet verdicts (for the review UI and for undo)."""
    rows = await fetch(
        """SELECT a.id, a.raw_normalized, a.kind, a.label, a.notes,
                  a.created_at, pm.fleet_number AS plant_fleet_number,
                  pm.description AS plant_description
           FROM project_fleet_aliases a
           LEFT JOIN plants_master pm ON pm.id = a.plant_id
           ORDER BY a.created_at DESC"""
    )
    return {"success": True, "data": rows}


@router.post("/unmapped-fleet-numbers/link")
async def link_unmapped_fleet_number(
    request: Request,
    background_tasks: BackgroundTasks,
    current_user: Annotated[CurrentUser, Depends(require_admin)],
    body: dict[str, Any] = None,
) -> dict[str, Any]:
    """Link a raw fleet number to a plant — backfills ALL historical rows
    in both project tables."""
    body = body or {}
    raw = (body.get("fleet_number_raw") or "").strip()
    plant_id = body.get("plant_id")
    if not raw or not plant_id:
        raise ValidationError("fleet_number_raw and plant_id are required")

    plant = await fetchrow(
        "SELECT id, fleet_number FROM plants_master WHERE id = $1::uuid",
        str(plant_id),
    )
    if plant is None:
        raise NotFoundError("Plant not found")

    updated = 0
    for table in ("project_plant_utilization", "project_diesel_consumption"):
        result = await execute(
            f"""UPDATE {table} SET plant_id = $1::uuid
                WHERE fleet_number_raw = $2 AND plant_id IS NULL""",
            str(plant_id), raw,
        )
        updated += int(result.split()[-1])

    # durable verdict: future imports resolve this raw number automatically
    from app.workers.etl_worker import normalize_fleet_number
    norm = normalize_fleet_number(raw)
    if norm:
        await execute(
            """INSERT INTO project_fleet_aliases
                   (raw_normalized, kind, plant_id, created_by)
               VALUES ($1, 'plant', $2::uuid, $3::uuid)
               ON CONFLICT (raw_normalized) DO UPDATE SET
                   kind = 'plant', plant_id = EXCLUDED.plant_id,
                   label = NULL, created_by = EXCLUDED.created_by""",
            norm, str(plant_id), current_user.id,
        )

    background_tasks.add_task(
        audit_service.log,
        user_id=current_user.id,
        user_email=current_user.email,
        action="update",
        table_name="project_plant_utilization",
        record_id=raw,
        new_values={"fleet_number_raw": raw, "plant_id": str(plant_id),
                    "rows_backfilled": updated, "alias": norm},
        ip_address=get_client_ip(request),
        description=f"Linked fleet {raw} → {plant['fleet_number']} ({updated} rows)",
    )
    return {"success": True,
            "data": {"linked_to": plant["fleet_number"], "rows_backfilled": updated}}


@router.post("/unmapped-fleet-numbers/mark-external")
async def mark_fleet_number_external(
    request: Request,
    background_tasks: BackgroundTasks,
    current_user: Annotated[CurrentUser, Depends(require_admin)],
    body: dict[str, Any] = None,
) -> dict[str, Any]:
    """Settle a raw fleet number as NOT company plant (hired vehicle,
    contractor kit). Rows keep their raw number with plant_id NULL and
    still count in site totals; the number leaves the queue for good."""
    body = body or {}
    raw = (body.get("fleet_number_raw") or "").strip()
    if not raw:
        raise ValidationError("fleet_number_raw is required")
    label = (body.get("label") or "External / hired").strip()

    from app.workers.etl_worker import normalize_fleet_number
    norm = normalize_fleet_number(raw)
    if not norm:
        raise ValidationError(f"{raw!r} does not normalize to a fleet number")

    await execute(
        """INSERT INTO project_fleet_aliases
               (raw_normalized, kind, label, notes, created_by)
           VALUES ($1, 'external', $2, $3, $4::uuid)
           ON CONFLICT (raw_normalized) DO UPDATE SET
               kind = 'external', plant_id = NULL,
               label = EXCLUDED.label, notes = EXCLUDED.notes,
               created_by = EXCLUDED.created_by""",
        norm, label, body.get("notes"), current_user.id,
    )
    background_tasks.add_task(
        audit_service.log,
        user_id=current_user.id,
        user_email=current_user.email,
        action="update",
        table_name="project_fleet_aliases",
        record_id=norm,
        new_values={"fleet_number_raw": raw, "kind": "external", "label": label},
        ip_address=get_client_ip(request),
        description=f"Marked fleet {raw} as external ({label})",
    )
    return {"success": True, "data": {"raw_normalized": norm, "kind": "external"}}


@router.delete("/fleet-aliases/{alias_id}")
async def delete_fleet_alias(
    alias_id: UUID,
    request: Request,
    background_tasks: BackgroundTasks,
    current_user: Annotated[CurrentUser, Depends(require_admin)],
) -> dict[str, Any]:
    """Undo a verdict. An external number returns to the queue; a plant
    alias stops auto-linking future uploads (already-linked rows keep
    their plant_id — use re-resolve after correcting plants_master)."""
    row = await fetchrow(
        "DELETE FROM project_fleet_aliases WHERE id = $1::uuid RETURNING raw_normalized, kind",
        str(alias_id),
    )
    if row is None:
        raise NotFoundError("Alias not found")
    background_tasks.add_task(
        audit_service.log,
        user_id=current_user.id,
        user_email=current_user.email,
        action="delete",
        table_name="project_fleet_aliases",
        record_id=row["raw_normalized"],
        new_values={"kind": row["kind"]},
        ip_address=get_client_ip(request),
        description=f"Removed fleet verdict for {row['raw_normalized']}",
    )
    return {"success": True, "data": {"raw_normalized": row["raw_normalized"]}}


@router.post("/unmapped-fleet-numbers/re-resolve")
async def re_resolve_fleet_numbers(
    request: Request,
    background_tasks: BackgroundTasks,
    current_user: Annotated[CurrentUser, Depends(require_admin)],
) -> dict[str, Any]:
    """Re-attempt deterministic matching for all unlinked fleet rows.

    Useful after a plant is added to (or corrected in) plants_master:
    any NULL plant_id row whose normalized fleet number now matches
    exactly gets linked, in both project tables.
    """
    updated = 0
    for table in ("project_plant_utilization", "project_diesel_consumption"):
        result = await execute(
            f"""UPDATE {table} t SET plant_id = pm.id
                FROM plants_master pm
                WHERE t.plant_id IS NULL
                  AND pm.fleet_number = upper(replace(trim(t.fleet_number_raw), ' ', ''))""",
        )
        updated += int(result.split()[-1])

    if updated:
        background_tasks.add_task(
            audit_service.log,
            user_id=current_user.id,
            user_email=current_user.email,
            action="update",
            table_name="project_plant_utilization",
            record_id="re-resolve",
            new_values={"rows_backfilled": updated},
            ip_address=get_client_ip(request),
            description=f"Re-resolved fleet numbers ({updated} rows linked)",
        )
    return {"success": True, "data": {"rows_backfilled": updated}}


@router.get("/submissions/{submission_id}/download")
async def download_submission_file(
    submission_id: UUID,
    current_user: Annotated[CurrentUser, Depends(require_projects_access)],
) -> dict[str, Any]:
    """One-hour signed link to the original workbook (the Bronze layer —
    stored untouched exactly as the site emailed it)."""
    sub = await fetchrow(
        """SELECT file_path, file_name FROM project_report_submissions
           WHERE id = $1::uuid""",
        str(submission_id),
    )
    if sub is None or not sub["file_path"]:
        raise NotFoundError("Submission file not found")

    from app.core.database import get_supabase_admin_client  # Storage only
    res = get_supabase_admin_client().storage.from_("reports").create_signed_url(
        sub["file_path"], 3600,
        options={"download": sub["file_name"] or True},
    )
    url = (res or {}).get("signedURL") or (res or {}).get("signedUrl")
    if not url:
        raise ValidationError("Could not create a download link for this file")
    return {"success": True,
            "data": {"url": url, "file_name": sub["file_name"]}}


@router.delete("/submissions/{submission_id}")
async def delete_project_submission(
    submission_id: UUID,
    request: Request,
    background_tasks: BackgroundTasks,
    current_user: Annotated[CurrentUser, Depends(require_admin)],
) -> dict[str, Any]:
    """Delete a submission and (if it produced one) that week's report data.

    The weekly-report header delete cascades every per-week operational
    table. Certificates touched only by this week are removed too — they
    come back on re-upload because each workbook carries the full ledger.
    """
    sub = await fetchrow(
        """SELECT s.*, p.short_name FROM project_report_submissions s
           JOIN projects p ON p.id = s.project_id
           WHERE s.id = $1::uuid""",
        str(submission_id),
    )
    if sub is None:
        raise NotFoundError("Submission not found")
    if sub["status"] in ("queued", "parsing"):
        raise ValidationError("Submission is currently processing — wait or retry later")

    pool = get_pool()
    async with pool.acquire() as conn, conn.transaction():
        # Wipe the week's data (cascades all operational tables). Target the
        # (project, year, week) triple, not just weekly_report_id — an older
        # failed submission may have NULL weekly_report_id while the data
        # exists from a later retry.
        deleted_report = await conn.fetchval(
            """DELETE FROM project_weekly_reports
               WHERE project_id = $1::uuid AND year = $2 AND week_number = $3
               RETURNING id""",
            str(sub["project_id"]), sub["year"], sub["week_number"],
        )
        await conn.execute(
            "DELETE FROM project_report_submissions WHERE id = $1::uuid",
            str(submission_id),
        )

    # Storage cleanup is best-effort — data consistency doesn't depend on it.
    try:
        from app.core.database import get_supabase_admin_client  # Storage only
        get_supabase_admin_client().storage.from_("reports").remove([sub["file_path"]])
    except Exception as exc:
        logger.warning("Could not remove storage file",
                       path=sub["file_path"], error=str(exc))

    background_tasks.add_task(
        audit_service.log,
        user_id=current_user.id,
        user_email=current_user.email,
        action="delete",
        table_name="project_report_submissions",
        record_id=str(submission_id),
        old_values={"project": sub["short_name"], "year": sub["year"],
                    "week_number": sub["week_number"], "status": sub["status"]},
        ip_address=get_client_ip(request),
        description=(f"Deleted W{sub['week_number']}/{sub['year']} submission for "
                     f"{sub['short_name']}"
                     + (" incl. weekly report data" if deleted_report else "")),
    )
    return {"success": True,
            "data": {"deleted_week_data": bool(deleted_report),
                     "year": sub["year"], "week_number": sub["week_number"]}}


# ============================================================================
# Operations (weekly-report derived) — the MD/GPM view
#
# Every number here is RECOMPUTED from atomic weekly facts. The workbook's
# own Previous/To-Date columns are never trusted (broken cross-workbook
# links); totals are sums of stored this-week rows, and to-date figures
# (works certified, payments) come from the latest report's ledger.
# ============================================================================


@router.get("/operations")
async def list_project_operations(
    current_user: Annotated[CurrentUser, Depends(require_projects_access)],
) -> dict[str, Any]:
    """Portfolio view: every project with ingested weekly data, with
    recomputed headline totals. Weeks can arrive in any order — all
    aggregates group over whatever (year, week) rows exist."""
    rows = await fetch(
        """
        WITH wr AS (
            -- first/last must be year-aware: W43 2025 precedes W2 2026
            SELECT project_id,
                   count(*)::int                AS weeks_received,
                   (min(year * 100 + week_number) % 100)::int AS first_week,
                   (max(year * 100 + week_number) % 100)::int AS last_week,
                   max(week_ending_date)        AS last_week_ending,
                   max(year)::int               AS latest_year
            FROM project_weekly_reports
            GROUP BY project_id
        ),
        util AS (
            SELECT project_id,
                   COALESCE(sum(hours_worked), 0)    AS hours_worked,
                   COALESCE(sum(breakdown_hours), 0) AS breakdown_hours,
                   COALESCE(sum(standby_hours), 0)   AS standby_hours,
                   COALESCE(sum(plant_cost), 0)      AS plant_cost_ngn,
                   count(DISTINCT fleet_number_raw)::int AS fleet_count
            FROM project_plant_utilization
            GROUP BY project_id
        ),
        diesel AS (
            SELECT project_id, COALESCE(sum(total_litres), 0) AS diesel_litres
            FROM project_diesel_consumption
            GROUP BY project_id
        ),
        latest_report AS (
            SELECT DISTINCT ON (project_id) project_id, id
            FROM project_weekly_reports
            ORDER BY project_id, year DESC, week_number DESC
        ),
        pay AS (
            -- payments are a per-report ledger: read the LATEST report only
            SELECT p.project_id,
                   COALESCE(sum(p.net_amount), 0) AS payments_net_ngn,
                   count(*)::int                  AS payments_count
            FROM project_payments p
            JOIN latest_report lr ON lr.id = p.weekly_report_id
            GROUP BY p.project_id
        ),
        snap AS (
            SELECT DISTINCT ON (project_id) project_id,
                   current_contract_amount, works_certified
            FROM project_contract_summary_snapshot
            ORDER BY project_id, year DESC, week_number DESC
        ),
        pct AS (
            SELECT DISTINCT ON (project_id) project_id, beme_pct_complete
            FROM project_weekly_reports
            WHERE beme_pct_complete IS NOT NULL
            ORDER BY project_id, year DESC, week_number DESC
        )
        SELECT p.id, p.short_name, p.project_name, p.status,
               l.name AS location_name,
               wr.weeks_received, wr.first_week, wr.last_week,
               wr.latest_year, wr.last_week_ending,
               (CURRENT_DATE - wr.last_week_ending)::int AS days_since_last_report,
               util.hours_worked, util.breakdown_hours, util.standby_hours,
               util.plant_cost_ngn, util.fleet_count,
               diesel.diesel_litres,
               pay.payments_net_ngn, pay.payments_count,
               snap.current_contract_amount, snap.works_certified,
               pct.beme_pct_complete
        FROM wr
        JOIN projects p ON p.id = wr.project_id
        LEFT JOIN locations l ON l.id = p.location_id
        LEFT JOIN util   ON util.project_id   = wr.project_id
        LEFT JOIN diesel ON diesel.project_id = wr.project_id
        LEFT JOIN pay    ON pay.project_id    = wr.project_id
        LEFT JOIN snap   ON snap.project_id   = wr.project_id
        LEFT JOIN pct    ON pct.project_id    = wr.project_id
        ORDER BY wr.last_week_ending DESC
        """
    )
    return {"success": True, "data": rows}


@router.get("/{project_id}/overview")
async def get_project_overview(
    project_id: UUID,
    current_user: Annotated[CurrentUser, Depends(require_projects_access)],
) -> dict[str, Any]:
    """The living Contract Summary — every figure computed from ledgers
    and atomic weekly facts per docs/WORKBOOK_ARITHMETIC.md, never from
    the workbook's fossil client-position block."""
    from app.services.project_overview import compute_project_overview

    data = await compute_project_overview(str(project_id))
    if not data:
        raise NotFoundError("Project", str(project_id))
    return {"success": True, "data": data}


@router.get("/{project_id}/work-done")
async def get_project_work_done(
    project_id: UUID,
    current_user: Annotated[CurrentUser, Depends(require_projects_access)],
) -> dict[str, Any]:
    """BEME drill-down: per-item cumulative (stored weeks + baseline/gap
    adjustments, == workbook to-date) + latest-week movement, grouped by
    bill. Quantity and amount progress are independent facts."""
    pid = str(project_id)
    rows = await fetch(
        """SELECT v.*, lm.qty_this_week AS latest_qty, lm.amount_this_week AS latest_amount
           FROM v_project_beme_cumulative v
           LEFT JOIN LATERAL (
               SELECT qty_this_week, amount_this_week
               FROM project_beme_progress p
               WHERE p.item_id = v.item_id
                 AND p.weekly_report_id = (
                     SELECT id FROM project_weekly_reports
                     WHERE project_id = $1::uuid
                     ORDER BY year DESC, week_number DESC LIMIT 1)
           ) lm ON TRUE
           WHERE v.project_id = $1::uuid""",
        pid,
    )
    if not rows:
        return {"success": True, "data": {"bills": []}}

    def code_key(code: str | None) -> tuple:
        parts = []
        for seg in (code or "").split("."):
            num = "".join(ch for ch in seg if ch.isdigit())
            parts.append((int(num) if num else 0, seg))
        return tuple(parts)

    bills: dict[str, dict[str, Any]] = {}
    for r in rows:
        b = bills.setdefault(r["bill_code"] or str(r["bill_no"]), {
            "bill_code": r["bill_code"],
            "bill_name": r["bill_name"],
            "sort_order": r["sort_order"],
            "contract_amount": 0.0,
            "amount_done": 0.0,
            "latest_amount": 0.0,
            "items": [],
        })
        b["contract_amount"] += float(r["contract_amount"] or 0)
        b["amount_done"] += float(r["amount_done"] or 0)
        b["latest_amount"] += float(r["latest_amount"] or 0)
        b["items"].append({
            "item_code": r["item_code"],
            "description": r["description"],
            "unit": r["unit"],
            "contract_qty": r["contract_qty"],
            "rate": r["rate"],
            "contract_amount": r["contract_amount"],
            "qty_done": r["qty_done"],
            "amount_done": r["amount_done"],
            "pct_complete": r["pct_complete"],
            "is_overrun": r["is_overrun"],
            "no_contract_qty": r["no_contract_qty"],
            "latest_qty": r["latest_qty"],
            "latest_amount": r["latest_amount"],
        })

    out = sorted(bills.values(), key=lambda b: (b["sort_order"] or 0, code_key(b["bill_code"])))
    for b in out:
        b["items"].sort(key=lambda i: code_key(i["item_code"]))
        b["pct_complete"] = (
            round(b["amount_done"] / b["contract_amount"], 4)
            if b["contract_amount"] else None
        )
    return {"success": True, "data": {"bills": out}}


@router.get("/{project_id}/costs/summary")
async def get_project_costs_summary(
    project_id: UUID,
    current_user: Annotated[CurrentUser, Depends(require_projects_access)],
) -> dict[str, Any]:
    """Per-category cost position: to-date from the LATEST week's own
    cumulative column (previous + this week — exact even with missing
    weeks), this-week movement, and the stored-weeks sum for trends."""
    pid = str(project_id)
    rows = await fetch(
        """WITH latest AS (
               SELECT id FROM project_weekly_reports
               WHERE project_id = $1::uuid
               ORDER BY year DESC, week_number DESC LIMIT 1)
           SELECT cost_category,
                  coalesce(sum(amount_to_date)  FILTER (WHERE weekly_report_id = (SELECT id FROM latest)), 0) AS to_date,
                  coalesce(sum(amount_this_week) FILTER (WHERE weekly_report_id = (SELECT id FROM latest)), 0) AS this_week,
                  coalesce(sum(amount_this_week), 0) AS stored_weeks
           FROM project_cost_report
           WHERE project_id = $1::uuid AND cost_category IS NOT NULL
           GROUP BY cost_category
           ORDER BY 2 DESC""",
        pid,
    )
    total_to_date = sum(float(r["to_date"] or 0) for r in rows)
    return {"success": True, "data": {
        "categories": rows,
        "total_to_date": round(total_to_date, 2),
    }}


@router.get("/{project_id}/site")
async def get_project_site(
    project_id: UUID,
    current_user: Annotated[CurrentUser, Depends(require_projects_access)],
) -> dict[str, Any]:
    """Site resources, latest week + trends: labour by department,
    subcontractor ledgers (latest report carries the truth), materials
    usage/stock, hired vehicles."""
    pid = str(project_id)
    latest_id = await fetchval(
        """SELECT id FROM project_weekly_reports WHERE project_id = $1::uuid
           ORDER BY year DESC, week_number DESC LIMIT 1""",
        pid,
    )
    if latest_id is None:
        return {"success": True, "data": None}

    labour = await fetch(
        """SELECT block, dept_slot, department, manning_this_week,
                  manning_previous_week, movement, comment
           FROM project_labour_strength
           WHERE weekly_report_id = $1::uuid
           ORDER BY block DESC, dept_slot""",
        str(latest_id),
    )
    labour_trend = await fetch(
        """SELECT wr.year, wr.week_number, wr.week_ending_date,
                  coalesce(sum(l.manning_this_week), 0)::int AS total
           FROM project_weekly_reports wr
           LEFT JOIN project_labour_strength l ON l.weekly_report_id = wr.id
           WHERE wr.project_id = $1::uuid
           GROUP BY wr.year, wr.week_number, wr.week_ending_date
           ORDER BY wr.year, wr.week_number""",
        pid,
    )
    subs = await fetch(
        """SELECT subcontractor_name, description, location, unit,
                  agreed_rate, assigned_qty, qty_to_date, balance_remaining,
                  qty_this_week, amount_this_week, amount_to_date,
                  value_to_completion
           FROM v_project_subcontractors_latest
           WHERE project_id = $1::uuid
           ORDER BY subcontractor_name, description""",
        pid,
    )
    materials = await fetch(
        """SELECT sheet_source, material_name, unit, unit_cost,
                  opening_stock, received, closing_stock, available_for_use,
                  used_works, used_precast, used_mobilisation, used_other,
                  used, variance_qty, variance_value, stock_maintained
           FROM project_materials_stock
           WHERE weekly_report_id = $1::uuid
           ORDER BY sheet_source, material_name""",
        str(latest_id),
    )
    hired = await fetch(
        """SELECT registration_no, description, section, owners,
                  days_worked, rate_ngn, amount_ngn, remarks
           FROM project_hired_vehicles
           WHERE weekly_report_id = $1::uuid
           ORDER BY section, description""",
        str(latest_id),
    )
    hired_to_date = await fetchval(
        """SELECT coalesce(sum(amount_ngn), 0) FROM project_hired_vehicles
           WHERE project_id = $1::uuid""",
        pid,
    )
    return {"success": True, "data": {
        "labour": labour,
        "labour_trend": labour_trend,
        "subcontractors": subs,
        "materials": materials,
        "hired_vehicles": hired,
        "hired_to_date_stored": float(hired_to_date or 0),
    }}


@router.get("/{project_id}/financials/ledgers")
async def get_project_financial_ledgers(
    project_id: UUID,
    current_user: Annotated[CurrentUser, Depends(require_projects_access)],
) -> dict[str, Any]:
    """The two client-money ledgers, verbatim: every certificate (all 19
    workbook columns) and the latest payments ledger."""
    pid = str(project_id)
    certs = await fetch(
        """SELECT cert_number, date_submitted, gross_value_works_done,
                  add_materials_on_site, less_materials_on_site,
                  general_bill_1, total_value_of_work_done,
                  value_of_works_per_cert, total_retention_held,
                  total_net_payment, retention_released, contingency_used,
                  contingency_deducted, fluctuation_materials,
                  advance_received, total_works_executed, advance_recovery,
                  new_total, less_previously_certified
           FROM project_certificates WHERE project_id = $1::uuid
           ORDER BY gross_value_works_done NULLS FIRST""",
        pid,
    )
    payments = await fetch(
        """SELECT payment_date, voucher_number, payment_type, gross_amount,
                  wht, vat, vetting_fee, stamp_duty, other_deductions,
                  net_amount
           FROM v_project_payments_latest WHERE project_id = $1::uuid
           ORDER BY payment_date NULLS FIRST, voucher_number""",
        pid,
    )
    return {"success": True, "data": {"certificates": certs, "payments": payments}}


@router.get("/{project_id}/operations/summary")
async def get_project_operations_summary(
    project_id: UUID,
    current_user: Annotated[CurrentUser, Depends(require_projects_access)],
) -> dict[str, Any]:
    """Drill-down headline for one project: recomputed totals + latest
    contract snapshot + certificate/payment ledger summaries."""
    pid = str(project_id)
    project = await fetchrow(
        """SELECT p.id, p.short_name, p.project_name, p.status,
                  p.current_contract_sum, l.name AS location_name
           FROM projects p LEFT JOIN locations l ON l.id = p.location_id
           WHERE p.id = $1::uuid""",
        pid,
    )
    if project is None:
        raise NotFoundError("Project not found")

    totals = await fetchrow(
        """
        WITH latest_report AS (
            SELECT id FROM project_weekly_reports
            WHERE project_id = $1::uuid
            ORDER BY year DESC, week_number DESC LIMIT 1
        )
        SELECT
          (SELECT count(*)::int FROM project_weekly_reports
            WHERE project_id = $1::uuid)                       AS weeks_received,
          (SELECT max(week_ending_date) FROM project_weekly_reports
            WHERE project_id = $1::uuid)                       AS last_week_ending,
          (SELECT COALESCE(sum(hours_worked), 0)
             FROM project_plant_utilization
            WHERE project_id = $1::uuid)                       AS hours_worked,
          (SELECT COALESCE(sum(breakdown_hours), 0)
             FROM project_plant_utilization
            WHERE project_id = $1::uuid)                       AS breakdown_hours,
          (SELECT COALESCE(sum(standby_hours), 0)
             FROM project_plant_utilization
            WHERE project_id = $1::uuid)                       AS standby_hours,
          (SELECT COALESCE(sum(plant_cost), 0)
             FROM project_plant_utilization
            WHERE project_id = $1::uuid)                       AS plant_cost_ngn,
          (SELECT count(DISTINCT fleet_number_raw)::int
             FROM project_plant_utilization
            WHERE project_id = $1::uuid)                       AS fleet_count,
          (SELECT COALESCE(sum(total_litres), 0)
             FROM project_diesel_consumption
            WHERE project_id = $1::uuid)                       AS diesel_litres,
          (SELECT COALESCE(sum(net_amount), 0) FROM project_payments
            WHERE weekly_report_id IN (SELECT id FROM latest_report))
                                                               AS payments_net_ngn,
          (SELECT count(*)::int FROM project_payments
            WHERE weekly_report_id IN (SELECT id FROM latest_report))
                                                               AS payments_count,
          (SELECT count(*)::int FROM project_certificates
            WHERE project_id = $1::uuid)                       AS certificates_count,
          -- total_net_payment is the sheet's CUMULATIVE running column:
          -- the highest-numbered cert carries the to-date figure
          (SELECT COALESCE(total_net_payment, 0)
             FROM project_certificates
            WHERE project_id = $1::uuid
            ORDER BY NULLIF(regexp_replace(cert_number, '\D', '', 'g'), '')::int
                     DESC NULLS LAST
            LIMIT 1)                                           AS certificates_net_ngn
        """,
        pid,
    )
    snapshot = await fetchrow(
        """SELECT year, week_number, original_contract_amount,
                  current_contract_amount, works_certified, retention_held,
                  advance_unrecovered
           FROM project_contract_summary_snapshot
           WHERE project_id = $1::uuid
           ORDER BY year DESC, week_number DESC LIMIT 1""",
        pid,
    )
    pct = await fetchrow(
        """SELECT year, week_number, beme_pct_complete
           FROM project_weekly_reports
           WHERE project_id = $1::uuid AND beme_pct_complete IS NOT NULL
           ORDER BY year DESC, week_number DESC LIMIT 1""",
        pid,
    )

    # ── commercial position from the LEDGERS (locked 2026-07-11) ────────
    # Contract Summary's client-position block is frozen (~2023); certified
    # and paid figures must come from the cert + payments ledgers.
    commercial = await fetchrow(
        """
        WITH latest_cert AS (
            SELECT gross_value_works_done, total_retention_held,
                   retention_released, advance_recovery
            FROM v_project_certificates
            WHERE project_id = $1::uuid
            ORDER BY cert_sort DESC NULLS LAST LIMIT 1
        ),
        pay AS (
            SELECT
                COALESCE(sum(gross_amount) FILTER
                    (WHERE payment_type ILIKE '%advance%'), 0) AS advances_gross,
                COALESCE(sum(gross_amount) FILTER
                    (WHERE payment_type NOT ILIKE '%advance%'
                        OR payment_type IS NULL), 0)           AS cert_payments_gross,
                COALESCE(sum(net_amount), 0)                   AS payments_net,
                max(payment_date)                              AS last_payment_date
            FROM v_project_payments_latest
            WHERE project_id = $1::uuid
        )
        SELECT lc.gross_value_works_done      AS certified_cumulative,
               lc.total_retention_held        AS retention_held,
               lc.retention_released,
               p.advances_gross,
               p.cert_payments_gross,
               p.payments_net,
               p.last_payment_date,
               lc.gross_value_works_done - p.cert_payments_gross
                                              AS certified_unpaid
        FROM latest_cert lc, pay p
        """,
        pid,
    )

    return {"success": True, "data": {
        "project": project,
        "totals": totals,
        "latest_snapshot": snapshot,
        "latest_pct": pct,
        "commercial": commercial,
    }}


@router.get("/{project_id}/operations/series")
async def get_project_operations_series(
    project_id: UUID,
    current_user: Annotated[CurrentUser, Depends(require_projects_access)],
    granularity: str = Query("week", pattern=r"^(week|month)$"),
) -> dict[str, Any]:
    """Per-week or per-month series, recomputed from atomic rows.

    Monthly buckets group weeks by the month of their week-ending date
    (company weeks run Sat–Fri), so a month is the sum of the weeks that
    END in it.
    """
    pid = str(project_id)
    exists = await fetchval(
        "SELECT 1 FROM projects WHERE id = $1::uuid", pid
    )
    if not exists:
        raise NotFoundError("Project not found")

    if granularity == "week":
        rows = await fetch(
            """
            SELECT wr.year, wr.week_number, wr.week_ending_date,
                   wr.beme_pct_complete,
                   COALESCE(u.hours_worked, 0)     AS hours_worked,
                   COALESCE(u.breakdown_hours, 0)  AS breakdown_hours,
                   COALESCE(u.standby_hours, 0)    AS standby_hours,
                   COALESCE(u.plant_cost_ngn, 0)   AS plant_cost_ngn,
                   COALESCE(u.plants_on_site, 0)   AS plants_on_site,
                   COALESCE(d.diesel_litres, 0)    AS diesel_litres,
                   COALESCE(lab.labour_total, 0)   AS labour_total,
                   snap.works_certified
            FROM project_weekly_reports wr
            LEFT JOIN (
                SELECT weekly_report_id,
                       sum(hours_worked)    AS hours_worked,
                       sum(breakdown_hours) AS breakdown_hours,
                       sum(standby_hours)   AS standby_hours,
                       sum(plant_cost)      AS plant_cost_ngn,
                       count(DISTINCT fleet_number_raw) AS plants_on_site
                FROM project_plant_utilization
                WHERE project_id = $1::uuid
                GROUP BY weekly_report_id
            ) u ON u.weekly_report_id = wr.id
            LEFT JOIN (
                SELECT weekly_report_id, sum(total_litres) AS diesel_litres
                FROM project_diesel_consumption
                WHERE project_id = $1::uuid
                GROUP BY weekly_report_id
            ) d ON d.weekly_report_id = wr.id
            LEFT JOIN (
                SELECT weekly_report_id, sum(manning_this_week) AS labour_total
                FROM project_labour_strength
                WHERE project_id = $1::uuid
                GROUP BY weekly_report_id
            ) lab ON lab.weekly_report_id = wr.id
            LEFT JOIN project_contract_summary_snapshot snap
                   ON snap.weekly_report_id = wr.id
            WHERE wr.project_id = $1::uuid
            ORDER BY wr.year, wr.week_number
            """,
            pid,
        )
    else:
        rows = await fetch(
            """
            SELECT to_char(date_trunc('month', wr.week_ending_date),
                           'YYYY-MM')                          AS month,
                   count(*)::int                               AS weeks_in_month,
                   max(wr.beme_pct_complete)                   AS beme_pct_complete,
                   COALESCE(sum(u.hours_worked), 0)            AS hours_worked,
                   COALESCE(sum(u.breakdown_hours), 0)         AS breakdown_hours,
                   COALESCE(sum(u.standby_hours), 0)           AS standby_hours,
                   COALESCE(sum(u.plant_cost_ngn), 0)          AS plant_cost_ngn,
                   COALESCE(sum(d.diesel_litres), 0)           AS diesel_litres,
                   max(snap.works_certified)                   AS works_certified
            FROM project_weekly_reports wr
            LEFT JOIN (
                SELECT weekly_report_id,
                       sum(hours_worked)    AS hours_worked,
                       sum(breakdown_hours) AS breakdown_hours,
                       sum(standby_hours)   AS standby_hours,
                       sum(plant_cost)      AS plant_cost_ngn
                FROM project_plant_utilization
                WHERE project_id = $1::uuid
                GROUP BY weekly_report_id
            ) u ON u.weekly_report_id = wr.id
            LEFT JOIN (
                SELECT weekly_report_id, sum(total_litres) AS diesel_litres
                FROM project_diesel_consumption
                WHERE project_id = $1::uuid
                GROUP BY weekly_report_id
            ) d ON d.weekly_report_id = wr.id
            LEFT JOIN project_contract_summary_snapshot snap
                   ON snap.weekly_report_id = wr.id
            WHERE wr.project_id = $1::uuid
            GROUP BY date_trunc('month', wr.week_ending_date)
            ORDER BY 1
            """,
            pid,
        )
    return {"success": True, "data": rows, "meta": {"granularity": granularity}}


@router.get("/{project_id}/operations/financials")
async def get_project_operations_financials(
    project_id: UUID,
    current_user: Annotated[CurrentUser, Depends(require_projects_access)],
) -> dict[str, Any]:
    """Weekly earnings vs costs — is the site gaining or losing?

    earnings = Works Completed subtotal + 7.5% VAT (the sheet's own
    formula, verified to the kobo against its Net Earnings row).
    costs    = recomputed from atomic cost-report rows (category rows
    only; the sheet's total row is used as a cross-check, not as data).
    """
    pid = str(project_id)
    if not await fetchval("SELECT 1 FROM projects WHERE id = $1::uuid", pid):
        raise NotFoundError("Project not found")

    rows = await fetch(
        """
        SELECT wr.year, wr.week_number, wr.week_ending_date,
               COALESCE(beme.works, 0)                          AS works_value,
               cost.by_category                                 AS cost_by_category,
               COALESCE(cost.total, 0)                          AS cost_total,
               ws_net.value                                     AS sheet_net,
               ago.qty                                          AS diesel_charged_litres,
               ago.rate                                         AS diesel_rate,
               ago.amount                                       AS diesel_cost,
               COALESCE(d.litres, 0)                            AS diesel_logged_litres,
               fl.flags                                         AS flags
        FROM project_weekly_reports wr
        -- earnings: recomputed from atomic BEME item movement, never
        -- from the sheet's own rollup (that is a cross-check now)
        LEFT JOIN LATERAL (
            SELECT sum(amount_this_week) AS works
            FROM project_beme_progress WHERE weekly_report_id = wr.id
        ) beme ON TRUE
        LEFT JOIN LATERAL (
            SELECT value FROM project_weekly_summary
            WHERE weekly_report_id = wr.id AND section = 'Costs to Date'
              AND item LIKE 'Net Earnings%' AND metric = 'this_week' LIMIT 1
        ) ws_net ON TRUE
        LEFT JOIN LATERAL (
            SELECT sum(cat_total) AS total,
                   jsonb_object_agg(cost_category, cat_total) AS by_category
            FROM (
                SELECT cost_category, sum(amount_this_week) AS cat_total
                FROM project_cost_report
                WHERE weekly_report_id = wr.id AND cost_category IS NOT NULL
                GROUP BY cost_category
            ) c
        ) cost ON TRUE
        -- diesel money truth: the Cost Report AGO row (litres charged @ rate)
        LEFT JOIN LATERAL (
            SELECT quantity_this_week AS qty, rate_ngn AS rate,
                   amount_this_week AS amount
            FROM project_cost_report
            WHERE weekly_report_id = wr.id AND description = 'Diesel'
              AND cost_category = 'AGO'
            ORDER BY amount_this_week DESC NULLS LAST LIMIT 1
        ) ago ON TRUE
        -- attribution log (may be a stale copy — see flags)
        LEFT JOIN LATERAL (
            SELECT sum(total_litres) AS litres
            FROM project_diesel_consumption WHERE weekly_report_id = wr.id
        ) d ON TRUE
        LEFT JOIN LATERAL (
            SELECT jsonb_agg(jsonb_build_object(
                       'sheet', sheet_name, 'type', flag_type,
                       'severity', severity, 'message', message)
                   ORDER BY severity DESC) AS flags
            FROM project_sheet_flags
            WHERE weekly_report_id = wr.id
              AND flag_type IN ('stale_copy', 'frozen_column',
                                'cross_check_fail', 'chain_break')
        ) fl ON TRUE
        WHERE wr.project_id = $1::uuid
        ORDER BY wr.year, wr.week_number
        """,
        pid,
    )

    weeks: list[dict[str, Any]] = []
    warnings: list[str] = []
    cum_net = 0.0
    for r in rows:
        works = float(r["works_value"] or 0)
        vat = round(works * 0.075, 2)
        earnings = works + vat
        cost_total = float(r["cost_total"] or 0)
        net = earnings - cost_total
        cum_net += net

        # cross-check: our BEME-computed net vs the sheet's Net Earnings
        sheet_net = r["sheet_net"]
        if sheet_net is not None and abs(float(sheet_net) - net) > 1.0:
            warnings.append(
                f"W{r['week_number']}: recomputed net ₦{net:,.2f} ≠ "
                f"sheet net ₦{float(sheet_net):,.2f}"
            )

        by_cat = r["cost_by_category"] or {}
        charged = float(r["diesel_charged_litres"] or 0)
        logged = float(r["diesel_logged_litres"] or 0)
        weeks.append({
            "year": r["year"],
            "week_number": r["week_number"],
            "week_ending_date": r["week_ending_date"],
            "works_value": works,
            "vat": vat,
            "earnings": earnings,
            "cost_total": cost_total,
            "cost_by_category": by_cat,
            "diesel_cost": float(r["diesel_cost"] or 0),
            "diesel_rate": float(r["diesel_rate"] or 0) or None,
            "diesel_litres": charged,          # the money truth (charged)
            "diesel_logged_litres": logged,    # attribution log
            "net": net,
            "cumulative_net": round(cum_net, 2),
            "sheet_net": float(sheet_net) if sheet_net is not None else None,
            "flags": r["flags"] or [],
        })

    totals = {
        "earnings": round(sum(w["earnings"] for w in weeks), 2),
        "cost_total": round(sum(w["cost_total"] for w in weeks), 2),
        "net": round(cum_net, 2),
        "diesel_cost": round(sum(w["diesel_cost"] for w in weeks), 2),
        "diesel_litres": round(sum(w["diesel_litres"] for w in weeks), 2),
        "weeks_gaining": sum(1 for w in weeks if w["net"] > 0),
        "weeks_losing": sum(1 for w in weeks if w["net"] < 0),
        "cost_by_category": {},
    }
    cat_totals: dict[str, float] = {}
    for w in weeks:
        for cat, v in (w["cost_by_category"] or {}).items():
            cat_totals[cat] = cat_totals.get(cat, 0.0) + float(v or 0)
    totals["cost_by_category"] = {k: round(v, 2) for k, v in
                                  sorted(cat_totals.items(), key=lambda x: -x[1])}

    # BEME bill-level progress from the latest week's Works Completed section
    bills = await fetch(
        """
        SELECT ws.item,
               max(ws.value) FILTER (WHERE ws.metric = 'this_week')   AS this_week,
               max(ws.value) FILTER (WHERE ws.metric = 'pct_complete') AS pct_complete
        FROM project_weekly_summary ws
        JOIN (SELECT id FROM project_weekly_reports
              WHERE project_id = $1::uuid
              ORDER BY year DESC, week_number DESC LIMIT 1) lr
          ON lr.id = ws.weekly_report_id
        WHERE ws.section = 'Works Completed'
          AND ws.item NOT ILIKE 'SUB-TOTAL%'
          AND ws.item NOT ILIKE 'Add VAT%'
          AND ws.item NOT ILIKE 'Total%'
        GROUP BY ws.item
        ORDER BY max(ws.value) FILTER (WHERE ws.metric = 'pct_complete')
                 DESC NULLS LAST
        """,
        pid,
    )

    return {"success": True, "data": {
        "weeks": weeks,
        "totals": totals,
        "bills": bills,
        "cross_check_warnings": warnings,
    }}


@router.get("/{project_id}/operations/plants")
async def get_project_operations_plants(
    project_id: UUID,
    current_user: Annotated[CurrentUser, Depends(require_projects_access)],
) -> dict[str, Any]:
    """Per-plant totals across every ingested week: hours, breakdowns,
    plant cost, diesel — joined to the fleet register where resolved."""
    pid = str(project_id)
    if not await fetchval("SELECT 1 FROM projects WHERE id = $1::uuid", pid):
        raise NotFoundError("Project not found")

    rows = await fetch(
        """
        WITH util AS (
            SELECT fleet_number_raw,
                   max(plant_id::text)              AS plant_id,
                   max(description)                 AS description,
                   max(plant_category)              AS plant_category,
                   count(DISTINCT week_number)::int AS weeks_seen,
                   COALESCE(sum(hours_worked), 0)   AS hours_worked,
                   COALESCE(sum(breakdown_hours), 0) AS breakdown_hours,
                   COALESCE(sum(standby_hours), 0)  AS standby_hours,
                   COALESCE(sum(plant_cost), 0)     AS plant_cost_ngn
            FROM project_plant_utilization
            WHERE project_id = $1::uuid
            GROUP BY fleet_number_raw
        ),
        diesel AS (
            SELECT fleet_number_raw,
                   max(plant_id::text)              AS plant_id,
                   max(description)                 AS description,
                   count(DISTINCT week_number)::int AS weeks_seen,
                   COALESCE(sum(total_litres), 0)   AS diesel_litres
            FROM project_diesel_consumption
            WHERE project_id = $1::uuid
            GROUP BY fleet_number_raw
        )
        -- FULL join: some equipment appears only in the diesel sheet
        -- (generators, service vehicles) and must still be listed
        SELECT COALESCE(u.fleet_number_raw, d.fleet_number_raw) AS fleet_number_raw,
               COALESCE(u.plant_id, d.plant_id)                 AS plant_id,
               pm.fleet_number,
               COALESCE(pm.description, u.description, d.description) AS description,
               u.plant_category,
               pm.condition,
               COALESCE(u.weeks_seen, d.weeks_seen)             AS weeks_seen,
               COALESCE(u.hours_worked, 0)                      AS hours_worked,
               COALESCE(u.breakdown_hours, 0)                   AS breakdown_hours,
               COALESCE(u.standby_hours, 0)                     AS standby_hours,
               COALESCE(u.plant_cost_ngn, 0)                    AS plant_cost_ngn,
               COALESCE(d.diesel_litres, 0)                     AS diesel_litres
        FROM util u
        FULL OUTER JOIN diesel d ON d.fleet_number_raw = u.fleet_number_raw
        LEFT JOIN plants_master pm
               ON pm.id = COALESCE(u.plant_id, d.plant_id)::uuid
        ORDER BY COALESCE(u.hours_worked, 0) DESC,
                 COALESCE(d.diesel_litres, 0) DESC
        """,
        pid,
    )
    return {"success": True, "data": rows}


@router.get("/review-queue")
async def get_review_queue(
    current_user: Annotated[CurrentUser, Depends(require_admin)],
    sheet: str | None = None,
    reason: str | None = None,
    field: str | None = None,
    resolved: bool | None = False,
    page: int = 1,
    page_size: int = 50,
) -> dict[str, Any]:
    """Paginated register review queue (admin). resolved=None → all."""
    from app.services import register_review_service as review

    page = max(1, page)
    page_size = min(max(1, page_size), 200)
    pool = get_pool()
    async with pool.acquire() as conn:
        data = await review.list_review_queue(
            conn, sheet=sheet, reason=reason, field=field,
            resolved=resolved, page=page, page_size=page_size,
        )
    return {"success": True, "data": data}


@router.get("/review-queue/summary")
async def get_review_queue_summary(
    current_user: Annotated[CurrentUser, Depends(require_admin)],
) -> dict[str, Any]:
    from app.services import register_review_service as review

    pool = get_pool()
    async with pool.acquire() as conn:
        data = await review.summarize_review_queue(conn)
    return {"success": True, "data": data}


@router.post("/review-queue/{item_id}/resolve")
async def resolve_review_queue_item(
    item_id: str,
    request: Request,
    background_tasks: BackgroundTasks,
    current_user: Annotated[CurrentUser, Depends(require_admin)],
    body: dict[str, Any] | None = None,
) -> dict[str, Any]:
    """Resolve one item. Body {"value": "..."} applies to the project;
    {"value": null} or empty dismisses."""
    from app.services import register_review_service as review

    value = (body or {}).get("value")
    pool = get_pool()
    async with pool.acquire() as conn:
        result = await review.resolve_review_item(conn, item_id, current_user.id, value)

    background_tasks.add_task(
        audit_service.log,
        user_id=current_user.id,
        user_email=current_user.email,
        action="update",
        table_name="project_register_review_queue",
        record_id=item_id,
        new_values=result,
        ip_address=get_client_ip(request),
        description=f"Resolved review item ({'dismissed' if result['dismissed'] else 'applied'})",
    )
    return {"success": True, "data": result}


@router.post("/review-queue/bulk-dismiss")
async def bulk_dismiss_review_items(
    request: Request,
    background_tasks: BackgroundTasks,
    current_user: Annotated[CurrentUser, Depends(require_admin)],
    body: dict[str, Any] = None,
) -> dict[str, Any]:
    """Dismiss all unresolved items with a given reason (+optional field)."""
    from app.services import register_review_service as review

    body = body or {}
    reason = body.get("reason")
    if not reason:
        raise ValidationError("'reason' is required")
    pool = get_pool()
    async with pool.acquire() as conn:
        count = await review.bulk_dismiss(
            conn, current_user.id, reason=reason, field=body.get("field")
        )

    background_tasks.add_task(
        audit_service.log,
        user_id=current_user.id,
        user_email=current_user.email,
        action="update",
        table_name="project_register_review_queue",
        record_id=reason,
        new_values={"dismissed": count, "reason": reason, "field": body.get("field")},
        ip_address=get_client_ip(request),
        description=f"Bulk-dismissed {count} review items (reason={reason})",
    )
    return {"success": True, "data": {"dismissed": count}}


@router.get("/benchmarks")
async def get_project_benchmarks(
    current_user: Annotated[CurrentUser, Depends(require_projects_access)],
) -> dict[str, Any]:
    """Register benchmarks: per-type contract-value quartiles + actual
    award→completion delivery times (from v_project_benchmarks_by_type)."""
    rows = await fetch(
        "SELECT * FROM v_project_benchmarks_by_type ORDER BY n_projects DESC"
    )
    return {"success": True, "data": rows}


@router.get("/stats")
async def get_project_stats(
    current_user: Annotated[CurrentUser, Depends(require_projects_access)],
    is_legacy: bool | None = Query(None, description="Filter stats by legacy status"),
) -> dict[str, Any]:
    """Dashboard summary: counts by status, total contract value, top clients."""
    where = "TRUE"
    params: list[Any] = []
    if is_legacy is not None:
        params.append(is_legacy)
        where = "is_legacy = $1::boolean"

    stats_row = await fetchrow(
        f"""SELECT
               count(*)::int AS total,
               count(*) FILTER (WHERE status = 'active')::int AS active,
               count(*) FILTER (WHERE status = 'completed')::int AS completed,
               count(*) FILTER (WHERE status = 'on_hold')::int AS on_hold,
               count(*) FILTER (WHERE status = 'retention_period')::int AS retention_period,
               count(*) FILTER (WHERE status = 'cancelled')::int AS cancelled,
               count(*) FILTER (WHERE is_legacy)::int AS legacy,
               count(*) FILTER (WHERE NOT is_legacy)::int AS non_legacy,
               COALESCE(SUM(current_contract_sum), 0)::float AS total_contract_value,
               count(DISTINCT client)::int AS total_clients
           FROM projects WHERE {where}""",
        *params,
    )

    top_clients = await fetch(
        f"""SELECT client,
                  count(*)::int AS project_count,
                  COALESCE(SUM(current_contract_sum), 0)::float AS total_value
           FROM projects WHERE {where}
           GROUP BY client
           ORDER BY project_count DESC
           LIMIT 10""",
        *params,
    )

    return {
        "success": True,
        "data": {
            "totals": stats_row,
            "top_clients": top_clients,
        },
    }


@router.get("/clients")
async def list_clients(
    current_user: Annotated[CurrentUser, Depends(require_projects_access)],
) -> dict[str, Any]:
    """List distinct client names for filter dropdowns."""
    rows = await fetch(
        "SELECT DISTINCT client FROM projects ORDER BY client"
    )
    return {
        "success": True,
        "data": [r["client"] for r in rows],
    }


@router.post("/import/award-letters")
async def import_award_letters(
    request: Request,
    background_tasks: BackgroundTasks,
    current_user: Annotated[CurrentUser, Depends(require_admin)],
    file: UploadFile = File(...),
) -> dict[str, Any]:
    """Import projects from an Award Letters Excel workbook.

    Deletes existing legacy projects first, then parses all sheets
    and batch-inserts in a single transaction for a clean reimport.
    """
    from app.services.award_letters_parser import parse_award_letters_excel

    if not file.filename:
        raise ValidationError("File name is required")
    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
    if ext not in ("xlsx", "xls"):
        raise ValidationError("Only .xlsx and .xls files are accepted")

    file_content = await file.read()
    if len(file_content) > 10 * 1024 * 1024:
        raise ValidationError("File too large (max 10MB)")

    # Client default states give the parser its final state-resolution
    # fallback (clients.default_state_id, seeded from client names).
    from app.services.award_letters_import import (
        fetch_client_default_states,
        persist_award_letters,
    )

    pool = get_pool()
    async with pool.acquire() as conn:
        client_defaults = await fetch_client_default_states(conn)

    parsed = parse_award_letters_excel(file_content, client_defaults)

    if not parsed["projects"]:
        return {
            "success": False,
            "error": "No projects found in the uploaded file",
            "data": {
                "errors": parsed["errors"][:20],
                "warnings": parsed["warnings"][:20],
            },
        }

    # Defensive persistence: the DB can be busy (pooler contention, lock
    # waits). Each attempt gets a FRESH copy of the parse (persist mutates
    # its input) and everything is transactional — a failed attempt saves
    # nothing, so retrying is always safe.
    import asyncio as _asyncio
    import copy as _copy

    import asyncpg as _asyncpg

    from app.core.exceptions import DatabaseError

    stats = None
    last_err: Exception | None = None
    for attempt in (1, 2, 3):
        try:
            async with pool.acquire() as conn:
                stats = await persist_award_letters(
                    conn, _copy.deepcopy(parsed), current_user.id
                )
            break
        except (TimeoutError, _asyncio.TimeoutError, _asyncpg.PostgresError, OSError) as exc:
            last_err = exc
            logger.warning(
                "Award letters import attempt failed",
                attempt=attempt,
                error=f"{type(exc).__name__}: {exc}",
            )
            if attempt < 3:
                await _asyncio.sleep(2 * attempt)
    if stats is not None and stats["created"] == 0 and parsed["projects"]:
        first_errors = "; ".join(
            f"{e.get('sheet')}: {e.get('error', '')[:120]}"
            for e in stats["insert_errors"][:3]
        ) or "no row-level errors captured"
        raise DatabaseError(
            message=(
                f"Import saved 0 of {len(parsed['projects'])} projects — "
                f"every row was rejected. First errors: {first_errors}"
            ),
            operation="import_award_letters",
            retryable=False,
        )

    if stats is None:
        raise DatabaseError(
            message=(
                "Import failed after 3 attempts — the database timed out or "
                f"was busy ({type(last_err).__name__}). Nothing was saved; "
                "it is safe to retry in a minute."
            ),
            operation="import_award_letters",
        )

    background_tasks.add_task(
        audit_service.log,
        user_id=current_user.id,
        user_email=current_user.email,
        action="import",
        table_name="projects",
        record_id=parsed["import_batch_id"],
        new_values={
            "file_name": file.filename,
            "sheets_processed": parsed["sheets_processed"],
            "total_parsed": len(parsed["projects"]),
            "created": stats["created"],
            "deleted": stats["deleted"],
            "clients_upserted": stats["clients_upserted"],
            "review_queued": stats["review_queued"],
            "errors": len(stats["insert_errors"]),
        },
        ip_address=get_client_ip(request),
        description=(
            f"Reimported {stats['created']} legacy projects "
            f"(deleted {stats['deleted']}, queued {stats['review_queued']} for review)"
        ),
    )

    if stats["created"] > 0:
        broadcast("projects", "import", f"{stats['created']} projects imported")

    return {
        "success": True,
        "data": {
            "import_batch_id": parsed["import_batch_id"],
            "sheets_processed": parsed["sheets_processed"],
            "total_parsed": len(parsed["projects"]),
            "created": stats["created"],
            "deleted": stats["deleted"],
            "clients_upserted": stats["clients_upserted"],
            "review_queued": stats["review_queued"],
            "errors": stats["insert_errors"][:20],
            "warnings": parsed["warnings"][:20],
            "parse_errors": parsed["errors"][:20],
        },
    }


@router.get("/linkable")
async def list_linkable_projects(
    current_user: Annotated[CurrentUser, Depends(require_projects_access)],
) -> dict[str, Any]:
    """List non-legacy projects that don't yet have a linked location."""
    rows = await fetch(
        """SELECT p.id, p.project_name, p.client, p.status
           FROM projects p
           WHERE NOT p.is_legacy
             AND NOT EXISTS (
                 SELECT 1 FROM locations l WHERE l.project_id = p.id
             )
           ORDER BY p.project_name"""
    )
    return {"success": True, "data": rows}


# ============================================================================
# Standard CRUD
# ============================================================================


@router.get("")
async def list_projects(
    current_user: Annotated[CurrentUser, Depends(require_projects_access)],
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=100),
    search: str | None = None,
    client: str | None = Query(None),
    state_id: UUID | None = None,
    status: str | None = Query(
        None, pattern=r"^(active|completed|on_hold|cancelled|retention_period|legacy)$"
    ),
    is_legacy: bool | None = Query(None, description="Filter by legacy status"),
    project_type: str | None = Query(
        None,
        pattern=r"^(road|bridge|drainage|building|airport|water|infrastructure|other)$",
    ),
    work_nature: str | None = Query(
        None,
        pattern=r"^(construction|dualization|rehabilitation|maintenance|emergency_repair|completion)$",
    ),
    register_source: str | None = Query(
        None, pattern=r"^(award_letters_workbook|manual|weekly_report_inferred)$"
    ),
    sort_by: str = Query("created_at"),
    sort_order: str = Query("desc", pattern=r"^(asc|desc)$"),
) -> dict[str, Any]:
    """List projects with filtering, search, and pagination.

    Single query with count(*) OVER() for pagination — no N+1.
    """
    conds: list[str] = []
    params: list[Any] = []

    if search:
        params.append(f"%{search}%")
        n = len(params)
        conds.append(
            f"(v.project_name ILIKE ${n} OR v.client ILIKE ${n} "
            f"OR v.short_name ILIKE ${n})"
        )
    if client:
        params.append(client)
        conds.append(f"upper(v.client) = upper(${len(params)})")
    if state_id:
        params.append(str(state_id))
        conds.append(f"v.state_id = ${len(params)}::uuid")
    if status:
        params.append(status)
        conds.append(f"v.status = ${len(params)}")
    if is_legacy is not None:
        params.append(is_legacy)
        conds.append(f"v.is_legacy = ${len(params)}::boolean")
    if project_type:
        params.append(project_type)
        conds.append(f"v.project_type = ${len(params)}")
    if work_nature:
        params.append(work_nature)
        conds.append(f"v.work_nature = ${len(params)}")
    if register_source:
        params.append(register_source)
        conds.append(f"v.register_source = ${len(params)}")

    where = " AND ".join(conds) if conds else "TRUE"
    safe_sort = sort_by if sort_by in _ALLOWED_SORT_COLUMNS else "created_at"
    safe_order = "DESC" if sort_order == "desc" else "ASC"

    offset = (page - 1) * limit
    params.append(limit)
    params.append(offset)

    rows = await fetch(
        f"""SELECT v.*, count(*) OVER() AS _total_count
            FROM v_projects_summary v
            WHERE {where}
            ORDER BY v.{safe_sort} {safe_order} NULLS LAST
            LIMIT ${len(params) - 1} OFFSET ${len(params)}""",
        *params,
    )

    total = rows[0].pop("_total_count", 0) if rows else 0
    for r in rows[1:]:
        r.pop("_total_count", None)

    return {
        "success": True,
        "data": rows,
        "meta": {
            "page": page,
            "limit": limit,
            "total": total,
            "total_pages": (total + limit - 1) // limit if total > 0 else 0,
            "has_more": page * limit < total,
        },
    }


@router.get("/{project_id}")
async def get_project(
    project_id: UUID,
    current_user: Annotated[CurrentUser, Depends(require_projects_access)],
) -> dict[str, Any]:
    """Get a single project by ID."""
    row = await fetchrow(
        """SELECT v.*
           FROM v_projects_summary v
           WHERE v.id = $1::uuid""",
        str(project_id),
    )
    if not row:
        raise NotFoundError("Project", str(project_id))
    return {"success": True, "data": row}


@router.get("/{project_id}/milestones")
async def get_project_milestones(
    project_id: UUID,
    current_user: Annotated[CurrentUser, Depends(require_projects_access)],
) -> dict[str, Any]:
    """Get project milestone timeline data."""
    row = await fetchrow(
        "SELECT * FROM projects WHERE id = $1::uuid", str(project_id)
    )
    if not row:
        raise NotFoundError("Project", str(project_id))

    today = date.today()
    milestone_defs = [
        ("award_date", "Award Date"),
        ("commencement_date", "Commencement"),
        ("original_completion_date", "Original Completion"),
        ("revised_completion_date", "Revised Completion"),
        ("substantial_completion_date", "Substantial Completion"),
        ("final_completion_date", "Final Completion"),
        ("maintenance_cert_date", "Maintenance Certificate"),
        ("retention_application_date", "Retention Application"),
    ]

    milestones = []
    for field, label in milestone_defs:
        val = row.get(field)
        if val is None:
            milestones.append({"key": field, "label": label, "date": None, "status": "not_set"})
        else:
            # _record_to_dict may return date as ISO string — parse if needed
            if isinstance(val, str):
                try:
                    val_date = date.fromisoformat(val)
                except ValueError:
                    val_date = None
            elif isinstance(val, date):
                val_date = val
            else:
                val_date = None

            if val_date:
                status = "completed" if val_date <= today else "upcoming"
                milestones.append({
                    "key": field, "label": label,
                    "date": val_date.isoformat(), "status": status,
                })
            else:
                milestones.append({"key": field, "label": label, "date": str(val), "status": "unknown"})

    orig = row.get("original_duration_months")
    ext = row.get("extension_of_time_months")
    duration = {
        "original_months": orig,
        "extension_months": ext,
        "total_months": (orig or 0) + (ext or 0) if orig else None,
    }

    return {"success": True, "data": {"milestones": milestones, "duration": duration}}


@router.post("", status_code=201)
async def create_project(
    project: ProjectCreate,
    request: Request,
    background_tasks: BackgroundTasks,
    current_user: Annotated[CurrentUser, Depends(require_admin)],
) -> dict[str, Any]:
    """Create a new project (admin only)."""
    # python-mode dump: asyncpg needs real date/UUID objects, not strings
    data = project.model_dump(exclude_none=True)
    data["created_by"] = current_user.id
    data["updated_by"] = current_user.id

    cols = list(data.keys())
    vals = list(data.values())
    placeholders = []
    for i, col in enumerate(cols):
        if col in _UUID_FIELDS:
            placeholders.append(f"${i + 1}::uuid")
        else:
            placeholders.append(f"${i + 1}")

    created = await fetchrow(
        f"INSERT INTO projects ({', '.join(cols)}) "
        f"VALUES ({', '.join(placeholders)}) RETURNING *",
        *vals,
    )

    background_tasks.add_task(
        audit_service.log,
        user_id=current_user.id,
        user_email=current_user.email,
        action="create",
        table_name="projects",
        record_id=str(created["id"]),
        new_values=project.model_dump(exclude_none=True, mode="json"),
        ip_address=get_client_ip(request),
        description=f"Created project: {project.project_name}",
    )

    broadcast("projects", "create")
    return {"success": True, "data": created}


@router.patch("/{project_id}")
async def update_project(
    project_id: UUID,
    project: ProjectUpdate,
    request: Request,
    background_tasks: BackgroundTasks,
    current_user: Annotated[CurrentUser, Depends(require_admin)],
) -> dict[str, Any]:
    """Update an existing project (admin only)."""
    # python-mode dump: asyncpg needs real date/UUID objects, not strings
    update_data = project.model_dump(exclude_none=True)
    if not update_data:
        raise ValidationError("No fields to update")

    existing = await fetchrow(
        "SELECT * FROM projects WHERE id = $1::uuid", str(project_id)
    )
    if not existing:
        raise NotFoundError("Project", str(project_id))

    old_values = {k: existing.get(k) for k in update_data if k in existing}
    update_data["updated_by"] = current_user.id

    set_parts: list[str] = []
    params: list[Any] = []
    for key, val in update_data.items():
        params.append(val)
        if key in _UUID_FIELDS:
            set_parts.append(f"{key} = ${len(params)}::uuid")
        else:
            set_parts.append(f"{key} = ${len(params)}")

    params.append(str(project_id))
    updated = await fetchrow(
        f"""UPDATE projects SET {', '.join(set_parts)}
            WHERE id = ${len(params)}::uuid RETURNING *""",
        *params,
    )

    background_tasks.add_task(
        audit_service.log,
        user_id=current_user.id,
        user_email=current_user.email,
        action="update",
        table_name="projects",
        record_id=str(project_id),
        old_values=old_values,
        new_values=project.model_dump(exclude_none=True, mode="json"),
        ip_address=get_client_ip(request),
        description=f"Updated project: {existing.get('project_name')}",
    )

    broadcast("projects", "update")
    return {"success": True, "data": updated}


@router.delete("/{project_id}")
async def delete_project(
    project_id: UUID,
    request: Request,
    background_tasks: BackgroundTasks,
    current_user: Annotated[CurrentUser, Depends(require_admin)],
) -> dict[str, Any]:
    """Delete a project (admin only)."""
    existing = await fetchrow(
        "SELECT id, project_name FROM projects WHERE id = $1::uuid",
        str(project_id),
    )
    if not existing:
        raise NotFoundError("Project", str(project_id))

    await execute("DELETE FROM projects WHERE id = $1::uuid", str(project_id))

    background_tasks.add_task(
        audit_service.log,
        user_id=current_user.id,
        user_email=current_user.email,
        action="delete",
        table_name="projects",
        record_id=str(project_id),
        old_values=existing,
        ip_address=get_client_ip(request),
        description=f"Deleted project: {existing.get('project_name')}",
    )

    broadcast("projects", "delete")
    return {"success": True, "message": "Project deleted successfully"}
