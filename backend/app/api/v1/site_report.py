"""Site Engineer reporting endpoints.

Site engineers are assigned to exactly one location. They fill weekly
plant-status reports directly in the browser (no Excel). Drafts are
DB-persisted and survive logout. Submission directly processes into
plant_weekly_records via process_direct_submission().
"""

import io
from datetime import date, datetime, timedelta
from typing import Annotated, Any
from uuid import UUID, uuid4

from fastapi import APIRouter, BackgroundTasks, Depends, HTTPException, Query, Request
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from app.core.exceptions import NotFoundError, ValidationError
from app.core.pool import fetch, fetchrow, fetchval, execute, get_pool
from app.core.security import CurrentUser, require_site_engineer
from app.monitoring.logging import get_logger
from app.services.audit_service import audit_service

router = APIRouter()
logger = get_logger(__name__)


# ============================================================================
# Pydantic models
# ============================================================================

class DraftRowUpsert(BaseModel):
    fleet_number: str
    condition: str | None = None
    physical_verification: bool | None = None
    hours_worked: float | None = None
    standby_hours: float | None = None
    breakdown_hours: float | None = None
    off_hire: bool = False
    transfer_to_location_id: UUID | None = None
    remarks: str | None = None
    is_new_plant: bool = False
    plant_description: str | None = None  # Required when is_new_plant=True


# ============================================================================
# Helpers
# ============================================================================

def _current_week_ending() -> date:
    """Return the most recent Friday (or today if today is Friday)."""
    today = date.today()
    # weekday(): Mon=0, Fri=4
    days_since_friday = (today.weekday() - 4) % 7
    return today - timedelta(days=days_since_friday)


_DRAFT_WITH_ROWS_SQL = """
    SELECT d.id::text, d.location_id::text, d.week_ending_date, d.status, d.updated_at,
           COALESCE(
               json_agg(
                   json_build_object(
                       'id', r.id::text,
                       'draft_id', r.draft_id::text,
                       'fleet_number', r.fleet_number,
                       'plant_id', r.plant_id::text,
                       'condition', r.condition,
                       'physical_verification', r.physical_verification,
                       'hours_worked', r.hours_worked,
                       'standby_hours', r.standby_hours,
                       'breakdown_hours', r.breakdown_hours,
                       'off_hire', r.off_hire,
                       'transfer_to_location_id', r.transfer_to_location_id::text,
                       'remarks', r.remarks,
                       'is_new_plant', r.is_new_plant,
                       'updated_at', r.updated_at
                   ) ORDER BY r.fleet_number
               ) FILTER (WHERE r.id IS NOT NULL),
               '[]'::json
           ) AS rows
    FROM weekly_report_drafts d
    LEFT JOIN weekly_report_draft_rows r ON r.draft_id = d.id
    WHERE {where}
    GROUP BY d.id
"""


async def _get_or_create_draft(user_id: str, location_id: str, week_ending_date: date) -> dict:
    """Get existing draft or create a new one pre-populated with site plants."""
    # Check if draft already exists (fast check before the expensive aggregation)
    existing_id = await fetchval(
        "SELECT id FROM weekly_report_drafts WHERE user_id = $1::uuid AND week_ending_date = $2",
        user_id, week_ending_date,
    )

    if not existing_id:
        # Create new draft + pre-populate rows in a single transaction
        pool = get_pool()
        async with pool.acquire() as conn:
            async with conn.transaction():
                new_draft = await conn.fetchrow(
                    """INSERT INTO weekly_report_drafts (user_id, location_id, week_ending_date)
                       VALUES ($1::uuid, $2::uuid, $3)
                       RETURNING id""",
                    user_id, location_id, week_ending_date,
                )
                existing_id = new_draft["id"]

                # Pre-populate rows from plants currently at this location
                plants = await conn.fetch(
                    """SELECT id, fleet_number FROM plants_master
                       WHERE current_location_id = $1::uuid
                         AND condition NOT IN ('scrap', 'off_hire')
                       ORDER BY fleet_number""",
                    location_id,
                )

                if plants:
                    await conn.executemany(
                        """INSERT INTO weekly_report_draft_rows (draft_id, plant_id, fleet_number)
                           VALUES ($1::uuid, $2::uuid, $3)
                           ON CONFLICT (draft_id, fleet_number) DO NOTHING""",
                        [(str(existing_id), str(p["id"]), p["fleet_number"]) for p in plants],
                    )

    # Fetch draft with rows using draft_id (PK lookup — fast)
    draft = await fetchrow(
        _DRAFT_WITH_ROWS_SQL.format(where="d.id = $1::uuid"),
        str(existing_id),
    )
    result = dict(draft) if draft else {}
    result["rows"] = result.get("rows") or []
    return result


# ============================================================================
# Endpoints
# ============================================================================

@router.get("/me")
async def get_my_site(
    current_user: Annotated[CurrentUser, Depends(require_site_engineer)],
) -> dict[str, Any]:
    """Return site stats for the logged-in engineer's assigned location."""
    row = await fetchrow(
        """SELECT
               l.id, l.name AS location_name, s.name AS state_name,
               COALESCE(p.total_plants, 0)::int AS total_plants,
               COALESCE(p.working, 0)::int AS working,
               COALESCE(p.standby, 0)::int AS standby,
               COALESCE(p.breakdown, 0)::int AS breakdown,
               COALESCE(p.missing, 0)::int AS missing,
               COALESCE(p.faulty, 0)::int AS faulty,
               COALESCE(p.scrap, 0)::int AS scrap,
               COALESCE(p.off_hire, 0)::int AS off_hire,
               COALESCE(p.unverified, 0)::int AS unverified,
               ls.last_submission
           FROM locations l
           LEFT JOIN states s ON s.id = l.state_id
           LEFT JOIN LATERAL (
               SELECT
                   count(*)::int AS total_plants,
                   count(*) FILTER (WHERE condition = 'working')::int AS working,
                   count(*) FILTER (WHERE condition = 'standby')::int AS standby,
                   count(*) FILTER (WHERE condition = 'breakdown')::int AS breakdown,
                   count(*) FILTER (WHERE condition = 'missing')::int AS missing,
                   count(*) FILTER (WHERE condition = 'faulty')::int AS faulty,
                   count(*) FILTER (WHERE condition = 'scrap')::int AS scrap,
                   count(*) FILTER (WHERE condition = 'off_hire')::int AS off_hire,
                   count(*) FILTER (WHERE condition = 'unverified')::int AS unverified
               FROM plants_master pm
               WHERE pm.current_location_id = l.id
           ) p ON TRUE
           LEFT JOIN LATERAL (
               SELECT MAX(week_ending_date) AS last_submission
               FROM weekly_report_submissions
               WHERE location_id = l.id AND status = 'completed'
           ) ls ON TRUE
           WHERE l.id = $1::uuid""",
        current_user.location_id,
    )
    if not row:
        raise NotFoundError("Location", current_user.location_id)
    return {"success": True, "data": row}


@router.get("/plants")
async def get_my_site_plants(
    current_user: Annotated[CurrentUser, Depends(require_site_engineer)],
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=200),
    search: str | None = None,
    condition: str | None = None,
) -> dict[str, Any]:
    """List plants at the engineer's assigned location."""
    conds = ["pm.current_location_id = $1::uuid"]
    params: list[Any] = [current_user.location_id]

    if search:
        params.append(f"%{search}%")
        conds.append(f"(pm.fleet_number ILIKE ${len(params)} OR pm.description ILIKE ${len(params)})")
    if condition:
        params.append(condition)
        conds.append(f"pm.condition = ${len(params)}")

    where = " AND ".join(conds)
    offset = (page - 1) * limit
    params.extend([limit, offset])

    rows = await fetch(
        f"""SELECT pm.id, pm.fleet_number, pm.description, pm.fleet_type,
                   pm.make, pm.model, pm.condition, pm.physical_verification,
                   pm.last_verified_date,
                   count(*) OVER() AS _total_count
            FROM plants_master pm
            WHERE {where}
            ORDER BY pm.fleet_number
            LIMIT ${len(params) - 1} OFFSET ${len(params)}""",
        *params,
    )

    total = rows[0].pop("_total_count", 0) if rows else 0
    for r in rows[1:]:
        r.pop("_total_count", None)

    return {
        "success": True,
        "data": rows,
        "meta": {
            "page": page, "limit": limit, "total": total,
            "total_pages": (total + limit - 1) // limit if total else 0,
        },
    }


@router.get("/draft")
async def get_draft(
    current_user: Annotated[CurrentUser, Depends(require_site_engineer)],
    week_ending_date: date = Query(default=None),
) -> dict[str, Any]:
    """Get (or create) the draft for the given week. Defaults to current week's Friday."""
    if week_ending_date is None:
        week_ending_date = _current_week_ending()
    draft = await _get_or_create_draft(
        current_user.id, current_user.location_id, week_ending_date
    )
    return {"success": True, "data": draft}


@router.put("/draft/rows")
async def upsert_draft_row(
    current_user: Annotated[CurrentUser, Depends(require_site_engineer)],
    body: DraftRowUpsert,
    week_ending_date: date = Query(default=None),
    draft_id: str | None = Query(default=None),
) -> dict[str, Any]:
    """Upsert a single plant row in the engineer's current draft.

    Accepts an optional draft_id to skip the draft lookup query.
    Creates the draft if it doesn't exist. Safe to call on every keystroke
    (after debounce on the frontend).
    """
    if week_ending_date is None:
        week_ending_date = _current_week_ending()

    if body.is_new_plant and not body.plant_description:
        raise ValidationError("plant_description is required when adding a new plant")

    # Resolve draft_id — skip lookup if provided by frontend
    if not draft_id:
        draft = await fetchrow(
            "SELECT id, status FROM weekly_report_drafts WHERE user_id = $1::uuid AND week_ending_date = $2",
            current_user.id, week_ending_date,
        )
        if not draft:
            draft = await _get_or_create_draft(current_user.id, current_user.location_id, week_ending_date)
            draft_id = draft["id"]
        else:
            if draft.get("status") == "submitted":
                raise ValidationError("This draft has already been submitted")
            draft_id = str(draft["id"])

    # Single CTE: resolve plant_id + upsert row + bump draft timestamp (1 round-trip)
    fleet = body.fleet_number.strip().upper()
    result = await fetchval(
        """
        WITH plant_lookup AS (
            SELECT id FROM plants_master WHERE fleet_number = $2
        ),
        upserted AS (
            INSERT INTO weekly_report_draft_rows
                (draft_id, plant_id, fleet_number, condition, physical_verification,
                 hours_worked, standby_hours, breakdown_hours, off_hire,
                 transfer_to_location_id, remarks, is_new_plant)
            VALUES ($1::uuid, (SELECT id FROM plant_lookup), $2, $3, $4, $5, $6, $7, $8, $9::uuid, $10, $11)
            ON CONFLICT (draft_id, fleet_number)
            DO UPDATE SET
                plant_id                = COALESCE((SELECT id FROM plant_lookup), weekly_report_draft_rows.plant_id),
                condition               = EXCLUDED.condition,
                physical_verification   = EXCLUDED.physical_verification,
                hours_worked            = EXCLUDED.hours_worked,
                standby_hours           = EXCLUDED.standby_hours,
                breakdown_hours         = EXCLUDED.breakdown_hours,
                off_hire                = EXCLUDED.off_hire,
                transfer_to_location_id = EXCLUDED.transfer_to_location_id,
                remarks                 = EXCLUDED.remarks,
                is_new_plant            = EXCLUDED.is_new_plant,
                updated_at              = now()
            RETURNING draft_id
        )
        UPDATE weekly_report_drafts SET updated_at = now()
        WHERE id = (SELECT draft_id FROM upserted LIMIT 1)
        RETURNING id::text
        """,
        draft_id, fleet,
        body.condition, body.physical_verification,
        body.hours_worked, body.standby_hours, body.breakdown_hours,
        body.off_hire,
        str(body.transfer_to_location_id) if body.transfer_to_location_id else None,
        body.remarks, body.is_new_plant,
    )

    return {"success": True, "draft_id": result or draft_id}


class BatchDraftRowUpsert(BaseModel):
    rows: list[DraftRowUpsert]


@router.put("/draft/rows/batch")
async def batch_upsert_draft_rows(
    current_user: Annotated[CurrentUser, Depends(require_site_engineer)],
    body: BatchDraftRowUpsert,
    week_ending_date: date = Query(default=None),
    draft_id: str | None = Query(default=None),
) -> dict[str, Any]:
    """Batch upsert multiple rows in a single request/transaction.

    Significantly faster than individual PUT calls when saving many rows.
    """
    if week_ending_date is None:
        week_ending_date = _current_week_ending()

    if not body.rows:
        return {"success": True, "draft_id": draft_id, "saved": 0}

    # Resolve draft_id
    if not draft_id:
        draft = await fetchrow(
            "SELECT id, status FROM weekly_report_drafts WHERE user_id = $1::uuid AND week_ending_date = $2",
            current_user.id, week_ending_date,
        )
        if not draft:
            draft = await _get_or_create_draft(current_user.id, current_user.location_id, week_ending_date)
            draft_id = draft["id"]
        else:
            if draft.get("status") == "submitted":
                raise ValidationError("This draft has already been submitted")
            draft_id = str(draft["id"])

    pool = get_pool()
    async with pool.acquire() as conn:
        async with conn.transaction():
            for row in body.rows:
                fleet = row.fleet_number.strip().upper()
                await conn.execute(
                    """
                    WITH plant_lookup AS (
                        SELECT id FROM plants_master WHERE fleet_number = $2
                    )
                    INSERT INTO weekly_report_draft_rows
                        (draft_id, plant_id, fleet_number, condition, physical_verification,
                         hours_worked, standby_hours, breakdown_hours, off_hire,
                         transfer_to_location_id, remarks, is_new_plant)
                    VALUES ($1::uuid, (SELECT id FROM plant_lookup), $2, $3, $4, $5, $6, $7, $8, $9::uuid, $10, $11)
                    ON CONFLICT (draft_id, fleet_number)
                    DO UPDATE SET
                        plant_id                = COALESCE((SELECT id FROM plant_lookup), weekly_report_draft_rows.plant_id),
                        condition               = EXCLUDED.condition,
                        physical_verification   = EXCLUDED.physical_verification,
                        hours_worked            = EXCLUDED.hours_worked,
                        standby_hours           = EXCLUDED.standby_hours,
                        breakdown_hours         = EXCLUDED.breakdown_hours,
                        off_hire                = EXCLUDED.off_hire,
                        transfer_to_location_id = EXCLUDED.transfer_to_location_id,
                        remarks                 = EXCLUDED.remarks,
                        is_new_plant            = EXCLUDED.is_new_plant,
                        updated_at              = now()
                    """,
                    draft_id, fleet,
                    row.condition, row.physical_verification,
                    row.hours_worked, row.standby_hours, row.breakdown_hours,
                    row.off_hire,
                    str(row.transfer_to_location_id) if row.transfer_to_location_id else None,
                    row.remarks, row.is_new_plant,
                )

            # Bump draft timestamp once
            await conn.execute(
                "UPDATE weekly_report_drafts SET updated_at = now() WHERE id = $1::uuid",
                draft_id,
            )

    return {"success": True, "draft_id": draft_id, "saved": len(body.rows)}


@router.delete("/draft/rows/{fleet_number}")
async def remove_draft_row(
    fleet_number: str,
    current_user: Annotated[CurrentUser, Depends(require_site_engineer)],
    week_ending_date: date = Query(default=None),
) -> dict[str, Any]:
    """Remove a plant row from the draft (e.g. to exclude a missing plant)."""
    if week_ending_date is None:
        week_ending_date = _current_week_ending()

    draft = await fetchrow(
        "SELECT id FROM weekly_report_drafts WHERE user_id = $1::uuid AND week_ending_date = $2",
        current_user.id, week_ending_date,
    )
    if not draft:
        raise NotFoundError("Draft", "current week")

    await execute(
        "DELETE FROM weekly_report_draft_rows WHERE draft_id = $1::uuid AND fleet_number = $2",
        str(draft["id"]), fleet_number.strip().upper(),
    )
    return {"success": True}


@router.post("/draft/submit")
async def submit_draft(
    current_user: Annotated[CurrentUser, Depends(require_site_engineer)],
    background_tasks: BackgroundTasks,
    request: Request,
    week_ending_date: date = Query(default=None),
) -> dict[str, Any]:
    """Submit the draft. Processes directly into DB (no Excel file).

    Creates a weekly_report_submissions record and updates plants_master,
    plant_weekly_records, and plant_location_history. Marks draft as submitted.
    """
    from app.workers.etl_worker import process_direct_submission
    from app.api.v1.auth import get_client_ip

    if week_ending_date is None:
        week_ending_date = _current_week_ending()

    draft = await fetchrow(
        """SELECT d.id, d.location_id, d.status
           FROM weekly_report_drafts d
           WHERE d.user_id = $1::uuid AND d.week_ending_date = $2""",
        current_user.id, week_ending_date,
    )
    if not draft:
        raise NotFoundError("Draft", "current week")
    if draft["status"] == "submitted":
        raise ValidationError("This draft has already been submitted")

    # Count rows
    row_count = await fetchval(
        "SELECT count(*) FROM weekly_report_draft_rows WHERE draft_id = $1::uuid",
        str(draft["id"]),
    )
    if not row_count:
        raise ValidationError("Draft has no plant rows to submit")

    # Compute ISO week
    iso = week_ending_date.isocalendar()
    year = week_ending_date.year
    week_number = iso[1]

    # Check for duplicate submission
    existing_sub = await fetchrow(
        """SELECT id FROM weekly_report_submissions
           WHERE location_id = $1::uuid AND year = $2 AND week_number = $3
             AND status = 'completed' AND source_type = 'site_engineer'""",
        str(draft["location_id"]), year, week_number,
    )
    if existing_sub:
        raise ValidationError(
            f"A report for week {week_number}/{year} has already been submitted for this site"
        )

    # Create submission record
    submission_id = str(uuid4())
    location_id = str(draft["location_id"])

    # Get engineer's name/email for submitted_by_name/email fields
    engineer = await fetchrow(
        "SELECT full_name, email FROM users WHERE id = $1::uuid", current_user.id
    )

    await execute(
        """INSERT INTO weekly_report_submissions
               (id, source_file_name, location_id, year, week_number, week_ending_date,
                status, submitted_by_name, submitted_by_email, source_type)
           VALUES ($1::uuid, $2, $3::uuid, $4, $5, $6, 'processing', $7, $8, 'site_engineer')""",
        submission_id,
        f"direct-submission-week{week_number}-{year}",
        location_id, year, week_number, week_ending_date,
        engineer["full_name"] if engineer else current_user.id,
        engineer["email"] if engineer else current_user.email,
    )

    try:
        result = await process_direct_submission(
            draft_id=str(draft["id"]),
            submission_id=submission_id,
            location_id=location_id,
            week_ending_date=week_ending_date,
            year=year,
            week_number=week_number,
            submitted_by=current_user.id,
        )
    except Exception as e:
        # Mark submission as failed
        await execute(
            "UPDATE weekly_report_submissions SET status = 'failed', errors = $1::jsonb WHERE id = $2::uuid",
            f'[{{"error": "{str(e)}"}}]', submission_id,
        )
        logger.error("Direct submission failed", draft_id=str(draft["id"]), error=str(e))
        raise ValidationError(f"Submission failed: {e}")

    background_tasks.add_task(
        audit_service.log,
        user_id=current_user.id,
        user_email=current_user.email,
        action="create",
        table_name="weekly_report_submissions",
        record_id=submission_id,
        new_values={
            "week_ending_date": str(week_ending_date),
            "plants_processed": result.get("plants_processed", 0),
            "source": "site_engineer",
        },
        description=f"Site engineer submitted weekly report for week {week_number}/{year}",
    )

    return {
        "success": True,
        "data": {
            "submission_id": submission_id,
            "plants_processed": result.get("plants_processed", 0),
            "plants_created": result.get("plants_created", 0),
            "transfers_pending": result.get("transfers_pending", 0),
        },
    }


@router.get("/submissions")
async def get_my_submissions(
    current_user: Annotated[CurrentUser, Depends(require_site_engineer)],
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=50),
) -> dict[str, Any]:
    """List past submissions for this location."""
    offset = (page - 1) * limit
    rows = await fetch(
        """SELECT s.id, s.year, s.week_number, s.week_ending_date,
                  s.status, s.plants_processed, s.plants_created, s.plants_updated,
                  s.source_type, s.created_at,
                  count(*) OVER() AS _total_count
           FROM weekly_report_submissions s
           WHERE s.location_id = $1::uuid
           ORDER BY s.week_ending_date DESC, s.created_at DESC
           LIMIT $2 OFFSET $3""",
        current_user.location_id, limit, offset,
    )
    total = rows[0].pop("_total_count", 0) if rows else 0
    for r in rows[1:]:
        r.pop("_total_count", None)
    return {
        "success": True,
        "data": rows,
        "meta": {
            "page": page, "limit": limit, "total": total,
            "total_pages": (total + limit - 1) // limit if total else 0,
        },
    }


@router.get("/submissions/{submission_id}/records")
async def get_submission_records(
    submission_id: str,
    current_user: Annotated[CurrentUser, Depends(require_site_engineer)],
) -> dict[str, Any]:
    """Get individual plant records for a completed submission."""
    sub = await fetchrow(
        "SELECT id, location_id FROM weekly_report_submissions WHERE id = $1::uuid",
        submission_id,
    )
    if not sub or str(sub["location_id"]) != current_user.location_id:
        raise NotFoundError("Submission not found")

    records = await fetch(
        """SELECT pm.fleet_number, pm.description, pm.fleet_type,
                  wr.condition, wr.physical_verification,
                  wr.hours_worked, wr.standby_hours, wr.breakdown_hours,
                  wr.off_hire, wr.remarks, wr.transfer_to
           FROM plant_weekly_records wr
           JOIN plants_master pm ON pm.id = wr.plant_id
           WHERE wr.submission_id = $1::uuid
           ORDER BY pm.fleet_number""",
        submission_id,
    )
    return {
        "success": True,
        "data": [
            {
                "fleet_number": r["fleet_number"],
                "description": r["description"],
                "fleet_type": r["fleet_type"],
                "condition": r["condition"],
                "physical_verification": r["physical_verification"],
                "hours_worked": float(r["hours_worked"]) if r["hours_worked"] is not None else None,
                "standby_hours": float(r["standby_hours"]) if r["standby_hours"] is not None else None,
                "breakdown_hours": float(r["breakdown_hours"]) if r["breakdown_hours"] is not None else None,
                "off_hire": r["off_hire"],
                "remarks": r["remarks"],
                "transfer_to": r["transfer_to"],
            }
            for r in records
        ],
    }


@router.get("/submissions/{submission_id}/export")
async def export_submission(
    submission_id: UUID,
    current_user: Annotated[CurrentUser, Depends(require_site_engineer)],
) -> StreamingResponse:
    """Export a past submission.

    For upload-portal submissions (source_file_path set): streams the original
    Excel file from Supabase Storage so the engineer gets back exactly what was
    uploaded rather than a system-generated file.

    For direct form submissions (no file path): generates a branded Excel on
    the fly from the stored plant_weekly_records.
    """
    submission = await fetchrow(
        "SELECT * FROM weekly_report_submissions WHERE id = $1::uuid AND location_id = $2::uuid",
        str(submission_id), current_user.location_id,
    )
    if not submission:
        raise NotFoundError("Submission", str(submission_id))

    file_path = submission.get("source_file_path")
    if file_path:
        # Upload-portal submission — stream the original file from Storage
        from app.core.database import get_supabase_admin_client
        try:
            storage_client = get_supabase_admin_client()
            file_data = storage_client.storage.from_("reports").download(file_path)
            file_name = submission.get("source_file_name") or "weekly-report.xlsx"
            return StreamingResponse(
                io.BytesIO(file_data),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f'attachment; filename="{file_name}"'},
            )
        except Exception as e:
            logger.error("Failed to download submission file from Storage", error=str(e))
            raise ValidationError(f"Could not retrieve original file: {str(e)}")

    # Site-engineer form submission — generate branded Excel
    return await _build_submission_excel(submission)


@router.get("/transfers/incoming")
async def get_incoming_transfers(
    current_user: Annotated[CurrentUser, Depends(require_site_engineer)],
) -> dict[str, Any]:
    """List pending incoming transfers for this location."""
    rows = await fetch(
        """SELECT
               t.id, t.status, t.transfer_date, t.created_at,
               t.source_remarks AS notes,
               pm.fleet_number, pm.description, pm.fleet_type,
               fl.name AS from_location_name
           FROM plant_transfers t
           JOIN plants_master pm ON pm.id = t.plant_id
           JOIN locations fl ON fl.id = t.from_location_id
           WHERE t.to_location_id = $1::uuid AND t.status = 'pending'
             AND t.is_pull_request = FALSE
           ORDER BY t.created_at DESC""",
        current_user.location_id,
    )
    return {"success": True, "data": rows}


class PullRequestBody(BaseModel):
    fleet_number: str


@router.post("/transfers/pull-request")
async def create_pull_request_transfer(
    body: PullRequestBody,
    current_user: Annotated[CurrentUser, Depends(require_site_engineer)],
) -> dict[str, Any]:
    """Request transfer of a plant that is currently at another site.

    Creates a plant_transfers record with is_pull_request=TRUE. The source-site
    engineer will see it in their GET /site/transfers/pull-requests list and
    can approve or reject it.
    """
    fn = body.fleet_number.strip().upper()
    plant = await fetchrow(
        """SELECT id, fleet_number, current_location_id
           FROM plants_master WHERE fleet_number = $1""",
        fn,
    )
    if not plant:
        raise NotFoundError("Plant", fn)

    if str(plant["current_location_id"]) == current_user.location_id:
        raise ValidationError("Plant is already registered at your site")

    # Avoid duplicate pending pull requests for the same plant
    existing = await fetchval(
        """SELECT id FROM plant_transfers
           WHERE plant_id = $1::uuid AND to_location_id = $2::uuid
             AND status = 'pending' AND is_pull_request = TRUE""",
        str(plant["id"]), current_user.location_id,
    )
    if existing:
        return {"success": True, "message": f"A transfer request for {fn} is already pending — waiting for the other site to approve"}

    await execute(
        """INSERT INTO plant_transfers
               (plant_id, from_location_id, to_location_id, status,
                direction, is_pull_request, detected_date)
           VALUES ($1::uuid, $2::uuid, $3::uuid, 'pending',
                   'outbound', TRUE, CURRENT_DATE)""",
        str(plant["id"]),
        str(plant["current_location_id"]),
        current_user.location_id,
    )
    return {"success": True, "message": f"Transfer request sent for {fn} — the other site will be notified to approve"}


@router.get("/transfers/pull-requests")
async def get_pull_requests(
    current_user: Annotated[CurrentUser, Depends(require_site_engineer)],
) -> dict[str, Any]:
    """List pending pull requests — plants that other sites are requesting FROM this site."""
    rows = await fetch(
        """SELECT
               t.id, t.status, t.created_at,
               pm.fleet_number, pm.description, pm.fleet_type,
               tl.name AS requesting_location_name
           FROM plant_transfers t
           JOIN plants_master pm ON pm.id = t.plant_id
           JOIN locations tl ON tl.id = t.to_location_id
           WHERE t.from_location_id = $1::uuid
             AND t.status = 'pending' AND t.is_pull_request = TRUE
           ORDER BY t.created_at DESC""",
        current_user.location_id,
    )
    return {"success": True, "data": rows}


@router.post("/transfers/{transfer_id}/confirm")
async def confirm_incoming_transfer(
    transfer_id: UUID,
    current_user: Annotated[CurrentUser, Depends(require_site_engineer)],
) -> dict[str, Any]:
    """Confirm a transfer — moves plant between locations.

    Handles two cases:
    - Regular incoming transfer: engineer at to_location_id confirms.
    - Pull request: engineer at from_location_id approves (releases the plant).
    In both cases the plant moves from from_location_id to to_location_id.
    """
    pool = get_pool()
    async with pool.acquire() as conn:
        async with conn.transaction():
            transfer = await conn.fetchrow(
                """SELECT * FROM plant_transfers
                   WHERE id = $1::uuid AND status = 'pending'
                     AND (
                       (to_location_id = $2::uuid AND is_pull_request = FALSE)
                       OR (from_location_id = $2::uuid AND is_pull_request = TRUE)
                     )""",
                str(transfer_id), current_user.location_id,
            )
            if not transfer:
                raise NotFoundError("Transfer", str(transfer_id))

            plant_id = str(transfer["plant_id"])
            from_loc = str(transfer["from_location_id"])
            to_loc = str(transfer["to_location_id"])  # Always use the DB value (correct for both regular + pull requests)

            # Confirm the transfer
            await conn.execute(
                """UPDATE plant_transfers
                   SET status = 'confirmed', confirmed_at = now(),
                       actual_arrival_date = CURRENT_DATE
                   WHERE id = $1::uuid""",
                str(transfer_id),
            )

            # Close old location history
            await conn.execute(
                """UPDATE plant_location_history SET end_date = now()
                   WHERE plant_id = $1::uuid AND location_id = $2::uuid AND end_date IS NULL""",
                plant_id, from_loc,
            )

            # Open new location history
            await conn.execute(
                """INSERT INTO plant_location_history (plant_id, location_id, start_date)
                   VALUES ($1::uuid, $2::uuid, CURRENT_DATE)""",
                plant_id, to_loc,
            )

            # Update plant's current location
            await conn.execute(
                "UPDATE plants_master SET current_location_id = $1::uuid, updated_at = now() WHERE id = $2::uuid",
                to_loc, plant_id,
            )

            fleet_number = await conn.fetchval(
                "SELECT fleet_number FROM plants_master WHERE id = $1::uuid", plant_id
            )

    is_pull = transfer.get("is_pull_request", False)
    msg = (
        f"Transfer approved — {fleet_number} released to requesting site"
        if is_pull
        else f"Transfer confirmed — {fleet_number} is now at your site"
    )
    return {"success": True, "message": msg}


@router.post("/transfers/{transfer_id}/reject")
async def reject_incoming_transfer(
    transfer_id: UUID,
    current_user: Annotated[CurrentUser, Depends(require_site_engineer)],
) -> dict[str, Any]:
    """Reject a transfer (incoming or pull request)."""
    transfer = await fetchrow(
        """SELECT id FROM plant_transfers
           WHERE id = $1::uuid AND status = 'pending'
             AND (
               (to_location_id = $2::uuid AND is_pull_request = FALSE)
               OR (from_location_id = $2::uuid AND is_pull_request = TRUE)
             )""",
        str(transfer_id), current_user.location_id,
    )
    if not transfer:
        raise NotFoundError("Transfer", str(transfer_id))

    await execute(
        "UPDATE plant_transfers SET status = 'rejected' WHERE id = $1::uuid",
        str(transfer_id),
    )
    return {"success": True, "message": "Transfer rejected"}


@router.get("/locations")
async def list_locations_for_transfer(
    current_user: Annotated[CurrentUser, Depends(require_site_engineer)],
) -> dict[str, Any]:
    """List all locations (for transfer-to dropdown). Excludes the engineer's own site."""
    rows = await fetch(
        """SELECT id::text, name FROM locations
           WHERE id != $1::uuid
           ORDER BY name""",
        current_user.location_id,
    )
    return {"success": True, "data": [{"id": r["id"], "name": r["name"]} for r in rows]}


@router.get("/new-plant-check/{fleet_number}")
async def check_new_plant(
    fleet_number: str,
    current_user: Annotated[CurrentUser, Depends(require_site_engineer)],
) -> dict[str, Any]:
    """Check if a fleet number is available to add.

    Returns available=True if:
    - Plant does not exist in the DB (brand new plant)
    - Plant exists but is already at this engineer's site (re-add)

    Returns available=False if:
    - Plant is registered at a different site (transfer must happen first)
    """
    fn_upper = fleet_number.strip().upper()
    existing = await fetchrow(
        """SELECT pm.id, pm.current_location_id, l.name AS current_location
           FROM plants_master pm
           LEFT JOIN locations l ON l.id = pm.current_location_id
           WHERE pm.fleet_number = $1""",
        fn_upper,
    )
    if not existing:
        return {"success": True, "data": {"available": True}}

    # Plant is at this engineer's own site — OK to re-add to report
    existing_loc_id = str(existing["current_location_id"]) if existing.get("current_location_id") else None
    if existing_loc_id == current_user.location_id:
        return {
            "success": True,
            "data": {
                "available": True,
                "message": f"{fn_upper} is registered at your site — you can re-add it to this report",
            },
        }

    # Plant is at another site
    loc = existing.get("current_location") or "another site"
    return {
        "success": True,
        "data": {
            "available": False,
            "current_location": existing.get("current_location"),
            "message": f"{fn_upper} is currently registered at {loc}. You can send a transfer request — they can approve to release the plant to you.",
        },
    }


# ============================================================================
# Excel export helper
# ============================================================================

async def _build_submission_excel(submission: dict) -> StreamingResponse:
    """Build an Excel file for a completed submission."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise ValidationError("openpyxl is required for Excel export")

    submission_id = str(submission["id"])
    location_name = await fetchval(
        "SELECT name FROM locations WHERE id = $1::uuid", str(submission["location_id"])
    )

    records = await fetch(
        """SELECT pm.fleet_number, pm.description, pm.fleet_type,
                  wr.physical_verification, wr.condition,
                  wr.hours_worked, wr.standby_hours, wr.breakdown_hours,
                  wr.off_hire, wr.remarks, wr.transfer_to
           FROM plant_weekly_records wr
           JOIN plants_master pm ON pm.id = wr.plant_id
           WHERE wr.submission_id = $1::uuid
           ORDER BY pm.fleet_number""",
        submission_id,
    )

    # Brand colours
    BRAND_GOLD = "FFBF36"       # P.W. Nigeria brand gold
    BRAND_GOLD_DARK = "101415"  # dark text on gold
    HEADER_FONT_COLOR = BRAND_GOLD_DARK

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Week {submission['week_number']}"

    # Column header row style — gold fill, dark text
    col_header_font = Font(bold=True, color=HEADER_FONT_COLOR, size=10)
    col_header_fill = PatternFill("solid", fgColor=BRAND_GOLD)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    center_mid = Alignment(horizontal="center", vertical="center")

    headers = [
        "S/N", "Fleet No.", "Description", "Fleet Type",
        "Physical Verification", "Condition",
        "Hrs Worked", "Standby Hrs", "Breakdown Hrs",
        "Off Hire", "Transfer To", "Remarks",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = col_header_font
        cell.fill = col_header_fill
        cell.alignment = center_mid

    # Alternating row fill for data rows
    alt_fill = PatternFill("solid", fgColor="FFF8E7")  # very light gold tint

    for i, r in enumerate(records, 1):
        ws.append([
            i,
            r["fleet_number"],
            r["description"] or "",
            r["fleet_type"] or "",
            "Yes" if r["physical_verification"] else "No",
            (r["condition"] or "").replace("_", " ").title(),
            float(r["hours_worked"] or 0),
            float(r["standby_hours"] or 0),
            float(r["breakdown_hours"] or 0),
            "Yes" if r["off_hire"] else "No",
            r["transfer_to"] or "",
            r["remarks"] or "",
        ])
        if i % 2 == 0:
            for cell in ws[ws.max_row]:
                cell.fill = alt_fill

    # Column widths
    col_widths = [6, 14, 35, 20, 18, 14, 12, 12, 16, 10, 20, 30]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    # Title rows (inserted above the column headers)
    num_cols = len(headers)
    last_col_letter = openpyxl.utils.get_column_letter(num_cols)

    ws.insert_rows(1)  # row 2 = week info
    ws.insert_rows(1)  # row 1 = company header
    ws.insert_rows(1)  # row 1 = company name (new top row)

    # Row 1: Company name — gold background
    ws["A1"] = "P.W. Nigeria Ltd."
    ws["A1"].font = Font(bold=True, size=14, color=HEADER_FONT_COLOR)
    ws["A1"].fill = PatternFill("solid", fgColor=BRAND_GOLD)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(f"A1:{last_col_letter}1")
    ws.row_dimensions[1].height = 22

    # Row 2: Report title — site name
    ws["A2"] = f"Weekly Plant Report  —  {location_name}"
    ws["A2"].font = Font(bold=True, size=12)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(f"A2:{last_col_letter}2")
    ws.row_dimensions[2].height = 18

    # Row 3: Week info
    ws["A3"] = f"Week Ending: {submission['week_ending_date']}     |     Week {submission['week_number']} / {submission['year']}"
    ws["A3"].font = Font(size=10, italic=True)
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(f"A3:{last_col_letter}3")
    ws.row_dimensions[3].height = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"weekly-report-{(location_name or 'site').lower().replace(' ', '-')}-week{submission['week_number']}-{submission['year']}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
