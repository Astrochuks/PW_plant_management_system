"""Anchor-based parsing toolkit for project weekly-report workbooks (T2.2/T2.3).

PRD v2 §8.1: parsers find labels and header rows by TEXT, never by fixed
cell addresses — the same template drifts by a few rows/columns between
sites, and fixed addresses are how the old plant parsers rotted.

Everything here is pure and deterministic: openpyxl worksheets in,
plain Python values out. Nothing raises on malformed input — missing
anchors return None and the caller decides (usually: mark the sheet
partial and report specifics).
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date, datetime
from typing import Any, Iterator

from openpyxl.worksheet.worksheet import Worksheet


# ---------------------------------------------------------------------------
# Normalization
# ---------------------------------------------------------------------------

def norm(text: Any) -> str:
    """Header/label matching key: lowercase, collapsed spaces, no punctuation
    edges. 'Fleet No.' == 'FLEET NO' == ' fleet  no '."""
    s = re.sub(r"\s+", " ", str(text or "")).strip().strip(".:;")
    return s.lower()


def cell_number(value: Any) -> float | None:
    """Numeric cell → float; '#VALUE!'/text/None → None. Never raises."""
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if not s or s.startswith("#"):  # #VALUE!, #REF!, #DIV/0! …
        return None
    try:
        return float(s.replace(",", ""))
    except ValueError:
        return None


def cell_date(value: Any) -> date | None:
    """Date-ish cell → date; anything else → None. Never raises."""
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def cell_text(value: Any, max_len: int = 500) -> str | None:
    if value is None:
        return None
    s = re.sub(r"\s+", " ", str(value)).strip()
    if not s or s.startswith("#"):
        return None
    return s[:max_len]


# ---------------------------------------------------------------------------
# Anchors
# ---------------------------------------------------------------------------

def find_label_cell(
    ws: Worksheet,
    label_pattern: str,
    *,
    max_row: int = 60,
    max_col: int = 20,
) -> tuple[int, int] | None:
    """(row, col) of the first cell whose text matches label_pattern
    (regex, case-insensitive, matched against the normalized text)."""
    rx = re.compile(label_pattern, re.IGNORECASE)
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        for cell in row:
            if cell.value is not None and rx.search(norm(cell.value)):
                return (cell.row, cell.column)
    return None


def find_label_value(
    ws: Worksheet,
    label_pattern: str,
    *,
    max_row: int = 60,
    max_col: int = 20,
    scan_right: int = 6,
) -> Any:
    """Value of the first non-empty cell to the RIGHT of a label.

    'Name of Contract: | <blank> | CONSTRUCTION OF …' → the construction
    text, wherever the label sits.
    """
    pos = find_label_cell(ws, label_pattern, max_row=max_row, max_col=max_col)
    if pos is None:
        return None
    r, c = pos
    for offset in range(1, scan_right + 1):
        val = ws.cell(row=r, column=c + offset).value
        if val is not None and str(val).strip() != "":
            return val
    return None


def find_header_row(
    ws: Worksheet,
    required_headers: list[str],
    *,
    max_row: int = 15,
    max_col: int = 40,
    min_matches: int | None = None,
) -> tuple[int, dict[str, int]] | None:
    """Locate the table header row by its column TEXTS.

    Returns (row_number, {normalized_header: column}) for the first row
    matching at least `min_matches` (default: all) of required_headers.
    Matching is startswith-tolerant: 'Total Fuel Taken' matches a
    required 'total fuel'.
    """
    need = [norm(h) for h in required_headers]
    threshold = len(need) if min_matches is None else min_matches

    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        found: dict[str, int] = {}
        for cell in row:
            if cell.value is None:
                continue
            text = norm(cell.value)
            if not text:
                continue
            for h in need:
                if h not in found and (text == h or text.startswith(h)):
                    found[h] = cell.column
        if len(found) >= threshold:
            # Also map every other non-empty header on the row
            full_map = dict(found)
            for cell in row:
                if cell.value is not None:
                    t = norm(cell.value)
                    if t and t not in full_map:
                        full_map[t] = cell.column
            return (row[0].row, full_map)
    return None


def iter_table_rows(
    ws: Worksheet,
    header_row: int,
    columns: dict[str, int],
    *,
    stop_after_blank: int = 10,
    max_rows: int = 5000,
) -> Iterator[dict[str, Any]]:
    """Yield one dict per data row under the header, keyed by the
    normalized header names in `columns`. Stops after `stop_after_blank`
    consecutive fully-blank rows (sheets have trailing junk)."""
    blanks = 0
    for r in range(header_row + 1, header_row + 1 + max_rows):
        values = {name: ws.cell(row=r, column=col).value for name, col in columns.items()}
        if all(v is None or str(v).strip() == "" for v in values.values()):
            blanks += 1
            if blanks >= stop_after_blank:
                return
            continue
        blanks = 0
        values["_row"] = r
        yield values


# ---------------------------------------------------------------------------
# Manifest / drift detection (T2.3)
# ---------------------------------------------------------------------------

@dataclass(frozen=True)
class SheetSpec:
    canonical: str
    aliases: tuple[str, ...] = ()
    required_headers: tuple[str, ...] = ()   # header-table sheets
    required_labels: tuple[str, ...] = ()    # label/value sheets
    min_header_matches: int | None = None


#: The 16-sheet Akwa Ibom template (PRD §10 — pinned for v1; a second
#: site's deviations surface as drift, never as silent mis-parses).
WORKBOOK_MANIFEST: tuple[SheetSpec, ...] = (
    SheetSpec("Weekly Summary", required_labels=("weekly work summary",)),
    SheetSpec("Contract Summary",
              required_labels=("name of contract", "original contract amount")),
    SheetSpec("BEME & Works Completed Fd", aliases=("BEME & Works Completed",),
              required_headers=("item", "description", "rate", "this week qty"),
              min_header_matches=3),
    SheetSpec("Certificate Status",
              required_headers=("cert number", "gross value of works don"),
              min_header_matches=1),
    SheetSpec("Payments Recieved", aliases=("Payments Received",),
              required_headers=("voucher number", "payment type", "gross amount"),
              min_header_matches=2),
    SheetSpec("Cost Report",
              required_headers=("description", "cost category", "amount"),
              min_header_matches=2),
    SheetSpec("Diesel Consumption",
              required_headers=("fleet no", "description", "total fuel"),
              min_header_matches=2),
    SheetSpec("Plant Return",
              required_headers=("fleet no", "description", "hours worked")),
    SheetSpec("Hired Vehicles",
              required_headers=("description", "days worked", "rate"),
              min_header_matches=2),
    SheetSpec("Labour Strength",
              required_headers=("department", "manning this week")),
    SheetSpec("Subcontractors",
              required_headers=("subcontractor", "description", "agreed rate"),
              min_header_matches=2),
    SheetSpec("Precast",
              required_headers=("description", "cast this week"),
              min_header_matches=1),
    SheetSpec("Materials & Civils",
              required_headers=("description", "opening stock", "closing stock"),
              min_header_matches=2),
    SheetSpec("Bill 1 Summary",
              required_headers=("description",), min_header_matches=1),
    SheetSpec("Bill 1 Payments",
              required_headers=("payee", "amount"), min_header_matches=2),
    SheetSpec("Lists", required_headers=("date", "week no"), min_header_matches=1),
)


def resolve_sheet(wb, spec: SheetSpec) -> Worksheet | None:
    """Find a sheet by canonical name or alias (normalized comparison)."""
    wanted = {norm(spec.canonical)} | {norm(a) for a in spec.aliases}
    for name in wb.sheetnames:
        if norm(name) in wanted:
            return wb[name]
    return None


@dataclass
class DriftReport:
    ok: dict[str, str] = field(default_factory=dict)       # canonical → actual name
    missing: list[str] = field(default_factory=list)
    drifted: dict[str, str] = field(default_factory=dict)  # canonical → what's wrong

    @property
    def clean(self) -> bool:
        return not self.missing and not self.drifted


def check_workbook(wb) -> DriftReport:
    """Verify every manifest sheet exists and its anchors are findable."""
    report = DriftReport()
    for spec in WORKBOOK_MANIFEST:
        ws = resolve_sheet(wb, spec)
        if ws is None:
            report.missing.append(spec.canonical)
            continue

        problems: list[str] = []
        for label in spec.required_labels:
            if find_label_cell(ws, re.escape(label)) is None:
                problems.append(f"label {label!r} not found")
        if spec.required_headers:
            hit = find_header_row(
                ws, list(spec.required_headers),
                min_matches=spec.min_header_matches,
            )
            if hit is None:
                problems.append(
                    f"header row with {list(spec.required_headers)} not found"
                )
        if problems:
            report.drifted[spec.canonical] = "; ".join(problems)
        else:
            report.ok[spec.canonical] = ws.title
    return report
