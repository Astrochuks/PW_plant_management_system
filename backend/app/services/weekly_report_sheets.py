"""Sheet parsers for project weekly-report workbooks (T2.4–T2.13).

One pure function per sheet: worksheet in → row dicts out (keys aligned
with the target table columns) + warnings. Everything anchor-based
(weekly_report_parsing) — no fixed cell addresses. Cumulative workbook
columns (Previous / To Date) are deliberately ignored per the PRD rule:
THIS WEEK is the only stored fact.

parse_workbook() orchestrates: manifest check first, then each sheet in
isolation — one broken sheet marks itself 'failed' with specifics and
never sinks the others.
"""

from __future__ import annotations

import re
from typing import Any

from openpyxl.workbook.workbook import Workbook

from app.services.weekly_report_parsing import (
    WORKBOOK_MANIFEST,
    cell_date,
    cell_number,
    cell_text,
    check_workbook,
    find_header_row,
    find_label_value,
    iter_table_rows,
    norm,
    resolve_sheet,
)

_NOISE_ROWS = re.compile(
    r"^(sub-?total|total|grand total|instructions?)\b", re.IGNORECASE
)


def _num(row: dict, *keys: str) -> float | None:
    """First parseable number among the row's keys (tolerant lookup)."""
    for key in keys:
        for k, v in row.items():
            if k != "_row" and k.startswith(key):
                n = cell_number(v)
                if n is not None:
                    return n
    return None


def _txt(row: dict, *keys: str) -> str | None:
    for key in keys:
        for k, v in row.items():
            if k != "_row" and k.startswith(key):
                t = cell_text(v)
                if t is not None:
                    return t
    return None


def _dt(row: dict, *keys: str):
    for key in keys:
        for k, v in row.items():
            if k != "_row" and k.startswith(key):
                d = cell_date(v)
                if d is not None:
                    return d
    return None


# ---------------------------------------------------------------------------
# T2.4 — Contract Summary → identity + snapshot
# ---------------------------------------------------------------------------

def parse_contract_summary(ws) -> dict[str, Any]:
    warnings: list[str] = []

    identity = {
        "client_raw": cell_text(find_label_value(ws, r"^client$")),
        "project_name": cell_text(find_label_value(ws, r"^name of contract")),
        "short_name": cell_text(find_label_value(ws, r"^short name")),
        "original_contract_amount": cell_number(
            find_label_value(ws, r"^original contract amount")
        ),
        "current_contract_amount": cell_number(
            find_label_value(ws, r"^current contract amount")
        ),
        "award_date": cell_date(find_label_value(ws, r"date of contract award")),
        "commencement_date": cell_date(
            find_label_value(ws, r"contract commencement date")
        ),
        "original_duration_months": cell_number(
            find_label_value(ws, r"original contract duration")
        ),
        "original_completion_date": cell_date(
            find_label_value(ws, r"original contract completion")
        ),
        "extension_of_time_months": cell_number(
            find_label_value(ws, r"extension of time granted")
        ),
        "revised_completion_date": cell_date(
            find_label_value(ws, r"^revised completion date")
        ),
    }
    for key in ("project_name", "original_contract_amount"):
        if identity[key] is None:
            warnings.append(f"Contract Summary: {key} not found")

    snapshot = {
        "original_contract_amount": identity["original_contract_amount"],
        "current_contract_amount": identity["current_contract_amount"],
        "client_name": identity["client_raw"],
        "contract_name": identity["project_name"],
        "short_name": identity["short_name"],
        "award_date": identity["award_date"],
        "commencement_date": identity["commencement_date"],
        "original_duration_months": identity["original_duration_months"],
        "eot_requested_months": cell_number(
            find_label_value(ws, r"extension of time requested")
        ),
        "eot_granted_months": identity["extension_of_time_months"],
        "revised_duration_months": cell_number(
            find_label_value(ws, r"^revised contract duration")
        ),
        "overdue_weeks": cell_number(
            find_label_value(ws, r"overdue to revised completion")
        ),
        "works_certified": cell_number(
            find_label_value(ws, r"works vetted & certified|works vetted and certified")
        ),
        "works_submitted_not_vetted": cell_number(
            find_label_value(ws, r"works submitted and not yet vetted")
        ),
        "total_works_submitted": cell_number(
            find_label_value(ws, r"total value of works submitted")
        ),
        "gross_certified": cell_number(
            find_label_value(ws, r"total gross value of works completed")
        ),
        "retention_held": cell_number(
            find_label_value(ws, r"retention money - deducted")
        ),
        "retention_released": cell_number(
            find_label_value(ws, r"retention money - released")
        ),
        "advance_recovered": cell_number(
            find_label_value(ws, r"advance payment recovered")
        ),
        "apg_amount": cell_number(
            find_label_value(ws, r"advance payment guarantee")
        ),
        "advance_unrecovered": cell_number(
            find_label_value(ws, r"advance payment yet to be recovered")
        ),
        "apg_expiry": cell_date(find_label_value(ws, r"apg.*expir")),
        "bill1_requested": cell_number(
            find_label_value(ws, r"total bill 1 requested")
        ),
        "bill1_paid": cell_number(
            find_label_value(ws, r"total paid out by pw")
        ),
        "bill1_outstanding": cell_number(
            find_label_value(ws, r"amount outstanding from amounts requested")
        ),
    }
    return {"identity": identity, "snapshot": snapshot, "warnings": warnings}


# ---------------------------------------------------------------------------
# T2.5 — Weekly Summary → long rows (this-week values only)
# ---------------------------------------------------------------------------

def parse_weekly_summary(ws) -> dict[str, Any]:
    warnings: list[str] = []
    rows: list[dict[str, Any]] = []

    # The sheet holds >=2 blocks, each headed by a row containing a
    # "This Week" column: Works Completed and Costs to Date.
    header_hits: list[tuple[int, dict[str, int]]] = []
    seen_rows: set[int] = set()
    for probe_start in range(1, 60):
        hit = find_header_row(
            _window(ws, probe_start), ["this week"], min_matches=1, max_row=1
        )
        if hit is not None:
            actual_row = probe_start
            if actual_row not in seen_rows:
                seen_rows.add(actual_row)
                header_hits.append((actual_row, hit[1]))

    if not header_hits:
        return {"rows": [], "warnings": ["Weekly Summary: no 'This Week' headers found"]}

    for block_idx, (hrow, cols) in enumerate(header_hits):
        section_label = None
        # section name = leftmost text on the header row
        for c in range(1, 6):
            t = cell_text(ws.cell(row=hrow, column=c).value)
            if t:
                section_label = t
                break
        section = section_label or f"BLOCK {block_idx + 1}"

        next_header = (
            header_hits[block_idx + 1][0]
            if block_idx + 1 < len(header_hits)
            else hrow + 40
        )
        this_week_col = cols.get("this week")
        pct_col = next(
            (col for name, col in cols.items() if name.startswith("%")), None
        )

        for r in range(hrow + 1, next_header):
            # item name: first text cell in cols A..C that isn't a bullet
            name = None
            for c in range(1, 4):
                t = cell_text(ws.cell(row=r, column=c).value)
                if t and t not in ("•", "-"):
                    name = t
                    break
            if not name:
                continue
            value = cell_number(ws.cell(row=r, column=this_week_col).value)
            pct = cell_number(ws.cell(row=r, column=pct_col).value) if pct_col else None
            if value is None and pct is None:
                continue  # section separators / prose
            if value is not None:
                rows.append({
                    "section": section, "item": name,
                    "metric": "this_week", "value": value, "raw_value": None,
                })
            if pct is not None:
                rows.append({
                    "section": section, "item": name,
                    "metric": "pct_complete", "value": pct, "raw_value": None,
                })
    return {"rows": rows, "warnings": warnings}


class _window:
    """Tiny worksheet view starting at a given row (for repeated header
    scans without re-walking the whole sheet)."""

    def __init__(self, ws, start_row: int):
        self._ws = ws
        self._start = start_row

    def iter_rows(self, min_row=1, max_row=1, max_col=40):
        yield from self._ws.iter_rows(
            min_row=self._start, max_row=self._start + max_row - 1, max_col=max_col
        )


# ---------------------------------------------------------------------------
# T2.6 — Plant Return
# ---------------------------------------------------------------------------

def parse_plant_return(ws) -> dict[str, Any]:
    """Full roster including idle plants (all-zero rows are the 'on site,
    idle' signal). Footer captured: TOTAL All, % allocated, and the
    'Internal Plant Cost Adjusted for' consumable lines — the deduction
    that becomes the Cost Report's Plant Internal figure."""
    warnings: list[str] = []
    hit = find_header_row(ws, ["fleet no", "description", "hours worked"])
    if hit is None:
        return {"rows": [], "footer": {}, "warnings": ["Plant Return: header row not found"]}
    hrow, cols = hit

    rows = []
    last_excel_row = hrow
    for r in iter_table_rows(ws, hrow, cols):
        fleet = _txt(r, "fleet no")
        if not fleet or _NOISE_ROWS.match(fleet):
            continue
        row = {
            "fleet_number_raw": fleet,
            "description": _txt(r, "description"),
            "plant_category": _txt(r, "plant category"),
            "hours_worked": _num(r, "hours worked") or 0,
            "standby_hours": _num(r, "s/b hours", "standby") or 0,
            "breakdown_hours": _num(r, "b/d hours", "breakdown") or 0,
            "rate_ngn": _num(r, "rate"),
            "plant_cost": _num(r, "plant cost"),
            "transferred_from": _txt(r, "transfered from", "transferred from"),
            "current_location": _txt(r, "current location"),
            "remarks": _txt(r, "remarks"),
        }
        rows.append(row)
        last_excel_row = r["_row"]
        # validation: plant cost = hours worked × rate (standby/BD never charged)
        if row["rate_ngn"] and row["plant_cost"] is not None \
                and abs(row["hours_worked"] * row["rate_ngn"] - row["plant_cost"]) > 1.0:
            warnings.append(
                f"Plant Return {fleet}: hours {row['hours_worked']:g} × rate "
                f"{row['rate_ngn']:g} != cost {row['plant_cost']:,.2f}"
            )

    # ── footer: total / % allocated / consumable adjustments ───────────
    footer: dict[str, Any] = {"adjustments": []}
    in_adjustments = False
    for r in range(last_excel_row + 1, min(last_excel_row + 30, ws.max_row + 1)):
        label = None
        for c in range(1, 8):
            t = cell_text(ws.cell(row=r, column=c).value)
            if t:
                label = t
                break
        if not label:
            continue
        nums = [cell_number(ws.cell(row=r, column=c).value)
                for c in range(2, 14)]
        nums = [n for n in nums if n is not None]
        lu = label.upper()
        if lu.startswith("TOTAL ALL"):
            footer["total_all"] = nums[-1] if nums else None
        elif "% ALLOCATED" in lu:
            footer["pct_allocated"] = nums[-1] if nums else None
        elif "COST ALLOCATED" in lu:
            footer["total_allocated"] = nums[-1] if nums else None
        elif "ADJUSTED FOR" in lu:
            in_adjustments = True
        elif in_adjustments and nums:
            footer["adjustments"].append({"label": label, "amount": nums[-1]})
        elif in_adjustments and not nums:
            # end of adjustment block unless it's a spacer label
            if "NET" in lu or "TOTAL" in lu:
                in_adjustments = False

    if not rows:
        warnings.append("Plant Return: no fleet rows parsed")
    if footer.get("total_all") is not None:
        ours = sum(float(x["plant_cost"] or 0) for x in rows)
        if abs(ours - float(footer["total_all"])) > 1.0:
            warnings.append(
                f"Plant Return: our cost sum {ours:,.2f} != footer TOTAL All "
                f"{float(footer['total_all']):,.2f}"
            )
    return {"rows": rows, "footer": footer, "warnings": warnings}


# ---------------------------------------------------------------------------
# T2.7 — Diesel Consumption
# ---------------------------------------------------------------------------

_COST_CENTRE_RX = re.compile(r"^[A-Z][A-Z /&.-]{2,}$")  # no digits, len>2


def parse_diesel(ws) -> dict[str, Any]:
    """Fuel EVENTS only — rows with fuel taken > 0. Zero rows are unused
    template lines, not idle declarations (that's Plant Return's job).
    Recipients without digits (MECHANICS, CIVIL, LAB/PRECAST) are marked
    cost-centres. The stock header line and the sheet's subtotal/total
    rows are captured for cross-checks."""
    warnings: list[str] = []
    hit = find_header_row(ws, ["fleet no", "description", "total fuel"], min_matches=2)
    if hit is None:
        return {"rows": [], "stock": {}, "sheet_totals": {},
                "warnings": ["Diesel: header row not found"]}
    hrow, cols = hit

    stock = {
        "opening": cell_number(find_label_value(ws, r"^opening stock")),
        "received": cell_number(find_label_value(ws, r"^received this week")),
        "used": cell_number(find_label_value(ws, r"^used this week")),
        "closing": cell_number(find_label_value(ws, r"^closing stock")),
    }

    day_keys = ("saturday", "sunday", "monday", "tuesday", "wednesday",
                "thursday", "friday")
    rows = []
    sheet_totals: dict[str, Any] = {}
    for r in iter_table_rows(ws, hrow, cols):
        fleet = _txt(r, "fleet no")
        if not fleet:
            continue
        total_claimed = _num(r, "total fuel")
        fu = fleet.upper()
        if fu.startswith("SUB TOTAL OTHER"):
            sheet_totals["other_used"] = total_claimed
            continue
        if fu.startswith("TOTAL ALL DIESEL"):
            sheet_totals["all_used"] = total_claimed
            continue
        if _NOISE_ROWS.match(fleet):
            continue

        day_values = {f"{d}_litres": (_num(r, d) or 0) for d in day_keys}
        day_sum = sum(day_values.values())
        litres = total_claimed if total_claimed is not None else day_sum
        if not litres:  # unused template line, not an event
            continue
        if total_claimed is not None and abs(day_sum - total_claimed) > 0.01:
            warnings.append(
                f"Diesel {fleet}: day sum {day_sum:g} != row total {total_claimed:g}"
            )
        rows.append({
            "fleet_number_raw": fleet,
            "description": _txt(r, "description"),
            "plant_category": _txt(r, "plant category"),
            **day_values,
            "amount_ngn": _num(r, "usage value"),
            "is_cost_centre": bool(_COST_CENTRE_RX.fullmatch(fu)),
        })

    # events must sum to the sheet's own total
    total_parsed = sum(sum(row[f"{d}_litres"] for d in day_keys) for row in rows)
    target = sheet_totals.get("all_used", stock.get("used"))
    if target is not None and abs(total_parsed - float(target)) > 0.01:
        warnings.append(
            f"Diesel: parsed events {total_parsed:g}L != sheet total {float(target):g}L"
        )
    return {"rows": rows, "stock": stock, "sheet_totals": sheet_totals,
            "warnings": warnings}


# ---------------------------------------------------------------------------
# T2.8 — Cost Report (this-week amounts only)
# ---------------------------------------------------------------------------

def parse_cost_report(ws) -> dict[str, Any]:
    """Category rows only become data. The sheet's Total row and its
    internal arithmetic (qty × rate, to-date = previous + this-week)
    become cross-checks. Reported-previous is captured per row for
    baseline/gap derivation."""
    warnings: list[str] = []
    hit = find_header_row(
        ws, ["description", "cost category", "amount"], min_matches=2
    )
    if hit is None:
        return {"rows": [], "sheet_total": None,
                "warnings": ["Cost Report: header row not found"]}
    hrow, cols = hit

    def col_for(*fragments: str) -> int | None:
        for name, col in cols.items():
            if all(f in name for f in fragments):
                return col
        return None

    this_week_col = col_for("amount", "this week")
    prev_col = col_for("amount", "previous")
    todate_col = col_for("amount", "up to date") or col_for("amount", "to date")

    rows: list[dict[str, Any]] = []
    sheet_total: dict[str, Any] | None = None
    section = None
    for r in iter_table_rows(ws, hrow, cols, stop_after_blank=15):
        desc = _txt(r, "description")
        category = _txt(r, "cost category")
        excel_row = r["_row"]
        amount_this_week = (cell_number(ws.cell(row=excel_row, column=this_week_col).value)
                            if this_week_col else _num(r, "amount this week"))
        amount_previous = (cell_number(ws.cell(row=excel_row, column=prev_col).value)
                           if prev_col else None)
        amount_to_date = (cell_number(ws.cell(row=excel_row, column=todate_col).value)
                          if todate_col else None)

        # the grand-total row: pure cross-check, never data
        if desc == "Total":
            sheet_total = {"previous": amount_previous,
                           "this_week": amount_this_week,
                           "to_date": amount_to_date}
            continue

        # Section banners ("PLANT DEPARTMENT") sit in the S/No column
        sno = _txt(r, "s/no")
        banner = next(
            (t for t in (sno, desc)
             if t and t.isupper() and len(t) > 3 and not t.replace(" ", "").isdigit()),
            None,
        )
        if banner and category is None and _num(r, "rate") is None:
            section = banner
            continue
        # unused template slots carry literal '0' descriptions
        if not desc or desc == "0":
            continue

        qty = _num(r, "quantity this week")
        rate = _num(r, "rate")
        row = {
            "section": section,
            "description": desc,
            "cost_category": category,
            "unit": _txt(r, "unit"),
            "quantity_this_week": qty,
            "rate_ngn": rate,
            "amount_previous_week": amount_previous,
            "amount_this_week": amount_this_week or 0,
            "amount_to_date": amount_to_date,
        }
        rows.append(row)

        # per-row arithmetic validation
        if qty and rate and amount_this_week is not None \
                and abs(qty * rate - amount_this_week) > 1.0:
            warnings.append(
                f"Cost Report {desc[:40]!r}: qty {qty:g} × rate {rate:g} = "
                f"{qty*rate:,.2f} but sheet says {amount_this_week:,.2f}"
            )
        if amount_previous is not None and amount_to_date is not None \
                and abs((amount_previous + (amount_this_week or 0)) - amount_to_date) > 1.0:
            warnings.append(
                f"Cost Report {desc[:40]!r}: previous + this-week != up-to-date"
            )

    # total cross-check: categorized rows must sum to the sheet's Total
    if sheet_total and sheet_total.get("this_week") is not None:
        ours = sum(float(x["amount_this_week"] or 0)
                   for x in rows if x["cost_category"])
        if abs(ours - float(sheet_total["this_week"])) > 1.0:
            warnings.append(
                f"Cost Report: our categorized sum {ours:,.2f} != sheet total "
                f"{float(sheet_total['this_week']):,.2f}"
            )
    if not rows:
        warnings.append("Cost Report: no rows parsed")
    return {"rows": rows, "sheet_total": sheet_total, "warnings": warnings}


# ---------------------------------------------------------------------------
# T2.9 — Certificates + Payments
# ---------------------------------------------------------------------------

def parse_certificates(ws) -> dict[str, Any]:
    warnings: list[str] = []
    hit = find_header_row(
        ws, ["cert number", "gross value of works don"], min_matches=1
    )
    if hit is None:
        return {"rows": [], "warnings": ["Certificates: header row not found"]}
    hrow, cols = hit

    rows = []
    prev_gross = 0.0
    for r in iter_table_rows(ws, hrow, cols):
        cert_no = _txt(r, "cert number")
        gross = _num(r, "gross value of works don")
        if cert_no is None or gross is None:
            continue
        row = {
            "cert_number": cert_no,
            "date_submitted": _dt(r, "date submitted"),
            "gross_value_works_done": gross,
            "add_materials_on_site": _num(r, "add materials on site"),
            "less_materials_on_site": _num(r, "less materials on site"),
            "general_bill_1": _num(r, "general bill 1"),
            "total_value_of_work_done": _num(r, "total value of work done"),
            "value_of_works_per_cert": _num(r, "value of works per cert"),
            "total_retention_held": _num(r, "total retention held"),
            "total_net_payment": _num(r, "total net payment"),
            # commercial columns L..Q (locked 2026-07-11)
            "retention_released": _num(r, "add release of retention"),
            "contingency_used": _num(r, "add value of contingency used"),
            "contingency_deducted": _num(r, "less value of contingency"),
            "fluctuation_materials": _num(r, "add fluctuation on materials"),
            "advance_received": _num(r, "add advance received"),
            "total_works_executed": _num(r, "total value of works executed"),
            "advance_recovery": _num(r, "deduct advance recovery"),
        }
        rows.append(row)

        # gross is CUMULATIVE: retention must be 5% of it, increments >= 0
        ret = row["total_retention_held"]
        if ret is not None and gross and abs(ret - gross * 0.05) > 1.0:
            warnings.append(
                f"Certificates cert {cert_no}: retention {ret:,.2f} != 5% of "
                f"cumulative gross ({gross * 0.05:,.2f})"
            )
        if gross < prev_gross - 1.0:
            warnings.append(
                f"Certificates cert {cert_no}: cumulative gross DECREASED "
                f"({prev_gross:,.2f} -> {gross:,.2f})"
            )
        elif abs(gross - prev_gross) <= 1.0 and prev_gross > 0:
            warnings.append(
                f"Certificates cert {cert_no}: zero increment (same cumulative "
                f"gross as previous cert — resubmission?)"
            )
        prev_gross = gross
    # an empty ledger is legitimate on a young project — no warning
    return {"rows": rows, "warnings": warnings}


def parse_payments(ws) -> dict[str, Any]:
    warnings: list[str] = []
    hit = find_header_row(
        ws, ["voucher number", "payment type", "gross amount"], min_matches=2
    )
    if hit is None:
        return {"rows": [], "warnings": ["Payments: header row not found"]}
    hrow, cols = hit

    rows = []
    for r in iter_table_rows(ws, hrow, cols):
        voucher = _txt(r, "voucher number")
        gross = _num(r, "gross amount")
        if voucher is None and gross is None:
            continue
        row = {
            "payment_date": _dt(r, "date"),
            "voucher_number": voucher,
            "payment_type": _txt(r, "payment type"),
            "gross_amount": gross,
            "wht": _num(r, "wht") or 0,
            "vat": _num(r, "vat") or 0,
            "vetting_fee": _num(r, "vetting fee") or 0,
            "stamp_duty": _num(r, "stamp duty") or 0,
            "other_deductions": _num(r, "other") or 0,
            "net_amount": _num(r, "net amount"),
        }
        if not (row["gross_amount"] or 0) and not (row["net_amount"] or 0):
            continue  # template placeholder (young project, no money moved)
        if row["gross_amount"] is not None and row["net_amount"] is not None:
            expected_net = row["gross_amount"] - (
                row["wht"] + row["vat"] + row["vetting_fee"]
                + row["stamp_duty"] + row["other_deductions"]
            )
            if abs(expected_net - row["net_amount"]) > 1.0:
                warnings.append(
                    f"Payments {voucher or '?'}: gross-deductions "
                    f"{expected_net:,.2f} != net {row['net_amount']:,.2f}"
                )
        rows.append(row)
    if not rows:
        warnings.append("Payments: no rows parsed")
    return {"rows": rows, "warnings": warnings}


# ---------------------------------------------------------------------------
# T2.10 — BEME (bills + items + this-week progress)
# ---------------------------------------------------------------------------

_DOTTED_CODE_RX = re.compile(r"^\d+(\.\d+)*$")           # bill: any depth
_ITEM_CODE_RX = re.compile(r"^\d+(\.\d+)+[a-z]?$")        # item: >= 2 segments
_BILL_TOTAL_RX = re.compile(r"^total bill", re.IGNORECASE)  # number optional


def _extends_by_one(bill_code: str, code: str) -> bool:
    """'4.05a' extends '4'; '5.3.12' extends '5.3'; '5.2.1' does NOT
    extend '5' — an item sits exactly one segment below its bill."""
    base = code[:-1] if code and code[-1].isalpha() else code
    if not base.startswith(bill_code + "."):
        return False
    rest = base[len(bill_code) + 1:]
    return bool(rest) and "." not in rest


def _normalize_item_code(raw: str | None) -> str | None:
    """'6.23999999999999' (Excel float corruption) → '6.24'; '1.1' stays."""
    if raw is None:
        return None
    code = raw.strip()
    if re.fullmatch(r"\d+\.\d{4,}", code):
        try:
            return f"{round(float(code), 2):g}"
        except ValueError:
            return code
    return code


def parse_beme(ws) -> dict[str, Any]:
    """Classify every row by the COMPANY standard (verified on Akwa Ibom
    and Kaduna Bridge): a bill is a dotted code at ANY depth with an
    ALL-CAPS name and no amounts; items sit exactly one code segment
    deeper. Only real items become data; the sheet's own totals become
    cross-checks. Emits reported-previous per item (baseline/gap inputs).
    """
    warnings: list[str] = []
    hit = find_header_row(
        ws, ["item", "description", "rate", "this week qty"], min_matches=3
    )
    if hit is None:
        return {"rows": [], "bills": [], "tail": {}, "summary_table": [],
                "warnings": ["BEME: header row not found"]}
    hrow, cols = hit

    rows: list[dict[str, Any]] = []
    bills: list[dict[str, Any]] = []
    tail: dict[str, Any] = {}
    summary_table: list[dict[str, Any]] = []
    current_bill: dict[str, Any] | None = None
    in_summary_table = False
    unclassified_priced = 0

    for r in iter_table_rows(ws, hrow, cols, stop_after_blank=25, max_rows=2000):
        desc = _txt(r, "description")
        code = _normalize_item_code(_txt(r, "item"))
        contract_amount = _num(r, "contract amount")
        this_week_amount = _num(r, "this week amount")
        rate = _num(r, "rate")

        code_s = code or ""
        desc_s = desc or ""
        desc_u = desc_s.upper()

        # ── repeated page headers ("ITEM | DESCRIPTION") ────────────────
        if norm(code_s) == "item":
            continue

        # ── bill total rows: "Total Bill No… Carried to Summary" ───────
        if _BILL_TOTAL_RX.search(code_s) or _BILL_TOTAL_RX.search(desc_s):
            if current_bill is not None:
                current_bill["sheet_total_contract"] = contract_amount
                current_bill["sheet_total_this_week"] = this_week_amount
                current_bill["sheet_total_previous"] = _num(r, "previous amount")
            continue

        # ── tail rows (after all bills) ─────────────────────────────────
        if desc_u.startswith("ADD CONTINGENC"):
            tail["contingency"] = {"contract": contract_amount,
                                   "this_week": this_week_amount}
            in_summary_table = True  # everything below is summary/markup
            continue
        if desc_u.startswith("ADD VOP"):
            tail["vop"] = {"contract": contract_amount,
                           "this_week": this_week_amount}
            continue
        if desc_u.startswith("ADD VAT"):
            tail["vat"] = {"contract": contract_amount,
                           "this_week": this_week_amount}
            continue
        if desc_u.startswith(("SUB – TOTAL", "SUB - TOTAL", "SUB-TOTAL", "SUB –TOTAL")):
            tail.setdefault("subtotals", []).append(
                {"contract": contract_amount, "this_week": this_week_amount})
            continue
        if desc_u == "TOTAL" or "GRAND TOTAL" in desc_u:
            tail["grand_total"] = {"contract": contract_amount,
                                   "this_week": this_week_amount}
            continue
        if code_s.startswith("Total Contingency"):
            tail["contingency_vop_total"] = {"contract": contract_amount,
                                             "this_week": this_week_amount}
            continue

        # ── bill headers: dotted code at ANY depth + CAPS + no money ───
        if (_DOTTED_CODE_RX.fullmatch(code_s) and desc_s and desc_s == desc_u
                and rate is None and contract_amount is None
                and this_week_amount is None and not in_summary_table):
            current_bill = {
                "bill_code": code_s,
                "bill_no": int(code_s) if code_s.isdigit() else None,
                "sort_order": len(bills) + 1,
                "name": desc_s,
                "sheet_total_contract": None,
                "sheet_total_this_week": None,
                "sheet_total_previous": None,
            }
            bills.append(current_bill)
            continue

        # ── summary-table restatement (after the tail begins) ──────────
        # covers coded bill lines AND its code-less rows (its own
        # "Contingency & VOP" line) — never data, kept for the preview
        if in_summary_table and desc_s and (
                _DOTTED_CODE_RX.fullmatch(code_s) or not code_s):
            summary_table.append({"bill_code": code_s or None, "name": desc_s,
                                  "contract": contract_amount,
                                  "this_week": this_week_amount})
            continue

        # ── real items: one segment below their bill (by CODE, not by
        # position — sites sometimes append an item inside another
        # bill's block, e.g. Akwa Ibom's 7.09 sitting in Bill 8) ────────
        owner = None
        if (_ITEM_CODE_RX.fullmatch(code_s) and desc_s
                and not in_summary_table and current_bill is not None):
            if _extends_by_one(current_bill["bill_code"], code_s):
                owner = current_bill
            else:
                owner = next((b for b in reversed(bills)
                              if _extends_by_one(b["bill_code"], code_s)), None)
                if owner is not None:
                    warnings.append(
                        f"BEME item {code_s} appears inside Bill "
                        f"{current_bill['bill_code']}'s block — assigned to "
                        f"Bill {owner['bill_code']} by its code"
                    )
        if owner is not None:
            rows.append({
                "bill_code": owner["bill_code"],
                "bill_no": owner["bill_no"],
                "item_code": code_s,
                "description": desc_s,
                "unit": _txt(r, "unit"),
                "contract_qty": _num(r, "contract qty"),
                "rate": rate,
                "contract_amount": contract_amount,
                "qty_previous_reported": _num(r, "previous qty"),
                "amount_previous_reported": _num(r, "previous amount"),
                "qty_this_week": _num(r, "this week qty"),
                "amount_this_week": this_week_amount,
            })
            continue

        # priced rows we could not classify must never vanish silently
        if contract_amount or this_week_amount:
            unclassified_priced += 1

    # ── duplicate identities: the site reused 3.07 for two different
    # rows — dup_seq (occurrence index) keeps both as distinct items ────
    seen_ids: dict[tuple, int] = {}
    for item in rows:
        key = (item["bill_code"], item["item_code"], item["description"])
        item["dup_seq"] = seen_ids.get(key, 0)
        seen_ids[key] = item["dup_seq"] + 1

    # ── arithmetic cross-checks (also power the upload preview) ────────
    checks: list[dict[str, Any]] = []
    for b in bills:
        mine_c = sum(float(i["contract_amount"] or 0)
                     for i in rows if i["bill_code"] == b["bill_code"])
        mine_w = sum(float(i["amount_this_week"] or 0)
                     for i in rows if i["bill_code"] == b["bill_code"])
        mine_p = sum(float(i["amount_previous_reported"] or 0)
                     for i in rows if i["bill_code"] == b["bill_code"])
        for label, mine, sheet in (("contract", mine_c, b["sheet_total_contract"]),
                                   ("this_week", mine_w, b["sheet_total_this_week"]),
                                   ("previous", mine_p, b["sheet_total_previous"])):
            if sheet is not None and abs(mine - float(sheet)) > 1.0:
                checks.append({
                    "check": f"bill_{b['bill_code']}_{label}",
                    "ours": round(mine, 2), "sheet": float(sheet),
                    "delta": round(mine - float(sheet), 2),
                })
                warnings.append(
                    f"BEME Bill {b['bill_code']} {label}: our sum "
                    f"{mine:,.2f} != sheet total {float(sheet):,.2f} — the "
                    f"sheet's own SUM range may be broken"
                )
    if unclassified_priced:
        warnings.append(
            f"BEME: {unclassified_priced} priced row(s) could not be "
            f"classified — layout may deviate from the company standard"
        )
    if not rows:
        warnings.append("BEME: no item rows parsed")
    if not bills:
        warnings.append("BEME: no bill headers found")
    return {"rows": rows, "bills": bills, "tail": tail,
            "summary_table": summary_table, "cross_checks": checks,
            "warnings": warnings}


# ---------------------------------------------------------------------------
# Promotions (dossiers 9-13, locked 2026-07-13)
# ---------------------------------------------------------------------------

def parse_hired_vehicles(ws) -> dict[str, Any]:
    """Content rows only — INCLUDING zero-day rows that carry a rate
    (standing hire arrangements, e.g. the ₦1M/day crane). Total row is a
    cross-check against the Cost Report's Hired Plant figure."""
    warnings: list[str] = []
    hit = find_header_row(ws, ["description", "owners", "days worked", "rate"],
                          min_matches=2)
    if hit is None:
        return {"rows": [], "sheet_total": None,
                "warnings": ["Hired Vehicles: header row not found"]}
    hrow, cols = hit

    rows: list[dict[str, Any]] = []
    sheet_total = None
    for r in iter_table_rows(ws, hrow, cols):
        desc = _txt(r, "description")
        owner = _txt(r, "owners")
        rate = _num(r, "rate")
        amount = _num(r, "amount")
        if owner and norm(owner) == "total":
            sheet_total = amount
            continue
        if not (desc or owner or rate):
            continue  # numbered template line
        days = _num(r, "days worked") or 0
        row = {
            "registration_no": _txt(r, "reg"),
            "description": desc,
            "section": _txt(r, "section"),
            "owners": owner,
            "days_worked": days,
            "rate_ngn": rate,
            "amount_ngn": amount if amount is not None else 0,
            "remarks": _txt(r, "remarks"),
        }
        rows.append(row)
        if rate and abs(days * rate - (amount or 0)) > 1.0:
            warnings.append(
                f"Hired Vehicles {desc or owner}: days {days:g} × rate "
                f"{rate:g} != amount {(amount or 0):,.2f}"
            )
    if sheet_total is not None:
        ours = sum(float(x["amount_ngn"] or 0) for x in rows)
        if abs(ours - float(sheet_total)) > 1.0:
            warnings.append(
                f"Hired Vehicles: our sum {ours:,.2f} != sheet total "
                f"{float(sheet_total):,.2f}"
            )
    return {"rows": rows, "sheet_total": sheet_total, "warnings": warnings}


def parse_labour(ws) -> dict[str, Any]:
    """Two blocks (permanent / CASUAL STAFF), departments keyed by their
    fixed template slot. Totals + movement are validators, never data."""
    warnings: list[str] = []
    hit = find_header_row(ws, ["department", "manning this week"], min_matches=2)
    if hit is None:
        return {"rows": [], "totals": {},
                "warnings": ["Labour Strength: header row not found"]}
    hrow, cols = hit

    rows: list[dict[str, Any]] = []
    totals: dict[str, Any] = {}
    block = "permanent"
    for r in iter_table_rows(ws, hrow, cols, stop_after_blank=10):
        slot = _txt(r, "s/no")
        dept = _txt(r, "department")
        this_w = _num(r, "manning this week")
        prev_w = _num(r, "manning previous")
        move = _num(r, "movement")
        if slot and norm(slot) == "casual staff":
            block = "casual"
            continue
        if dept and norm(dept).startswith("total"):
            totals[block] = this_w
            continue
        if slot and norm(slot) == "s/no":
            continue  # repeated casual-block header
        if not dept:
            continue
        rows.append({
            "block": block,
            "dept_slot": int(float(slot)) if slot and slot.replace(".", "").isdigit() else None,
            "department": dept,
            "manning_this_week": this_w if this_w is not None else 0,
            "manning_previous_week": prev_w,
            "movement": move,
            "comment": _txt(r, "comment"),
        })
        if (this_w is not None and prev_w is not None and move is not None
                and abs((this_w - prev_w) - move) > 0.01):
            warnings.append(
                f"Labour {dept}: movement {move:g} != this-week {this_w:g} "
                f"- previous {prev_w:g}"
            )
    for blk, sheet_total in totals.items():
        if sheet_total is None:
            continue
        ours = sum(x["manning_this_week"] for x in rows if x["block"] == blk)
        if abs(ours - float(sheet_total)) > 0.01:
            warnings.append(
                f"Labour {blk}: our head count {ours:g} != sheet total "
                f"{float(sheet_total):g}"
            )
    if not rows:
        warnings.append("Labour Strength: no department rows parsed")
    return {"rows": rows, "totals": totals, "warnings": warnings}


def parse_subcontractors(ws) -> dict[str, Any]:
    """Per-report ledger (read latest, like payments). Names carry forward
    from their group row; work items = rows with a description + rate.
    Zero-quantity rate-card rows are kept — standing agreements."""
    warnings: list[str] = []
    hit = find_header_row(ws, ["subcontractor", "description", "agreed rate"],
                          min_matches=2)
    if hit is None:
        return {"rows": [], "warnings": ["Subcontractors: header row not found"]}
    hrow, cols = hit

    rows: list[dict[str, Any]] = []
    current_sub: str | None = None
    for r in iter_table_rows(ws, hrow, cols, stop_after_blank=12):
        sub = _txt(r, "subcontractor")
        desc = _txt(r, "description")
        rate = _num(r, "agreed rate")
        if sub:
            current_sub = sub
        if not desc or rate is None:
            continue
        prev_q = _num(r, "previous qty")
        week_q = _num(r, "qty executed for week")
        total_q = _num(r, "total qty executed")
        row = {
            "subcontractor_name": current_sub,
            "description": desc,
            "location": _txt(r, "location"),
            "unit": _txt(r, "unit"),
            "agreed_rate": rate,
            "assigned_qty": _num(r, "assigned qty"),
            "previous_qty": prev_q,
            "qty_this_week": week_q,
            "qty_to_date": total_q,
            "amount_this_week": _num(r, "value completed this week"),
            "value_previous": _num(r, "previous value completed"),
            "amount_to_date": _num(r, "total value completed"),
            "balance_remaining": _num(r, "balance to complete"),
            "value_to_completion": _num(r, "value to completion"),
            "remarks": _txt(r, "remarks"),
        }
        rows.append(row)
        if (prev_q is not None and total_q is not None
                and abs((prev_q + (week_q or 0)) - total_q) > 0.01):
            warnings.append(
                f"Subcontractors {desc[:36]!r}: previous + this-week != total qty"
            )
    if not rows:
        warnings.append("Subcontractors: no work items parsed")
    return {"rows": rows, "warnings": warnings}


def parse_materials(ws) -> dict[str, Any]:
    """Stock cycle + usage split + the sheet's own loss detector.
    stock_maintained tells the truth about whether the site keeps the
    ledger (Kaduna: yes, zero discrepancies; Akwa: usage only)."""
    warnings: list[str] = []
    hit = find_header_row(ws, ["description", "opening stock", "received"],
                          min_matches=2)
    if hit is None:
        return {"rows": [], "sheet_total": None, "stock_maintained": None,
                "warnings": ["Materials & Civils: header row not found"]}
    hrow, cols = hit

    rows: list[dict[str, Any]] = []
    sheet_total = None
    section = "materials"
    for r in iter_table_rows(ws, hrow, cols, stop_after_blank=10):
        slot = _txt(r, "#") or _txt(r, "s/no")
        desc = _txt(r, "description")
        if slot and slot.isupper() and not slot.replace(" ", "").isdigit()                 and len(slot) > 3:
            # "QUARRY MATERIALS" banner rides the # column → canonical value
            section = "quarry" if "QUARRY" in slot else "materials"
            continue
        if desc and norm(desc).startswith("total all"):
            sheet_total = _num(r, "total used") or _num(r, "used")
            continue
        if not desc:
            continue
        rows.append({
            "sheet_source": section,
            "material_name": " ".join(desc.split()),  # strip nbsp padding
            "unit": _txt(r, "unit"),
            "unit_cost": _num(r, "current price"),
            "opening_stock": _num(r, "opening stock"),
            "received": _num(r, "received"),
            "closing_stock": _num(r, "closing stock"),
            "available_for_use": _num(r, "available for use"),
            "used_works": _num(r, "on site works"),
            "used_precast": _num(r, "precast"),
            "used_mobilisation": _num(r, "mobil"),
            "used": _num(r, "total used"),
            "discrepancy_qty": None,   # computed below
            "discrepancy_value": None,
            "remarks": _txt(r, "remarks"),
        })

    # the site maintains stock only if closings are actually entered
    stock_maintained = any(x["closing_stock"] not in (None, 0) for x in rows)
    for x in rows:
        o, rcv, c = x["opening_stock"], x["received"], x["closing_stock"]
        used = x["used"]
        if stock_maintained and o is not None and c is not None:
            avail = (o or 0) + (rcv or 0) - (c or 0)
            x["discrepancy_qty"] = round(avail - (used or 0), 3)
            if x["unit_cost"]:
                x["discrepancy_value"] = round(
                    x["discrepancy_qty"] * x["unit_cost"], 2)
            if abs(x["discrepancy_qty"]) > 0.01:
                warnings.append(
                    f"Materials {x['material_name'][:30]}: stock discrepancy "
                    f"{x['discrepancy_qty']:g} {x['unit'] or ''}"
                )
        x["stock_maintained"] = stock_maintained
    if not stock_maintained and rows:
        warnings.append(
            "Materials: stock side not maintained (openings/closings empty) "
            "— usage recorded, loss detection unavailable"
        )
    return {"rows": rows, "sheet_total": sheet_total,
            "stock_maintained": stock_maintained, "warnings": warnings}


# ---------------------------------------------------------------------------
# Lists (master data — company calendar + reference lists)
# ---------------------------------------------------------------------------

def parse_lists(ws) -> dict[str, Any]:
    """Cols A–C: the company's daily calendar (date, week no). Remaining
    headed columns: reference lists (UOMs, categories, rates…)."""
    warnings: list[str] = []

    # calendar: build {(year, week_number): max date seen} = week-ending map
    week_endings: dict[tuple[int, int], Any] = {}
    for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
        d = cell_date(row[0])
        wk = cell_number(row[1])
        if d is None or wk is None:
            continue
        key = (d.year, int(wk))
        if key not in week_endings or d > week_endings[key]:
            week_endings[key] = d

    # reference lists from headed columns beyond C
    reference: list[dict[str, Any]] = []
    headers = {}
    for cell in next(ws.iter_rows(min_row=1, max_row=1, max_col=45)):
        if cell.column > 3 and cell.value is not None:
            t = cell_text(cell.value)
            if t:
                headers[cell.column] = t
    for col, list_name in headers.items():
        order = 0
        for r in range(2, 200):
            item = cell_text(ws.cell(row=r, column=col).value)
            if item is None:
                continue
            order += 1
            detail = cell_text(ws.cell(row=r, column=col + 1).value)
            reference.append({
                "list_name": list_name, "item": item,
                "detail": detail, "sort_order": order,
            })
    if not week_endings:
        warnings.append("Lists: calendar columns not parsed")
    return {
        "week_endings": week_endings,
        "reference": reference,
        "warnings": warnings,
    }


# ---------------------------------------------------------------------------
# Orchestrator
# ---------------------------------------------------------------------------

_SHEET_PARSERS = {
    "Contract Summary": parse_contract_summary,
    "Weekly Summary": parse_weekly_summary,        # cross-check only
    "Plant Return": parse_plant_return,
    "Diesel Consumption": parse_diesel,
    "Cost Report": parse_cost_report,
    "Certificate Status": parse_certificates,
    "Payments Recieved": parse_payments,
    "BEME & Works Completed Fd": parse_beme,
    "Hired Vehicles": parse_hired_vehicles,
    "Labour Strength": parse_labour,
    "Subcontractors": parse_subcontractors,
    "Materials & Civils": parse_materials,
    "Lists": parse_lists,
}

# Sheets we deliberately do NOT parse — stored in Storage, shown raw in
# the upload preview. Their money already auto-posts into the Cost Report.
# Dormant sheets: stored in Storage, raw-previewed, watched for life.
STORED_ONLY_SHEETS = (
    "Bill 1 Summary", "Bill 1 Payments", "Precast",
)



def _extract_declaration(wb: Workbook) -> dict[str, Any]:
    """The workbook declares its own week: 11+ sheets carry 'Week No:'
    and 'Date:' headers. Consensus across sheets beats any typed form;
    disagreement between sheets is a hard integrity problem (a workbook
    assembled from different weeks' sheets)."""
    from collections import Counter
    from datetime import date as _date, datetime as _dt

    votes: list[tuple[str, int, Any]] = []  # (sheet, week, date)
    for name in wb.sheetnames:
        ws = wb[name]
        wk = None
        dt = None
        for r in range(1, 5):
            for c in range(1, 22):
                v = ws.cell(r, c).value
                if not isinstance(v, str):
                    continue
                label = v.strip().rstrip(":").lower()
                if label == "week no":
                    nxt = ws.cell(r, c + 1).value
                    if isinstance(nxt, (int, float)) and 1 <= int(nxt) <= 53:
                        wk = int(nxt)
                elif label == "date":
                    nxt = ws.cell(r, c + 1).value
                    if isinstance(nxt, (_dt, _date)):
                        dt = nxt.date() if isinstance(nxt, _dt) else nxt
        if wk is not None and dt is not None:
            votes.append((name, wk, dt))

    if not votes:
        return {"week_number": None, "year": None, "week_ending_date": None,
                "consistent": False, "votes": 0, "disagreements": []}

    week_counts = Counter(w for _, w, _ in votes)
    date_counts = Counter(d for _, _, d in votes)
    week, _ = week_counts.most_common(1)[0]
    wdate, _ = date_counts.most_common(1)[0]
    disagreements = [
        {"sheet": n, "week": w, "date": d.isoformat()}
        for n, w, d in votes if w != week or d != wdate
    ]
    # year of the report-week (Dec/Jan boundary guard)
    year = wdate.year
    if week >= 50 and wdate.month == 1:
        year -= 1
    elif week == 1 and wdate.month == 12:
        year += 1
    return {
        "week_number": week,
        "year": year,
        "week_ending_date": wdate,
        "consistent": not disagreements,
        "votes": len(votes),
        "disagreements": disagreements,
    }


def parse_workbook(wb: Workbook) -> dict[str, Any]:
    """Parse all 16 sheets. One failing sheet never sinks the rest:
    each gets status ok | partial (parsed with warnings) | failed."""
    drift = check_workbook(wb)
    sheets: dict[str, dict[str, Any]] = {}

    for spec in WORKBOOK_MANIFEST:
        canonical = spec.canonical
        parser = _SHEET_PARSERS.get(canonical)
        if parser is None:
            continue
        ws = resolve_sheet(wb, spec)
        if ws is None:
            sheets[canonical] = {
                "status": "failed", "error": "sheet missing", "warnings": [],
            }
            continue
        try:
            result = parser(ws)
            warnings = result.get("warnings", [])
            payload = {k: v for k, v in result.items() if k != "warnings"}
            n_rows = len(result.get("rows", [])) if "rows" in result else None
            sheets[canonical] = {
                "status": "partial" if warnings else "ok",
                "warnings": warnings,
                "rows_parsed": n_rows,
                **payload,
            }
        except Exception as exc:  # one sheet must never sink the workbook
            sheets[canonical] = {
                "status": "failed",
                "error": f"{type(exc).__name__}: {exc}",
                "warnings": [],
            }

    return {
        "sheets": sheets,
        "drift": {
            "clean": drift.clean,
            "missing": drift.missing,
            "drifted": drift.drifted,
        },
        "identity": sheets.get("Contract Summary", {}).get("identity"),
        "declared": _extract_declaration(wb),
    }
