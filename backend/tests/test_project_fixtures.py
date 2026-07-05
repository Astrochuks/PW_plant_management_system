"""T0.3 — Projects module test fixtures sanity checks.

Verifies the fixture workbooks exist, open cleanly, and contain the sheets
every downstream parser test depends on. If these fail, nothing in the
projects parser suite can be trusted.
"""

from pathlib import Path

import openpyxl
import pytest

FIXTURES_DIR = Path(__file__).parent / "fixtures" / "projects"

AWARD_LETTERS = FIXTURES_DIR / "award_letters_2017.xlsx"
WEEKLY_REPORTS = [
    FIXTURES_DIR / "week_02_akwa_ibom_2026.xlsx",
    FIXTURES_DIR / "week_10_akwa_ibom_2026.xlsx",
]

# The 16 sheets the weekly-report parser must handle (vendor spelling kept).
EXPECTED_WEEKLY_SHEETS = {
    "Weekly Summary",
    "Contract Summary",
    "BEME & Works Completed Fd",
    "Certificate Status",
    "Payments Recieved",
    "Cost Report",
    "Diesel Consumption",
    "Plant Return",
    "Hired Vehicles",
    "Labour Strength",
    "Subcontractors",
    "Precast",
    "Materials & Civils",
    "Bill 1 Summary",
    "Bill 1 Payments",
    "Lists",
}

# 17 client/state sheets in the 2017 register workbook.
EXPECTED_REGISTER_SHEET_COUNT = 17


def _load(path: Path) -> openpyxl.Workbook:
    assert path.exists(), f"fixture missing: {path}"
    try:
        return openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # pragma: no cover - failure is the diagnostic
        pytest.fail(f"fixture {path.name} failed to open: {exc}")


def test_award_letters_fixture_opens_with_expected_sheets():
    wb = _load(AWARD_LETTERS)
    try:
        assert len(wb.sheetnames) == EXPECTED_REGISTER_SHEET_COUNT, wb.sheetnames
        # Spot-check known sheets survive the copy intact
        for sheet in ("PLATEAU", "TARABA", "Akwa Ibom", "PRIVATE CLIENTS"):
            assert sheet in wb.sheetnames
    finally:
        wb.close()


@pytest.mark.parametrize("path", WEEKLY_REPORTS, ids=lambda p: p.stem)
def test_weekly_report_fixtures_contain_all_16_sheets(path: Path):
    wb = _load(path)
    try:
        missing = EXPECTED_WEEKLY_SHEETS - set(wb.sheetnames)
        assert not missing, f"{path.name} missing sheets: {sorted(missing)}"
    finally:
        wb.close()


@pytest.mark.parametrize("path", WEEKLY_REPORTS, ids=lambda p: p.stem)
def test_weekly_report_fixtures_have_readable_data(path: Path):
    """The Plant Return sheet must yield real fleet rows — guards against
    corrupted copies that open but read empty."""
    wb = _load(path)
    try:
        ws = wb["Plant Return"]
        rows_with_fleet = sum(
            1
            for row in ws.iter_rows(min_row=4, max_col=1, values_only=True)
            if row[0] not in (None, "")
        )
        assert rows_with_fleet > 50, (
            f"{path.name}: only {rows_with_fleet} fleet rows in Plant Return"
        )
    finally:
        wb.close()
