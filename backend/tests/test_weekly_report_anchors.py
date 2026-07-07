"""T2.2/T2.3 — anchor toolkit + manifest, against REAL workbooks and
synthetically shifted copies (the whole point: survive drift)."""

from pathlib import Path

import openpyxl
import pytest

from app.services.weekly_report_parsing import (
    WORKBOOK_MANIFEST,
    cell_date,
    cell_number,
    cell_text,
    check_workbook,
    find_header_row,
    find_label_value,
    iter_table_rows,
    norm,
    resolve_sheet,
)

FIXTURES = Path(__file__).parent / "fixtures" / "projects"


@pytest.fixture(scope="module")
def wb10():
    return openpyxl.load_workbook(
        FIXTURES / "week_10_akwa_ibom_2026.xlsx", data_only=True
    )


@pytest.fixture(scope="module")
def wb2():
    return openpyxl.load_workbook(
        FIXTURES / "week_02_akwa_ibom_2026.xlsx", data_only=True
    )


def _fresh_wb10():
    """Independent copy to mutate (re-SAVING the workbook trips on an
    embedded WMF image openpyxl can't round-trip — so clone via reload)."""
    return openpyxl.load_workbook(
        FIXTURES / "week_10_akwa_ibom_2026.xlsx", data_only=True
    )


@pytest.fixture(scope="module")
def shifted_wb():
    """Week 10's Plant Return with 2 rows and 1 column inserted — the
    drift that breaks fixed-address parsers."""
    wb = _fresh_wb10()
    ws = wb["Plant Return"]
    ws.insert_rows(2, amount=2)
    ws.insert_cols(1, amount=1)
    return wb


class TestCellCoercion:
    def test_numbers(self):
        assert cell_number("2,000,982.7") == 2000982.7
        assert cell_number(15) == 15.0
        assert cell_number("#VALUE!") is None
        assert cell_number("#REF!") is None
        assert cell_number("abc") is None
        assert cell_number(None) is None
        assert cell_number(True) is None

    def test_norm(self):
        assert norm(" Fleet  No. ") == "fleet no"
        assert norm("PAYMENTS RECIEVED") == "payments recieved"
        assert norm(None) == ""


class TestAnchorsOnRealSheets:
    def test_contract_summary_labels(self, wb10):
        ws = wb10["Contract Summary"]
        name = find_label_value(ws, r"name of contract")
        assert "CONSTRUCTION" in str(name).upper()

        short = find_label_value(ws, r"short name")
        assert "AKWA IBOM" in str(short).upper()

        amount = cell_number(find_label_value(ws, r"original contract amount"))
        assert amount == pytest.approx(10621359979.09)

    def test_plant_return_header(self, wb10):
        ws = wb10["Plant Return"]
        hit = find_header_row(ws, ["fleet no", "description", "hours worked"])
        assert hit is not None
        row, cols = hit
        assert row == 3
        assert "fleet no" in cols and "hours worked" in cols

    def test_plant_return_rows(self, wb10):
        ws = wb10["Plant Return"]
        row, cols = find_header_row(ws, ["fleet no", "description", "hours worked"])
        rows = list(iter_table_rows(ws, row, cols))
        fleet_nums = [cell_text(r["fleet no"]) for r in rows if cell_text(r["fleet no"])]
        assert len(fleet_nums) > 100
        assert "AC163" in fleet_nums

    def test_diesel_header_week2(self, wb2):
        ws = wb2["Diesel Consumption"]
        hit = find_header_row(ws, ["fleet no", "description", "total fuel"], min_matches=2)
        assert hit is not None


class TestDriftResilience:
    def test_shifted_sheet_still_parses(self, shifted_wb):
        """2 inserted rows + 1 inserted column — anchors must not care."""
        ws = shifted_wb["Plant Return"]
        hit = find_header_row(ws, ["fleet no", "description", "hours worked"])
        assert hit is not None
        row, cols = hit
        assert row == 5  # shifted by the 2 inserted rows

        rows = list(iter_table_rows(ws, row, cols))
        fleet_nums = [cell_text(r["fleet no"]) for r in rows if cell_text(r["fleet no"])]
        assert "AC163" in fleet_nums  # same data, new coordinates


class TestManifest:
    def test_real_workbooks_are_clean(self, wb10, wb2):
        for wb in (wb10, wb2):
            report = check_workbook(wb)
            assert report.clean, {
                "missing": report.missing, "drifted": report.drifted
            }
            assert len(report.ok) == len(WORKBOOK_MANIFEST)

    def test_missing_sheet_detected(self, wb10):
        wb = _fresh_wb10()
        del wb["Diesel Consumption"]
        report = check_workbook(wb)
        assert "Diesel Consumption" in report.missing
        assert not report.clean

    def test_renamed_headers_detected_as_drift(self, wb10):
        wb = _fresh_wb10()
        ws = wb["Plant Return"]
        # vandalize the header row
        row, cols = find_header_row(ws, ["fleet no", "description", "hours worked"])
        for col in cols.values():
            ws.cell(row=row, column=col).value = "XXX"
        report = check_workbook(wb)
        assert "Plant Return" in report.drifted
        assert "header row" in report.drifted["Plant Return"]

    def test_vendor_spelling_alias(self, wb10):
        """'Payments Recieved' (sic) and a corrected 'Payments Received'
        must both resolve."""
        spec = next(s for s in WORKBOOK_MANIFEST if s.canonical == "Payments Recieved")
        assert resolve_sheet(wb10, spec) is not None

        wb = _fresh_wb10()
        wb["Payments Recieved"].title = "Payments Received"
        assert resolve_sheet(wb, spec) is not None
